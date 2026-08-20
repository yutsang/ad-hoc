#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
勘查工具 —— 這個專案唯一的一支, 有新問題就改這支, 不要再開新檔。

只做「還沒答案」的檢查。已經問完的 (欄位結構分組、腳本 AST 比對、
全掃 FX 命中率) 都拿掉了, 不要再花那 37 分鐘。

目前在查的兩件事:

  1. 指定檔案的**每一個欄位**裝什麼 —— 非空筆數、數值範圍、文字值分佈。
     用來判斷新系統那幾個檔有沒有哪一欄其實對應得上舊格式。

  2. 空列是夾在資料**中間**還是在**尾端**。
     pandas 的 read_excel 會自己裁掉尾端空列, 中間的則變成整列 NaN ——
     差別是「完全沒影響」還是「統計分母虛胖」。

  3. --holes: 查客戶那三條重複偵測的六個疑點 (只讀不寫, 不碰客戶的輸出):
       H1 Has_Coupon 是逐檔判定的, 有 COUPON 欄的檔走第 1 條、沒有的走第 2 條,
          兩邊互斥 -> 同一個票號若橫跨兩種檔, 永遠配不到
       H2 四支各跑一個資料夾, 跨資料夾的重複完全不看
       H3 第 2 條只比 TKT_NUM + FLIGHT_NO, 不比日期 -> 同航班號不同日期的
          兩次離境會被當成重複
       H4 第 3 條只比 TKT_SET_NUM + 級距, 不比航班 -> 同上
       H5 四支都沒有 sorted(), 檔案順序靠 glob -> 誰被判 Keep 可能不可重現
       H6 票號/coupon 的字串形式若各檔不一致 ("1" vs "1.0"), 會靜靜配不到

  4. --outputs: 不碰原始資料, 只讀客戶四支跑出來的 Duplicate_Analysis.xlsx 與
     Duplicate_Details.csv, 回答交付前要確認的三件事:
       a. C1 生效了嗎 —— Rate Range 裡不該再有 "Missing FX"
       b. C2 生效了嗎 —— 2023/2024/2025 該有 A/B/C 級距與非零金額, 不是空白與 0
       c. Proposed Adjustment 的分佈 —— Keep/Adjust/TBC 各多少, 這就是交付內容

  5. --tkt: 票號的前導零到底補不補得回來 —— 用 IATA check digit 驗, 不用猜。
     13 碼票號的最後一碼 = 前 12 碼對 7 取餘。先算 13 碼票號的通過率當**基準線**
     (規則若在這批資料上成立, 該接近 100%; 亂猜是 1/7 = 14.3%), 再看非 13 碼的
     票號補零之後通過率有沒有跳上去。跳上去 = 補零是正確還原。
     最後量客戶的損失: 他們用字串比對, 所以同一張票寫成 12 碼和 13 碼就配不到,
     字串交集與正規化後交集的差, 就是他們漏掉的跨資料夾重複。
     (舊版 --dig 那個「補零後配到 N 個」是錯的 —— int(0043…) == int(43…),
      每個短票號都配到自己, 恆等於票號數。已移除。)

  6. --ara: 全掃驗證 CHANGES.local.md 的 C2 那三個假設。
     預設模式只取樣前 5000 列, 而那幾個檔是按月排序的, 5000 列連第一個月
     都跑不完 —— 等於只看了一個月就下結論。這個模式全掃, 專門看:
       a. 稅額幣別是不是真的全都是 HKD (不能只看第一個月)
       b. 稅額欄的空值比例 (2022_ARA 取樣時是 93.4%, 全檔未知)
       c. 稅額落在 Rate Range 各級距的分佈 (A<=140 / B<=260 / C>260)
       d. 「應申報」與「實際代收」差多少 —— 這就是 payment discrepancy 本身

輸出只有欄名與統計值。個資欄 (見 PII) 連高頻值都不印, 只給筆數與種類數 ——
2023-2025 的檔含 Matching Key (票號+旅客全名)、filter_pax (全名)、姓名、座位號,
而 profile 本來會印文字欄的高頻值、數值欄的 min/max, 兩種都會把它們洩出去。

用法:
    python survey.py                        # 預設: 新系統那幾個檔 + 空列檢查
    python survey.py --ara                  # 全掃驗證 C2 的假設 (見下)
    python survey.py --outliers             # 只看稅額異常大的那些列有沒有規律
    python survey.py --holes                # 查客戶的重複偵測有沒有漏抓/誤判
    python survey.py --outputs              # 讀客戶跑出來的成果, 驗 C1/C2 有沒有生效
    python survey.py --dig --files "KA tkt stk,CX Operate"  # 可只跑指定資料夾
    python survey.py --tkt --files "Payment Discrepancy"  # 前導零 (可只跑一個資料夾)
    python survey.py --intake               # 客戶給新資料就先跑這個
    python survey.py --cols --files 2017    # 只印欄名, 回答「那一欄到底叫什麼」
    python survey.py --pair --files "2022_ARA,2022_Capital"   # 兩個檔的票號交集
    python survey.py --months --files KA_Tkt_CX   # 哪幾個月沒有資料
    python survey.py --recon --files KA_Tkt_CX    # 列數對帳: 有沒有檔案被漏讀
    python survey.py --intake --files 2019  # 只看某個資料夾/檔名片段 (大小寫不拘)
    python survey.py --files 2016,2021      # 檔名含這些片段的
    python survey.py --all                  # 全部 (慢)
    python survey.py --sample 20000         # 統計取樣列數 (預設 5000)
"""
from __future__ import annotations

import argparse
import ast
import csv
import re
import sys
import time
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("需要 openpyxl:  pip install openpyxl")

# ================================================================= 設定
DATA_DIRNAME = "data"
REPORT = "survey_{mode}.txt"
DEFAULT_FILES = ["2022_ARA", "2023.xlsx", "2024.xlsx", "2025.xlsx", "2016.xlsx"]
# --ara: (檔名片段, 實際代收欄, 幣別欄或 None, 應申報欄)
ARA_CHECK = [
    ("2022_ARA",  "Collected Tax", "collect_ccy", "remit_amount"),
    ("2023.xlsx", "Collected Tax", None,          "amount"),
    ("2024.xlsx", "Collected Tax", None,          "Amount"),
    ("2025.xlsx", "Collected Tax", None,          "amount"),
]
BANDS = [(140.0, "A <=140"), (260.0, "B <=260"), (float("inf"), "C >260")]
OUTLIER = 260.0        # 超過這個值就當離群列來看
# 離群列要看哪些「情境欄」有沒有集中 —— 若是外幣, 應該會依航線/辦事處成群
CTX = ["Flight YearMonth", "Category", "Category 2", "Office ID",
       "Inbound Region", "Arrival Region", "inbound_dep_station",
       "flight_arr_station", "FLIGHT_ARR_STATION", "O/D", "Tkt Stk",
       "collect_ccy", "collect_amount_type", "With Collected Tax"]
# 常見幣別對港幣的粗略匯率, 用來認比值
FXHINT = [(19, "JPY"), (170, "KRW"), (2050, "IDR"), (3250, "VND"),
          (11500, "LBP"), (1.15, "CNY"), (0.13, "USD"), (0.11, "EUR"),
          (4.1, "TWD"), (0.55, "MYR"), (16, "NPR"), (11, "INR")]
SAMPLE = 5000
# 這些欄的**值**不能印出來 —— 印了就等於把個資寫進報告檔。
# profile 對文字欄會印高頻值、對數值欄會印 min/max, 兩種都會洩漏:
#   Matching Key = 票號 + 旅客全名, filter_pax = 全名,
#   TKT_NUM 的 min/max 本身就是兩個真實票號。
# 命中的欄只印「有值幾筆 / 幾種」, 值一律不出現。比對用小寫。
PII = {"matching key", "filter_pax", "pax_title", "pax_surname",
       "pax_given_name", "seat_number", "boarding_number",
       "booking_reference", "eb_emd_no", "endorsement",
       "tkt_num", "tkt no", "ticket number", "ticket_no", "ticket no",
       "tkt_nbr", "orig_issue_tkt_num", "original orig issue tkt num",
       "tkt_set_num", "original tkt set num", "coupon_number"}


def is_pii(name: str) -> bool:
    return name.strip().lower() in PII
TOP_TEXT = 6          # 文字欄最多印幾個高頻值
MAXKEEP = 60          # 文字值最多追蹤幾種, 超過就只算「還有更多」
# ================================================================= 設定結束

L: list[str] = []


def add(s: str = "") -> None:
    L.append(s)
    print(s, flush=True)


def txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return str(v)
    return str(v).strip()


def num(v):
    s = txt(v).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def profile(f: Path, rel: str, sample: int) -> None:
    add("")
    add("=" * 72)
    add(f"{rel}")
    add("=" * 72)
    t0 = time.time()
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    declared = ws.max_row
    it = ws.iter_rows(values_only=True)
    hdr = [txt(v) for v in next(it, ())]
    while hdr and not hdr[-1]:
        hdr.pop()
    st = [{"n": 0, "num": 0, "lo": None, "hi": None, "sum": 0.0,
           "vals": Counter(), "over": 0} for _ in hdr]

    rows = 0            # 有資料的列
    idx = 1             # 目前列號 (表頭是 1)
    last_data = 1
    blanks = 0
    blanks_before_last = 0
    for r in it:
        idx += 1
        if r is None or all(v is None for v in r):
            blanks += 1
            continue
        last_data = idx
        blanks_before_last = blanks
        rows += 1
        if rows > sample:
            continue                      # 只數列, 不再算統計
        for i in range(len(hdr)):
            v = r[i] if i < len(r) else None
            t = txt(v)
            if not t:
                continue
            s = st[i]
            s["n"] += 1
            x = num(v)
            if x is None:
                if len(s["vals"]) < MAXKEEP:
                    s["vals"][t] += 1
                else:
                    s["over"] += 1
            else:
                s["num"] += 1
                s["sum"] += x
                s["lo"] = x if s["lo"] is None else min(s["lo"], x)
                s["hi"] = x if s["hi"] is None else max(s["hi"], x)
    wb.close()

    shown = min(rows, sample)
    add(f"  {len(hdr)} 欄   有資料 {rows:,} 列   (dimension 宣告 {declared:,})")
    add(f"  以下統計取樣自前 {shown:,} 列")
    add("")
    for i, c in enumerate(hdr):
        s = st[i]
        name = c if c else f"(第{i+1}欄無標題)"
        if not s["n"]:
            add(f"    {name:<34} 全空")
            continue
        pct = s["n"] / shown if shown else 0
        head = f"    {name:<34} 有值 {s['n']:>6,} ({pct:>5.1%})"
        if is_pii(c):
            kinds = len(s["vals"]) + (s["num"] and 1 or 0)
            add(head + f"  [個資欄, 值不輸出]  不同值 {kinds}{'+' if s['over'] else ''} 種")
            continue
        if s["num"]:
            add(head + f"  數值 {s['num']:>6,}  "
                       f"{s['lo']:,.2f} ~ {s['hi']:,.2f}  平均 {s['sum']/s['num']:,.2f}")
        if s["vals"]:
            top = ", ".join(f"{v}×{n}" for v, n in s["vals"].most_common(TOP_TEXT))
            more = f"  (共 {len(s['vals'])}+ 種)" if s["over"] else f"  (共 {len(s['vals'])} 種)"
            add((head if not s["num"] else " " * len(head.split("有值")[0]) + " " * 8)
                + f"  文字: {top}{more}")

    add("")
    trailing = blanks - blanks_before_last
    add(f"  空列 {blanks:,}   中間 {blanks_before_last:,}   尾端 {trailing:,}")
    if blanks == 0:
        add("  -> 沒有空列。")
    elif blanks_before_last == 0:
        add("  -> 全部在尾端。pandas 的 read_excel 會自己裁掉, 對客戶結果沒有影響。")
    else:
        add(f"  -> 有 {blanks_before_last:,} 列夾在資料中間, pandas 會讀成整列 NaN。")
        add("     那些列 TKT_NUM 是空的, 三條偵測都有 TKT_NUM<>'' 條件, 不會被誤判成")
        add("     重複; 但 Total Rows 與 by-file 統計的分母會虛胖。")
    add(f"  ({time.time() - t0:.0f}s)")


def check_ara(data: Path) -> None:
    """全掃那幾個檔, 驗證 C2 的三個假設。只輸出計數與合計。"""
    add("")
    add("=" * 72)
    add("ARA 全掃 —— 驗證 C2 的假設 (不取樣)")
    add("=" * 72)
    for frag, amt_c, ccy_c, alt_c in ARA_CHECK:
        f = next((x for x in data.rglob("*.xlsx")
                  if frag in x.name and not x.name.startswith("~$")), None)
        if f is None:
            add(f"  !! 找不到含 {frag!r} 的檔")
            continue
        t0 = time.time()
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = [txt(v) for v in next(it, ())]
        ix = {c: i for i, c in enumerate(hdr)}
        low = {c.lower(): i for i, c in enumerate(hdr)}

        def col(name):
            return ix.get(name, low.get(name.lower()))

        ia = col(amt_c)
        ic = col(ccy_c) if ccy_c else None
        il = col(alt_c)
        rows = amt_empty = alt_empty = 0
        amt_bands = Counter()
        alt_bands = Counter()
        ccy = Counter()
        amt_lo = amt_hi = None
        amt_sum = alt_sum = 0.0
        diff_n = 0
        diff_sum = 0.0
        for r in it:
            if r is None or all(v is None for v in r):
                continue
            rows += 1
            a = num(r[ia]) if ia is not None and ia < len(r) else None
            l = num(r[il]) if il is not None and il < len(r) else None
            if ic is not None and ic < len(r):
                ccy[txt(r[ic]) or "(空)"] += 1
            if a is None:
                amt_empty += 1
            else:
                amt_sum += a
                amt_lo = a if amt_lo is None else min(amt_lo, a)
                amt_hi = a if amt_hi is None else max(amt_hi, a)
                amt_bands[next(n for b, n in BANDS if a <= b)] += 1
            if l is None:
                alt_empty += 1
            else:
                alt_sum += l
                alt_bands[next(n for b, n in BANDS if l <= b)] += 1
            if a is not None and l is not None and abs(a - l) > 0.005:
                diff_n += 1
                diff_sum += a - l
        wb.close()
        add("")
        add(f"--- {f.relative_to(data).as_posix()}   {rows:,} 列 ---")
        if ccy_c:
            add(f"  {ccy_c} (全檔): " + ", ".join(f"{k}×{v:,}" for k, v in ccy.most_common(10)))
        else:
            add(f"  (這個檔沒有幣別欄 —— C2 假設它已是港幣等值)")
        got = rows - amt_empty
        add(f"  {amt_c:<16} 有值 {got:,} ({got/rows:.1%})   空 {amt_empty:,}")
        if got:
            add(f"  {'':<16} 範圍 {amt_lo:,.2f} ~ {amt_hi:,.2f}   合計 {amt_sum:,.2f}")
            for _, n in BANDS:
                add(f"  {'':<16}   {n:<9} {amt_bands[n]:>9,} 列 ({amt_bands[n]/got:>6.1%})")
        gl = rows - alt_empty
        add(f"  {alt_c:<16} 有值 {gl:,} ({gl/rows:.1%})   合計 {alt_sum:,.2f}")
        for _, n in BANDS:
            add(f"  {'':<16}   {n:<9} {alt_bands[n]:>9,} 列")
        add(f"  兩者不同的列: {diff_n:,} ({diff_n/rows:.1%})   "
            f"差額合計 {diff_sum:,.2f}   <- payment discrepancy")
        add(f"  ({time.time() - t0:.0f}s)")


def load_client_mapping(root: Path) -> dict:
    """把**所有**客戶腳本的 column_mapping 聯集起來。

    以前是「取筆數最多的那一支」, 問題是哪一支中選會隨著腳本被改動而變 ——
    改了其中一支, 勘查用的字典就整個換掉了。而各支的字典並不相同(例如
    "FLIGHT MONTH" 那一筆只加在其中一支), 用錯的後果是把有的欄看成「沒有」,
    某幾個年度的月份欄就會被誤報成缺。
    勘查的目的是**盤點資料**, 不是重現某一支的行為, 所以認得越多欄名越好。
    """
    merged: dict = {}
    srcs: list[str] = []
    clash: list[str] = []
    for p in sorted(root.glob("*.py")):
        if p.name in ("survey.py", "run_all.py"):
            continue
        try:
            tree = ast.parse(p.read_text(encoding="utf-8", errors="replace"))
        except SyntaxError:
            continue
        for n in ast.walk(tree):
            if isinstance(n, ast.Assign) and getattr(n.targets[0], "id", "") == "column_mapping":
                try:
                    m = ast.literal_eval(n.value)
                except Exception:
                    continue
                srcs.append(f"{p.name[:28]}({len(m)})")
                for k, v in m.items():
                    if k in merged and merged[k] != v:
                        clash.append(f"{k}: {merged[k]} vs {v}")
                    merged.setdefault(k, v)
    if not merged:
        add("!! 找不到 column_mapping, 欄位辨識會不準")
        return merged
    add(f"欄名對映: {len(merged)} 筆, 聯集自 {len(srcs)} 支客戶腳本")
    add("  " + ", ".join(srcs))
    if clash:
        add(f"  !! 同一個欄名在不同腳本對到不同結果: {'; '.join(clash[:5])}")
    return merged


def tkt_key(s: str):
    """票號正規化成可比對的鍵。"1234567890123" 與 "1234567890123.0" 會歸一,
    這正是 H6 要問的 —— 客戶的程式沒有做這步, 所以那兩者在他們那邊配不到。"""
    s = s.strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return s


def check_holes(root: Path, data: Path) -> None:
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")},
                     key=lambda d: str(d).lower())
    per_folder_tkt: dict[str, set] = {}

    add("")
    add("=" * 72)
    add("H5 — 檔案順序是不是穩定的")
    add("=" * 72)
    add("  客戶四支都沒有 sorted(), 檔案順序直接用 glob 的結果。")
    add("  glob 順序決定 _Original_Row_Order, 而 Case 1 排序的最後一層就是它")
    add("  -> 順序若不穩, 同一組裡誰被判 Keep、誰被判 Adjust 可能對調。")
    for d in folders:
        g = [f.name for f in d.glob("*.xlsx") if not f.name.startswith("~$")]
        same = g == sorted(g)
        add(f"  [{d.relative_to(data).as_posix()}]  glob 與 sorted "
            + ("一致 (這台機器上暫時安全)" if same else "!! 不一致"))

    for d in folders:
        rel = d.relative_to(data).as_posix()
        add("")
        add("=" * 72)
        add(f"[{rel}]")
        add("=" * 72)
        t0 = time.time()
        y_tkt, n_tkt = set(), set()
        y_rows = n_rows = 0
        y_files, n_files = [], []
        raw_forms: dict[str, Counter] = {}
        # H3: (票號, 航班號) -> 見過的航班日期
        pair_dates: dict[tuple, set] = {}
        # H4: 票證組號 -> 見過的 (航班號, 航班日期)
        set_flights: dict[object, set] = {}

        for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower()):
            if f.name.startswith("~$"):
                continue
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            hdr = rn([txt(v) for v in next(it, ())])
            ix = {c: i for i, c in enumerate(hdr)}
            has_cpn = "COUPON_NUMBER" in ix
            it_, ifn, ifd = ix.get("TKT_NUM"), ix.get("FLIGHT_NO"), ix.get("FLIGHT_DATE")
            isn = ix.get("TKT_SET_NUM")
            rc = Counter()
            n = 0
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                n += 1
                raw = txt(r[it_]) if it_ is not None and it_ < len(r) else ""
                k = tkt_key(raw)
                if k is None:
                    continue
                if n <= 200000:          # 樣態只看前面, 夠代表了
                    rc["含小數點" if "." in raw else "純數字/文字"] += 1
                    rc[f"長度{len(raw)}"] += 1
                (y_tkt if has_cpn else n_tkt).add(k)
                fno = txt(r[ifn]) if ifn is not None and ifn < len(r) else ""
                fdt = txt(r[ifd]) if ifd is not None and ifd < len(r) else ""
                if not has_cpn and fno:
                    pair_dates.setdefault((k, fno), set()).add(fdt)
                sn = tkt_key(txt(r[isn])) if isn is not None and isn < len(r) else None
                if sn is not None:
                    set_flights.setdefault(sn, set()).add((fno, fdt))
            wb.close()
            (y_files if has_cpn else n_files).append(f.name)
            if has_cpn:
                y_rows += n
            else:
                n_rows += n
            raw_forms[f.name] = rc

        add("")
        add("  H6 — 票號的字串樣態 (各檔若不一致, 客戶的程式會配不到)")
        for fn, rc in raw_forms.items():
            forms = ", ".join(f"{k}×{v:,}" for k, v in rc.most_common(4))
            add(f"    {fn[:52]:<54} {forms}")
        dots = [fn for fn, rc in raw_forms.items() if rc.get("含小數點")]
        if dots:
            add(f"    -> !! 有 {len(dots)} 個檔的票號含小數點, 與其他檔配不到")
        else:
            add("    -> 所有檔的票號都沒有小數點")

        add("")
        add("  H1 — Has_Coupon 把資料切成兩個互不相通的偵測宇宙")
        add(f"    有 COUPON 欄 (走第 1 條): {len(y_files)} 個檔, {y_rows:,} 列")
        add(f"    沒 COUPON 欄 (走第 2 條): {len(n_files)} 個檔, {n_rows:,} 列")
        if y_files and n_files:
            both = y_tkt & n_tkt
            add(f"    !! 兩邊都出現的票號: {len(both):,} 個")
            add(f"       這些票號的列分屬兩條偵測, 客戶的程式**永遠配不到**")
            add(f"       (第 1 條 .eq('Y')、第 2 條 .eq('N'), 兩者互斥)")
            add(f"       沒 COUPON 那側的檔: {', '.join(fn[:40] for fn in n_files)}")
        else:
            add("    這個資料夾只有一種, 不會被切開")

        add("")
        add("  H3 — 第 2 條偵測不比日期")
        multi = {k: v for k, v in pair_dates.items() if len(v) > 1}
        if pair_dates:
            add(f"    (票號, 航班號) 組合 {len(pair_dates):,} 個")
            add(f"    !! 其中 {len(multi):,} 個對應到**多個不同航班日期** "
                f"({len(multi)/len(pair_dates):.2%})")
            add(f"       那是同一張票飛同一航班號的不同天 = 兩次離境 = 該收兩次稅,")
            add(f"       但第 2 條只比 TKT_NUM + FLIGHT_NO, 會判成重複。")
        else:
            add("    這個資料夾沒有走第 2 條的檔")

        add("")
        add("  H4 — 第 3 條偵測不比航班")
        ms = {k: v for k, v in set_flights.items() if len(v) > 1}
        if set_flights:
            add(f"    TKT_SET_NUM 共 {len(set_flights):,} 個")
            add(f"    !! 其中 {len(ms):,} 個對應到**多個不同航班** "
                f"({len(ms)/len(set_flights):.2%})")
            add(f"       同一票證組號的不同航段本來就該各收一次稅,")
            add(f"       但第 3 條只比 TKT_SET_NUM + 級距, 級距又幾乎都是 A。")
        add(f"    ({time.time() - t0:.0f}s)")
        per_folder_tkt[rel] = y_tkt | n_tkt
        del pair_dates, set_flights, y_tkt, n_tkt

    add("")
    add("=" * 72)
    add("H2 — 跨資料夾的重複, 四支都看不到")
    add("=" * 72)
    names = list(per_folder_tkt)
    for i in range(len(names)):
        add(f"  [{names[i]}]  不重複票號 {len(per_folder_tkt[names[i]]):,} 個")
    add("")
    add("  兩兩交集:")
    found = False
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            inter = per_folder_tkt[names[i]] & per_folder_tkt[names[j]]
            if inter:
                found = True
                add(f"    !! {names[i][:34]:<36} ∩ {names[j][:34]:<36} {len(inter):,} 個")
    if not found:
        add("    四個資料夾的票號完全不重疊 -> 跨資料夾重複不存在, H2 不成立")
    else:
        add("    這些票號同時出現在多個資料夾, 而四支各跑一個資料夾,")
        add("    所以跨資料夾的重複收費**沒有任何一支看得到**。")


def _find(hdr: list[str], *names):
    """欄名在各年度大小寫/底線不一致, 用寬鬆比對找欄位位置。"""
    low = {c.strip().lower().replace("_", " "): i for i, c in enumerate(hdr)}
    for n in names:
        k = n.strip().lower().replace("_", " ")
        if k in low:
            return low[k]
    return None


# --------------------------------------------------- H2 / H4 / H6 深入
GAPCOLS = ["Day between Sch.Dep and Inbound Dep",
           "Day between Inbound Dep & Flight Date"]


def day_of(v) -> str:
    """把各種寫法的航班日期收斂成日曆日。跨不跨日是轉機豁免的判準,
    所以只比日期不比時間。"""
    s = txt(v)
    if not s:
        return ""
    try:
        f = float(s)
        if 20000 < f < 60000:        # Excel 日期序號
            return (date(1899, 12, 30) + timedelta(days=int(f))).isoformat()
    except ValueError:
        pass
    return s[:10]


def check_dig(root: Path, data: Path, pats: list[str] | None = None) -> None:
    """H6 -> H4 -> H2 依序做, 因為 H6 的結論會影響 H2 的解讀。"""
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")
                      and "duplicate_analysis" not in f.name.lower()},
                     key=lambda d: str(d).lower())
    if pats:
        folders = [d for d in folders
                   if any(p.lower() in d.relative_to(data).as_posix().lower()
                          for p in pats)]
        if not folders:
            _no_hit(pats)
            return

    tkt_by_folder: dict[str, set] = {}
    short_raw: dict[str, Counter] = {}      # H6: 非 13 碼的原始字串
    all_raw: dict[int, set] = {}            # H6: 長度 -> 該長度的票號集合

    add("")
    add("=" * 72)
    add("第一趟 —— H6 票號樣態 + H4 轉機豁免")
    add("=" * 72)

    for d in folders:
        rel = d.relative_to(data).as_posix()
        t0 = time.time()
        tkts: set = set()
        lens: Counter = Counter()
        shorts: Counter = Counter()
        stk_of_short: Counter = Counter()
        # H4: 票證組號 -> [列數, 首個(航班,日期)雜湊, 是否全同, 首個日曆日, 是否同日]
        grp: dict = {}
        gap_hist: Counter = Counter()
        gap_col_found = set()
        rows_total = 0

        for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower()):
            if f.name.startswith("~$") or "duplicate_analysis" in f.name.lower():
                continue
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            hdr = rn([txt(v) for v in next(it, ())])
            ix = {c: i for i, c in enumerate(hdr)}
            it_, ifn = ix.get("TKT_NUM"), ix.get("FLIGHT_NO")
            ifd, isn = ix.get("FLIGHT_DATE"), ix.get("TKT_SET_NUM")
            istk = ix.get("Tkt Stk", ix.get("Ticket Stk"))
            igap = next((ix[c] for c in GAPCOLS if c in ix), None)
            if igap is not None:
                gap_col_found.add(hdr[igap])
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                rows_total += 1
                raw = txt(r[it_]) if it_ is not None and it_ < len(r) else ""
                k = tkt_key(raw)
                if k is not None:
                    tkts.add(k)
                    lens[len(raw)] += 1
                    if len(raw) and len(raw) != 13 and len(shorts) < 40000:
                        shorts[raw] += 1
                        if istk is not None and istk < len(r):
                            stk_of_short[txt(r[istk]) or "(空)"] += 1
                if igap is not None and igap < len(r):
                    g = num(r[igap])
                    if g is not None and -1 < g < 40:
                        gap_hist[int(g)] += 1
                sn = tkt_key(txt(r[isn])) if isn is not None and isn < len(r) else None
                if sn is None:
                    continue
                fno = txt(r[ifn]) if ifn is not None and ifn < len(r) else ""
                dy = day_of(r[ifd]) if ifd is not None and ifd < len(r) else ""
                fk = hash((fno, dy))
                e = grp.get(sn)
                if e is None:
                    grp[sn] = [1, fk, 1, dy, 1]
                else:
                    e[0] += 1
                    if e[1] != fk:
                        e[2] = 0
                    if e[3] != dy:
                        e[4] = 0
            wb.close()

        add("")
        add("-" * 72)
        add(f"[{rel}]   {rows_total:,} 列")
        add("-" * 72)
        add("  H6 — 票號長度分佈")
        for L, n in sorted(lens.items()):
            mark = "" if L == 13 else "   <- 非 13 碼"
            add(f"    {L} 碼   {n:>10,} 列{mark}")
        n_short = sum(v for L, v in lens.items() if L != 13)
        if n_short and shorts:
            add(f"    非 13 碼共 {n_short:,} 列, {len(shorts):,} 個不同票號")
            add(f"    這些票號的 Tkt Stk: "
                + ", ".join(f"{k}×{v:,}" for k, v in stk_of_short.most_common(6)))
            add("    (補零能不能還原、客戶因此漏掉多少 -> 用 --tkt, 這裡量不出來)")

        add("")
        add("  H4 — 轉機豁免 (修正版: 按日曆日分)")
        if gap_col_found:
            add(f"    找到日差欄: {', '.join(sorted(gap_col_found))}")
            tg = sum(gap_hist.values())
            for g in sorted(gap_hist):
                lab = "  <- 當日, 屬轉機, 應豁免" if g == 0 else ""
                add(f"      日差 {g}: {gap_hist[g]:>10,} 列 ({gap_hist[g]/tg:>6.1%}){lab}")
            add(f"    !! 日差 0 共 {gap_hist.get(0, 0):,} 列 "
                f"({gap_hist.get(0, 0)/tg:.1%}) —— 這些是當日轉機")
            add("       客戶的字典沒收這個欄名, 三條偵測都沒用到它")
        else:
            add("    這個資料夾的檔案沒有日差欄")

        multi = {k: v for k, v in grp.items() if v[0] > 1}
        same_all = sum(1 for v in multi.values() if v[2])
        same_day = sum(1 for v in multi.values() if not v[2] and v[4])
        diff_day = sum(1 for v in multi.values() if not v[2] and not v[4])
        add(f"    TKT_SET_NUM 共 {len(grp):,} 個, 其中 {len(multi):,} 個有多列:")
        add(f"      同航班同日期 (真重複)          {same_all:>9,}"
            + (f" ({same_all/len(multi):.1%})" if multi else ""))
        add(f"      同一日曆日、不同航班 (當日接駁) {same_day:>9,}"
            + (f" ({same_day/len(multi):.1%})" if multi else ""))
        add(f"      不同日曆日 (各自獨立的離境)     {diff_day:>9,}"
            + (f" ({diff_day/len(multi):.1%})" if multi else ""))
        add("      -> 客戶第 3 條偵測把這三種**全部**當成重複")
        add(f"    ({time.time() - t0:.0f}s)")

        tkt_by_folder[rel] = tkts
        del grp

    # ---- H2 第二趟: 只針對有交集的資料夾對 ----
    add("")
    add("=" * 72)
    add("第二趟 —— H2 跨資料夾的交集到底是什麼")
    add("=" * 72)
    names = list(tkt_by_folder)
    pairs = []
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            inter = tkt_by_folder[names[i]] & tkt_by_folder[names[j]]
            if inter:
                pairs.append((names[i], names[j], inter))
                add(f"  {names[i][:32]:<34} ∩ {names[j][:32]:<34} {len(inter):,} 個")
    if not pairs:
        add("  沒有交集")
        return
    del tkt_by_folder

    for a, b, inter in pairs:
        add("")
        add("-" * 72)
        add(f"{a}\n  ∩ {b}   {len(inter):,} 個共同票號")
        add("-" * 72)
        seen: dict = {}          # 票號 -> {資料夾: set of (航班, 日曆日)}
        for rel in (a, b):
            d = data / rel
            t0 = time.time()
            for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower()):
                if f.name.startswith("~$") or "duplicate_analysis" in f.name.lower():
                    continue
                wb = load_workbook(f, read_only=True, data_only=True)
                ws = wb.worksheets[0]
                it = ws.iter_rows(values_only=True)
                hdr = rn([txt(v) for v in next(it, ())])
                ix = {c: i for i, c in enumerate(hdr)}
                it_, ifn, ifd = ix.get("TKT_NUM"), ix.get("FLIGHT_NO"), ix.get("FLIGHT_DATE")
                # codeshare: 同一次離境在兩邊可能掛不同航班號 (一邊 CX100、
                # 另一邊 KA5100)。只比航班號會把它判成「不同航班」而漏掉,
                # 所以同時用航線 (出發站, 到達站) 當第二種判準 —— 航線不受
                # 航班號怎麼掛影響。這個資料夾組合本身就是 KA 票證 / CX 承運。
                idep = _find(hdr, "FLIGHT_DEP_STATION", "flight_dep_station")
                iarr = _find(hdr, "FLIGHT_ARR_STATION", "flight_arr_station")
                for r in it:
                    if r is None or all(v is None for v in r):
                        continue
                    k = tkt_key(txt(r[it_])) if it_ is not None and it_ < len(r) else None
                    if k not in inter:
                        continue
                    fno = txt(r[ifn]) if ifn is not None and ifn < len(r) else ""
                    dy = day_of(r[ifd]) if ifd is not None and ifd < len(r) else ""
                    dep = txt(r[idep]) if idep is not None and idep < len(r) else ""
                    arr = txt(r[iarr]) if iarr is not None and iarr < len(r) else ""
                    e = seen.setdefault(k, {}).setdefault(rel, [set(), set()])
                    e[0].add((fno, dy))
                    e[1].add((dy, dep, arr))
                wb.close()
            add(f"  掃完 {rel[:44]}  ({time.time() - t0:.0f}s)")
        both = exact = sameday = diff = 0
        route = route_only = 0
        for k, d2 in seen.items():
            if len(d2) < 2:
                continue
            both += 1
            sa, sb = d2[a][0], d2[b][0]
            ra_, rb_ = d2[a][1], d2[b][1]
            hit_fno = bool(sa & sb)
            # 航線判準: 同一天、同一個出發站與到達站 (兩站都要有值才算數)
            hit_rt = bool({x for x in ra_ if x[1] and x[2]}
                          & {x for x in rb_ if x[1] and x[2]})
            if hit_fno:
                exact += 1
            elif {x[1] for x in sa} & {x[1] for x in sb}:
                sameday += 1
            else:
                diff += 1
            if hit_rt:
                route += 1
                if not hit_fno:
                    route_only += 1
        add("")
        add(f"  兩邊都有紀錄的票號 {both:,} 個:")
        add(f"    完全相同的 (航班, 日期)  {exact:>9,}"
            + (f" ({exact/both:.1%})" if both else "") + "   <- 同一次離境被記兩次")
        add(f"    同一日曆日、不同航班      {sameday:>9,}"
            + (f" ({sameday/both:.1%})" if both else "") + "   <- 當日接駁")
        add(f"    完全不同的日期            {diff:>9,}"
            + (f" ({diff/both:.1%})" if both else "") + "   <- 各自獨立的離境")
        add("    -> 第一類是真的跨資料夾重複收費, 而四支腳本沒有一支看得到。")
        add("")
        add("  改用航線判準 (同一天 + 同出發站 + 同到達站, 不看航班號):")
        add(f"    同一天飛同一條航線      {route:>9,}"
            + (f" ({route/both:.1%})" if both else ""))
        add(f"    !! 其中航班號對不上的   {route_only:>9,}"
            + (f" ({route_only/both:.1%})" if both else "")
            + "   <- codeshare, 用航班號會漏掉")
        add("    -> 這個資料夾組合是 KA 票證 / CX 承運, 同一次離境在兩邊掛不同")
        add("       航班號是預期中的事。route_only 若明顯 > 0, 代表用航班號當")
        add("       判準會低估跨資料夾的重複。")
        del seen


# --------------------------------------------------- 票號前導零能不能還原
# check digit 的候選規則。哪一條在這批資料上成立**由資料決定**, 不由我猜 ——
# 第一版寫死「前 12 碼 % 7」, 實測 13 碼票號的通過率只有 10.0% (比亂猜的
# 14.3% 還低), 規則根本不成立。所以改成全部試一遍, 取通過率最高的。
CD_RULES = [
    ("前12碼 % 7 == 第13碼",    lambda s: int(s[:12]) % 7 == int(s[12])),
    ("第4~12碼 % 7 == 第13碼",  lambda s: int(s[3:12]) % 7 == int(s[12])),
    ("第4~13碼 % 7 == 0",       lambda s: int(s[3:]) % 7 == 0),
    ("整個13碼 % 7 == 0",       lambda s: int(s) % 7 == 0),
    ("前12碼 % 11 == 第13碼",   lambda s: int(s[:12]) % 11 == int(s[12])),
]
CD_LIMIT = 300000      # check digit 是統計量, 每個資料夾取這麼多列就夠


def cd_tally(s: str, acc: dict) -> None:
    """把一個 13 碼票號記進各條候選規則的計數。"""
    if len(s) != 13 or not s.isdigit():
        return
    for name, fn in CD_RULES:
        e = acc.setdefault(name, [0, 0])
        e[1] += 1
        if fn(s):
            e[0] += 1


def cd_report(acc: dict, label: str) -> float:
    """印出各規則的通過率, 回傳最高的那個。"""
    if not acc:
        add(f"    {label}: 沒有可驗的票號")
        return 0.0
    best = 0.0
    for name, fn in CD_RULES:
        p, n = acc.get(name, [0, 0])
        if not n:
            continue
        r = p / n
        best = max(best, r)
        add(f"    {label:<12} {name:<24} {p:>10,} / {n:<10,} = {r:>6.1%}")
    return best


def pad13(s: str) -> str:
    return "0" * (13 - len(s)) + s if 0 < len(s) < 13 else s


def check_tkt(root: Path, data: Path, pats: list[str] | None = None) -> None:
    """補前導零到底還原得對不對 —— 用 check digit 當硬證據, 不用猜。

    先算 13 碼票號的 check digit 通過率當**基準線**: 若基準線接近 100%,
    這條規則在這批資料上成立, 才有資格拿它去驗補零; 若基準線接近 1/7
    (14.3% = 亂猜), 代表規則不適用, 下面的數字一律不能用。

    然後對非 13 碼的票號各驗一次: 原樣 vs 補零後。補零後通過率若跳到
    基準線, 就證明「Excel 吃掉前導零」是對的, 補回去 = 正確還原。

    最後量客戶的損失: 他們用字串比對 (clean_text 只做 strip), 所以
      字串交集 = 兩邊都寫成同一種長度的才配得到
      int 交集   = 補零後真正是同一張票的
    兩者的差, 就是客戶因為前導零而**漏掉**的跨資料夾重複。
    """
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")
                      and "duplicate_analysis" not in f.name.lower()},
                     key=lambda d: str(d).lower())
    if pats:
        folders = [d for d in folders
                   if any(p.lower() in d.relative_to(data).as_posix().lower()
                          for p in pats)]
        if not folders:
            _no_hit(pats)
            return

    std: dict[str, set] = {}       # 資料夾 -> 有 13 碼寫法的票號 (int key)
    nonstd: dict[str, set] = {}    # 資料夾 -> 有非 13 碼寫法的票號 (int key)

    add("")
    add("=" * 72)
    add("票號前導零 —— 補得回來嗎? 客戶漏掉多少?")
    add("=" * 72)

    for d in folders:
        rel = d.relative_to(data).as_posix()
        t0 = time.time()
        s13: set = set()
        sx: set = set()
        base: dict = {}                  # 13 碼: 規則 -> [通過, 可驗總數]
        pad_cd: dict = {}                # 非 13 碼補前導零後: 同上
        n_base = 0                       # 已納入 base 的列數 (受 CD_LIMIT 限制)
        pfx: Counter = Counter()         # 補零後前 3 碼 × Tkt Stk
        pfx_raw: Counter = Counter()     # 補零**前**的前 3 碼 (對照組)
        stk_all: Counter = Counter()     # Tkt Stk 的完整值分佈 (全部列)
        car_short: Counter = Counter()   # 非 13 碼那些列的承運航空
        car_all: Counter = Counter()     # 全部列的承運航空 (當分母)
        lens: Counter = Counter()
        odd: Counter = Counter()         # 非純數字的票號樣本
        no_stk: list[str] = []           # 根本沒有 Tkt Stk 欄的檔

        for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower()):
            if f.name.startswith("~$") or "duplicate_analysis" in f.name.lower():
                continue
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            hdr = rn([txt(v) for v in next(it, ())])
            ix = {c: i for i, c in enumerate(hdr)}
            it_ = ix.get("TKT_NUM")
            istk = _find(hdr, "Tkt Stk", "Ticket Stk", "TKT_STK")
            icar = _find(hdr, "FLIGHT_CARRIER", "flight_carrier")
            if istk is None:
                no_stk.append(f.name)
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                raw = txt(r[it_]) if it_ is not None and it_ < len(r) else ""
                if not raw:
                    continue
                k = tkt_key(raw)
                if k is None:
                    continue
                lens[len(raw)] += 1
                if istk is not None and istk < len(r):
                    stk_all[txt(r[istk]) or "(空值)"] += 1
                else:
                    stk_all["(無此欄)"] += 1
                car = (txt(r[icar]) if icar is not None and icar < len(r) else "") \
                    or ("(空值)" if icar is not None else "(無此欄)")
                car_all[car] += 1
                if len(raw) != 13:
                    car_short[car] += 1
                if not raw.isdigit():
                    if len(odd) < 5:
                        odd[raw[:20]] += 1
                    continue
                if len(raw) == 13:
                    s13.add(k)
                    if n_base < CD_LIMIT:
                        n_base += 1
                        cd_tally(raw, base)
                    continue
                sx.add(k)
                p = pad13(raw)
                cd_tally(p, pad_cd)
                if istk is None:
                    stk = "(無此欄)"
                else:
                    stk = txt(r[istk]) if istk < len(r) else ""
                    stk = stk or "(欄在, 值空)"
                pfx[(p[:3], stk)] += 1
                pfx_raw[raw[:3]] += 1
            wb.close()

        add("")
        add("-" * 72)
        add(f"[{rel}]")
        add("-" * 72)
        add("  長度分佈: " + ", ".join(f"{L}碼×{n:,}" for L, n in sorted(lens.items())))
        if odd:
            add("  !! 非純數字的票號: " + ", ".join(f"{k!r}×{n}" for k, n in odd.items()))

        add("")
        add("  check digit — 五條候選規則各驗一次, 哪條成立由資料決定")
        b = cd_report(base, "13碼(基準線)")
        if base:
            add(f"    -> 基準線最高 {b:.1%}: "
                + ("有一條規則成立, 可以拿來驗補零" if b > 0.95 else
                   "!! 沒有一條成立 (亂猜是 14.3%) —— check digit 這條路"
                   "在這批資料上走不通, 補零的判斷改看下面的前 3 碼"
                   if b < 0.5 else
                   "介於中間 —— 要嘛規則對但資料有雜訊, 要嘛規則不對, 人工看過再用"))
        else:
            add("    13碼(基準線): 這個資料夾沒有 13 碼票號, 沒有基準線可比")
        if pad_cd:
            cd_report(pad_cd, "補零後")
        add("    注意: check digit 是整數運算, 對前導零不敏感 —— 補不補零算出來")
        add("    一樣。所以它就算成立也只證明「這是合法票號」, 證明不了「原本是")
        add("    13 碼」。決定性證據一直是下面的前 3 碼對照。")

        if pfx:
            add("")
            add("  前 3 碼: 補零前 vs 補零後 (IATA 航空公司代碼)")
            add("    補零前: " + ", ".join(f"{p}×{n:,}" for p, n in pfx_raw.most_common(8)))
            add("    補零後 × Tkt Stk:")
            for (p, stk), n in pfx.most_common(12):
                add(f"      {p}  ×  {stk:<14} {n:>10,} 列")
            add("    -> 補零前的前 3 碼若不是合法航空公司代碼、補零後才是,")
            add("       而且與同一列的 Tkt Stk 對得上, 補零就是正確還原。")
            add("       (前 3 碼是 IATA 的三位數航空公司代碼, 對照 IATA 代碼表)")
            add("")
            add("  Tkt Stk 的值分佈 (全部列):")
            add("    " + ", ".join(f"{k}×{v:,}" for k, v in stk_all.most_common(12)))
            add("  承運航空 (flight_carrier):")
            add("    全部列:   " + ", ".join(f"{k}×{v:,}"
                                            for k, v in car_all.most_common(8)))
            add("    非13碼列: " + ", ".join(f"{k}×{v:,}"
                                            for k, v in car_short.most_common(8)))
            add("    -> 票號前綴 = 誰**開票**, flight_carrier = 誰**承運**。")
            add("       非 13 碼的票若前綴是別家、但承運全是自家, 那是 interline/")
            add("       codeshare (別家開票、自家飛), 出現在這裡是正常的。")
        if no_stk:
            add(f"    !! 有 {len(no_stk)} 個檔根本沒有 Tkt Stk 欄: "
                + ", ".join(n[:40] for n in no_stk[:4]))

        # 同一個資料夾內, 同一張票有沒有被寫成兩種長度
        mixed = s13 & sx
        add("")
        add(f"  資料夾內同時有 13 碼與非 13 碼兩種寫法的票號: {len(mixed):,} 個")
        add("    -> 0 代表長度在這個資料夾內是一致的, 資料夾**內部**的比對不受影響。")

        std[rel] = s13
        nonstd[rel] = sx
        add(f"  ({time.time() - t0:.0f}s)")

    # ---- 客戶用字串比對, 因為前導零漏掉多少跨資料夾重複 ----
    add("")
    add("=" * 72)
    add("客戶用字串比對, 因為前導零漏掉的跨資料夾重複")
    add("=" * 72)
    names = list(std)
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            a, b = names[i], names[j]
            ai, bi = std[a] | nonstd[a], std[b] | nonstd[b]
            inter = ai & bi
            if not inter:
                continue
            # 字串配得到 = 兩邊都有 13 碼寫法, 或兩邊都有非 13 碼寫法
            hit = (std[a] & std[b]) | (nonstd[a] & nonstd[b])
            lost = inter - hit
            add("")
            add(f"  {a}")
            add(f"    ∩ {b}")
            add(f"    正規化後真正共同的票號   {len(inter):,}")
            add(f"    客戶的字串比對配得到     {len(hit):,}")
            add(f"    因前導零而漏掉           {len(lost):,} "
                f"({len(lost)/len(inter):.1%})"
                + ("   <- 0 = 兩邊長度一致, 前導零沒造成損失" if not lost else "   !!"))


def _no_hit(pats) -> None:
    add("")
    add(f"  !! 沒有任何檔案符合 --files {pats}")
    add("     檔名片段要跟實際檔名對得上 (大小寫不拘, 但空格和底線要一致);")
    add("     不加 --files 就是全掃。")


def check_recon(data: Path, pats: list[str] | None) -> None:
    """原始檔的列數 vs 產出 Summary Pivot 的列數, 逐檔對帳。

    抓三種「有東西沒對上」的情況, 這些都不會讓程式報錯:
      a. 資料夾裡有的檔, Summary Pivot 沒有 -> 被漏讀 (或不在該腳本的清單內)
      b. Summary Pivot 有的名字, 資料夾裡不是原始檔 -> 產出被當輸入讀進去
      c. 兩邊都有但列數不同 -> 讀進去的跟檔案裡的不一樣

    原始列數是逐列數出來的, 不用 ws.max_row —— 那是 dimension 宣告,
    實測會虛胖 (例如宣告 274,183 但實際 274,182)。
    """
    add("")
    add("=" * 72)
    add("列數對帳 —— 原始檔 vs 產出的 Summary Pivot")
    add("=" * 72)
    done = 0

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")},
                     key=lambda d: str(d).lower())
    for d in folders:
        rel = d.relative_to(data).as_posix()
        low = rel.lower()
        src = [f for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower())
               if not f.name.startswith("~$")
               and "duplicate_analysis" not in f.name.lower()
               and (pats is None
                    or any(p.lower() in f.name.lower() or p.lower() in low
                           for p in pats))]
        outs = [f for f in sorted(d.glob("*.xlsx"))
                if "duplicate_analysis" in f.name.lower()
                and not f.name.startswith("~$")]
        if not src or not outs:
            continue
        done += 1
        add("")
        add("-" * 72)
        add(f"[{rel}]")
        add("-" * 72)

        real: dict = {}
        for f in src:
            t0 = time.time()
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            next(it, None)                      # 表頭
            n = sum(1 for r in it
                    if r is not None and not all(v is None for v in r))
            wb.close()
            real[f.name] = n
            add(f"  數完 {f.name}   {n:,} 列   ({time.time() - t0:.0f}s)")

        for o in outs:
            wb = load_workbook(o, read_only=True, data_only=True)
            piv: dict = {}
            for nm in wb.sheetnames:
                if "summary pivot" not in nm.lower():
                    continue
                ws = wb[nm]
                it = ws.iter_rows(values_only=True)
                hdr = [txt(v) for v in next(it, ())]
                isf = _find(hdr, "Source_File", "Source File")
                ict = _find(hdr, "Total_Ticket_Number_Count",
                            "Total Ticket Number Count")
                for r in it:
                    if r is None or all(v is None for v in r):
                        continue
                    sf = txt(r[isf]) if isf is not None and isf < len(r) else ""
                    if not sf or sf.lower().startswith("grand"):
                        continue
                    piv[sf] = int(num(r[ict]) or 0) if ict is not None else 0
            wb.close()

            add("")
            add(f"  === {o.name} ===")
            add(f"    {'檔名':<44}{'原始':>10}{'Pivot':>10}{'差':>9}")
            bad = 0
            for k in sorted(set(real) | set(piv)):
                a, b = real.get(k), piv.get(k)
                if a is None:
                    add(f"    {k[:44]:<44}{'—':>10}{b:>10,}{'':>9}"
                        "   !! Pivot 有但不是原始檔 -> 產出被當輸入讀進去")
                    bad += 1
                elif b is None:
                    add(f"    {k[:44]:<44}{a:>10,}{'—':>10}{'':>9}"
                        "   <- 沒被讀 (或不在該腳本的檔案清單內)")
                elif a != b:
                    add(f"    {k[:44]:<44}{a:>10,}{b:>10,}{b - a:>+9,}   !!")
                    bad += 1
                else:
                    add(f"    {k[:44]:<44}{a:>10,}{b:>10,}{0:>9}")
            add(f"    -> {'全部對得上' if not bad else f'!! 有 {bad} 筆對不上'}")

    if not done:
        _no_hit(pats)


def _mrange(a: str, b: str):
    y, m = int(a[:4]), int(a[5:7])
    ye, me = int(b[:4]), int(b[5:7])
    while (y, m) <= (ye, me):
        yield f"{y:04d}-{m:02d}"
        m += 1
        if m > 12:
            m, y = 1, y + 1


def check_months(root: Path, data: Path, pats: list[str] | None) -> None:
    """每個檔涵蓋哪些月份, 區間內缺了哪幾個月。

    「某年某月沒有結果」有三種原因, 混在一起就查不出來:
      a. 那個月的資料在檔案裡, 但月份欄是空的 (欄名沒被字典收 -> Month 全空)
      b. 那個月本來就沒有列 (停飛、系統轉換空窗)
      c. 整個年度的檔案根本沒給
    這支把三者分開: 逐檔列出實際月份、標出區間內的缺口, 最後在資料夾層級
    列出所有檔合起來涵蓋的範圍與缺口。
    """
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")
                      and "duplicate_analysis" not in f.name.lower()},
                     key=lambda d: str(d).lower())
    add("")
    add("=" * 72)
    add("月份覆蓋 —— 哪幾個月沒有資料")
    add("=" * 72)
    done = 0

    for d in folders:
        rel = d.relative_to(data).as_posix()
        low = rel.lower()
        files = [f for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower())
                 if not f.name.startswith("~$")
                 and "duplicate_analysis" not in f.name.lower()
                 and (pats is None
                      or any(p.lower() in f.name.lower() or p.lower() in low
                             for p in pats))]
        if not files:
            continue
        done += 1
        add("")
        add("-" * 72)
        add(f"[{rel}]")
        add("-" * 72)
        seen_all: Counter = Counter()

        for f in files:
            t0 = time.time()
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            hdr = rn([txt(v) for v in next(it, ())])
            im = hdr.index("Flight YearMonth") if "Flight YearMonth" in hdr else None
            mc: Counter = Counter()
            n = 0
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                n += 1
                v = txt(r[im]) if im is not None and im < len(r) else ""
                mc[v[:7] if len(v) >= 7 else (v or "(空)")] += 1
            wb.close()

            add("")
            add(f"  --- {f.name} ---   {n:,} 列   ({time.time() - t0:.0f}s)")
            if im is None:
                add("    !! 這個檔沒有 Flight YearMonth 欄 (字典也對映不到)")
                add("       -> 月份會是空的, 該檔所有列都排不進任何月份")
                continue
            ok = sorted(k for k in mc if len(k) == 7 and k[4] == "-"
                        and k[:4].isdigit() and k[5:7].isdigit())
            bad = {k: v for k, v in mc.items() if k not in ok}
            if not ok:
                add("    !! 沒有任何一列的月份是 YYYY-MM 格式")
            else:
                for i in range(0, len(ok), 4):
                    add("    " + "   ".join(f"{k} {mc[k]:>7,}" for k in ok[i:i + 4]))
                gap = [m for m in _mrange(ok[0], ok[-1]) if m not in set(ok)]
                add(f"    涵蓋 {ok[0]} ~ {ok[-1]}, 共 {len(ok)} 個月"
                    + (f", **區間內缺 {len(gap)} 個月: {', '.join(gap)}**"
                       if gap else ", 中間沒有斷"))
            for k, v in bad.items():
                add(f"    !! 月份非 YYYY-MM: {k!r} × {v:,} 列")
            seen_all.update(mc)

        ok = sorted(k for k in seen_all if len(k) == 7 and k[4] == "-"
                    and k[:4].isdigit() and k[5:7].isdigit())
        if ok:
            gap = [m for m in _mrange(ok[0], ok[-1]) if m not in set(ok)]
            add("")
            add(f"  這個資料夾合計涵蓋 {ok[0]} ~ {ok[-1]}, 共 {len(ok)} 個月")
            if gap:
                add(f"  !! 整體缺 {len(gap)} 個月: {', '.join(gap)}")
                add("     (要分辨是「那幾個月停飛/沒交易」還是「檔案沒給」——")
                add("      若缺口正好是一整個年度, 多半是後者。)")
            else:
                add("  中間沒有斷月。")

        # 產出的 Details 只收 flagged 的列, 跟原始資料並排才分得出
        # 「那個月沒資料」和「有資料但沒抓到重複」。
        for csvf in sorted(d.glob("*.csv")):
            if "duplicate_details" not in csvf.name.lower():
                continue
            fc: Counter = Counter()
            rows = 0
            for enc in ("utf-8-sig", "cp1252"):
                try:
                    with open(csvf, encoding=enc, newline="") as fh:
                        rd = csv.reader(fh)
                        hdr2 = next(rd, [])
                        imo = _find(hdr2, "Month", "Flight YearMonth")
                        for r in rd:
                            rows += 1
                            v = (r[imo].strip() if imo is not None and imo < len(r)
                                 else "")
                            fc[v[:7] if len(v) >= 7 else (v or "(空)")] += 1
                    break
                except UnicodeDecodeError:
                    continue
            add("")
            add(f"  === 產出 {csvf.name} ===   {rows:,} 列 (只含 flagged)")
            if imo is None:
                add("    !! 這個 CSV 沒有月份欄, 無法按月比對")
                continue
            keys = sorted(set(ok) | {k for k in fc if len(k) == 7 and k[4] == "-"})
            add(f"    {'月份':<9}{'原始列數':>10}{'flagged':>10}")
            for k in keys:
                o, g = seen_all.get(k, 0), fc.get(k, 0)
                mark = ("   <- 有資料但沒有任何 flagged" if o and not g else
                        "   <- 原始資料就沒有這個月" if not o else "")
                add(f"    {k:<9}{o:>10,}{g:>10,}{mark}")
            odd = {k: v for k, v in fc.items() if k not in keys}
            for k, v in odd.items():
                add(f"    !! Details 裡月份異常: {k!r} × {v:,} 列")

    if not done:
        _no_hit(pats)


def check_samples(root: Path, data: Path, pats: list[str] | None, n: int) -> None:
    """從 Details CSV 抽幾組出來寫成檔案, 讓人用肉眼確認判斷對不對。

    ⚠️ 這是唯一會輸出票號的模式。票號結合姓名可識別到個人, 所以:
      - 明細寫進獨立的 dup_samples.txt, **不進** survey_*.txt
      - 主控台與報告只說「抽了幾組、寫到哪」
    分兩類各抽 n 組: 組內日期全同 (可能是同一次離境被記兩次) 與
    日期不同 (不同天各自從香港出發, 依規定該各收一次稅)。
    """
    outs = sorted((x for x in data.rglob("*.csv")
                   if "duplicate_details" in x.name.lower()
                   and (pats is None
                        or any(p.lower() in str(x).lower() for p in pats))),
                  key=lambda p: str(p).lower())
    add("")
    add("=" * 72)
    add(f"抽樣 —— 每份 Details 各抽 {n} 組同日 / {n} 組不同日")
    add("=" * 72)
    if not outs:
        _no_hit(pats)
        return

    lines: list[str] = [
        "# 重複組抽樣 —— 含票號, 屬個人資料, 不要外傳或貼進聊天視窗",
        "# 每一組是客戶第三條偵測判為「重複」的一組列。",
        "# 日期不同 = 不同天各自從香港出發, 依規定該各收一次稅。",
        "",
    ]
    for f in outs:
        # 第一趟只收日期, 記憶體可控 (Details 可能有幾十萬列)
        dates: dict = {}
        enc_ok = "utf-8-sig"
        for enc in ("utf-8-sig", "cp1252"):
            try:
                with open(f, encoding=enc, newline="") as fh:
                    rd = csv.reader(fh)
                    hdr = next(rd, [])
                    ik = _find(hdr, "Duplicate_Set_Key", "TKT_SET_NUM",
                               "Duplicate Set Key")
                    ifd = _find(hdr, "FLIGHT_DATE", "Flight Date")
                    if ik is None or ifd is None:
                        break
                    for r in rd:
                        k = r[ik].strip() if ik < len(r) else ""
                        if k:
                            dates.setdefault(k, set()).add(
                                day_of(r[ifd]) if ifd < len(r) else "")
                enc_ok = enc
                break
            except UnicodeDecodeError:
                continue
        if not dates:
            add(f"  {f.name}: 沒有 Duplicate_Set_Key 或 FLIGHT_DATE 欄, 跳過")
            continue

        same = [k for k, v in dates.items() if len(v) == 1][:n]
        diff = [k for k, v in dates.items() if len(v) > 1][:n]
        want = set(same) | set(diff)
        del dates

        rows: dict = {}
        with open(f, encoding=enc_ok, newline="") as fh:
            rd = csv.reader(fh)
            hdr = next(rd, [])
            ik = _find(hdr, "Duplicate_Set_Key", "TKT_SET_NUM", "Duplicate Set Key")
            cols = [(_find(hdr, *c), c[0]) for c in
                    (("Source_File", "Source File"), ("TKT_NUM",),
                     ("COUPON_NUMBER",), ("FLIGHT_NO",), ("FLIGHT_DATE",),
                     ("Total Amount(HKD)", "Total Amount HKD"),
                     ("Proposed Adjustment",))]
            for r in rd:
                k = r[ik].strip() if ik < len(r) else ""
                if k in want:
                    rows.setdefault(k, []).append(
                        [(r[i].strip() if i is not None and i < len(r) else "")
                         for i, _ in cols])

        lines.append("=" * 78)
        lines.append(f"{f.relative_to(data).as_posix()}")
        lines.append("=" * 78)
        for lab, keys in (("組內日期全同 (可能是同一次離境被記兩次)", same),
                          ("組內日期不同 (不同天各自離境, 應各收一次稅)", diff)):
            lines.append("")
            lines.append(f"--- {lab} —— 抽 {len(keys)} 組 ---")
            for k in keys:
                lines.append("")
                lines.append(f"  Duplicate_Set_Key = {k}   ({len(rows.get(k, []))} 列)")
                lines.append("    " + " | ".join(c for _, c in cols))
                for rec in rows.get(k, []):
                    lines.append("    " + " | ".join(rec))
        add(f"  {f.relative_to(data).as_posix()}: 抽了 "
            f"{len(same)} 組同日 + {len(diff)} 組不同日")

    out = root / "dup_samples.txt"
    out.write_text("\n".join(lines), encoding="utf-8")
    add("")
    add(f"  明細已寫到 {out.name} (含票號, 不要外傳)")


def check_pair(root: Path, data: Path, pats: list[str] | None) -> None:
    """比對兩個**檔案**的票號交集, 並看那些共同票號是不是同一次離境。

    用途: 2022 年新舊系統並行, 舊系統那個檔被放進來就是為了查
    「同一張票有沒有在兩套系統各收一次稅」。客戶的三條偵測是資料夾內
    全域的, 理論上抓得到 —— 但第三條需要 Rate Range ∈ {A,B,C},
    而新版 Conflict 3 把舊系統那個檔的金額歸零了, 等於這個目的失效。
    這支不依賴客戶的產出, 直接從原始資料回答。

    判準跟跨資料夾那段一致:
      完全相同的 (航班, 日曆日) = 同一次離境被記兩次 = 真的重複收費
      同一日曆日、不同航班     = 當日接駁
      不同日曆日               = 各自獨立的離境, 本來就該各收一次
    """
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    files = sorted((f for f in data.rglob("*.xlsx")
                    if not f.name.startswith("~$")
                    and "duplicate_analysis" not in f.name.lower()
                    and (pats is None
                         or any(p.lower() in f.name.lower() for p in pats))),
                   key=lambda p: str(p).lower())
    add("")
    add("=" * 72)
    add("兩個檔之間的票號交集")
    add("=" * 72)
    if len(files) != 2:
        add(f"  !! --files 要正好篩出 2 個檔, 現在是 {len(files)} 個:")
        for f in files[:12]:
            add(f"     {f.name}")
        return

    seen: dict = {}          # 票號 -> {檔名: set of (航班, 日曆日)}
    for f in files:
        t0 = time.time()
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = rn([txt(v) for v in next(it, ())])
        ix = {c: i for i, c in enumerate(hdr)}
        it_, ifn, ifd = ix.get("TKT_NUM"), ix.get("FLIGHT_NO"), ix.get("FLIGHT_DATE")
        n = 0
        for r in it:
            if r is None or all(v is None for v in r):
                continue
            n += 1
            k = tkt_key(txt(r[it_])) if it_ is not None and it_ < len(r) else None
            if k is None:
                continue
            fno = txt(r[ifn]) if ifn is not None and ifn < len(r) else ""
            dy = day_of(r[ifd]) if ifd is not None and ifd < len(r) else ""
            seen.setdefault(k, {}).setdefault(f.name, set()).add((fno, dy))
        wb.close()
        add(f"  掃完 {f.name}   {n:,} 列   ({time.time() - t0:.0f}s)")

    a, b = files[0].name, files[1].name
    both = exact = sameday = diff = 0
    for k, d2 in seen.items():
        if len(d2) < 2:
            continue
        both += 1
        sa, sb = d2[a], d2[b]
        if sa & sb:
            exact += 1
        elif {x[1] for x in sa} & {x[1] for x in sb}:
            sameday += 1
        else:
            diff += 1
    add("")
    add(f"  {a}")
    add(f"    ∩ {b}")
    add(f"    兩邊都有紀錄的票號   {both:,}")
    if both:
        add(f"      完全相同的 (航班, 日期) {exact:>9,} ({exact/both:.1%})"
            "   <- 同一次離境被記兩次 = 真的重複收費")
        add(f"      同一日曆日、不同航班    {sameday:>9,} ({sameday/both:.1%})"
            "   <- 當日接駁")
        add(f"      完全不同的日期          {diff:>9,} ({diff/both:.1%})"
            "   <- 各自獨立的離境")
    else:
        add("    -> 沒有共同票號, 這兩個檔之間不存在重複收費。")


def check_cols(root: Path, data: Path, pats: list[str] | None) -> None:
    """只讀表頭, 印出每個檔的完整欄名。秒級, 不讀任何資料列。

    --intake 回答的是「這個欄有沒有」, 這支回答「它到底叫什麼」——
    當某個 required 欄被判定為缺, 要先看清楚該檔實際的欄名長怎樣,
    才知道是「真的沒有」還是「換了個名字而字典沒收」。
    """
    mapping = load_client_mapping(root)
    add("")
    add("=" * 72)
    add("欄名清單 (只讀表頭)")
    add("=" * 72)
    done = 0

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")
                      and "duplicate_analysis" not in f.name.lower()},
                     key=lambda d: str(d).lower())
    for d in folders:
        rel = d.relative_to(data).as_posix()
        low = rel.lower()
        files = [f for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower())
                 if not f.name.startswith("~$")
                 and "duplicate_analysis" not in f.name.lower()
                 and (pats is None
                      or any(p.lower() in f.name.lower() or p.lower() in low
                             for p in pats))]
        if not files:
            continue
        done += 1
        add("")
        add("-" * 72)
        add(f"[{rel}]")
        add("-" * 72)
        for f in files:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            hdr = [txt(v) for v in next(ws.iter_rows(values_only=True), ())]
            wb.close()
            while hdr and not hdr[-1]:
                hdr.pop()
            # 三分類, 不是兩分類 —— 客戶的 df.rename(columns=...) 是**精確比對**,
            # 而這支勘查工具為了認得多一點, 比對前會先 .upper()。
            # 兩者的差就是「字典寫了但客戶那邊不會生效」的欄, 那是靜默失效:
            # 沒有警告、沒有例外, 只是月份整欄空掉。單獨列出來。
            exact, upper_only, miss = [], [], []
            for c in hdr:
                k = c.strip()
                if k in mapping:
                    exact.append(f"{k} -> {mapping[k]}")
                elif k.upper() in mapping:
                    upper_only.append(f"{k} -> {mapping[k.upper()]}  "
                                      f"(字典的 key 是 {k.upper()!r})")
                else:
                    miss.append(k)
            add("")
            add(f"  --- {f.name} ---   {len(hdr)} 欄")
            add(f"    字典精確認得的 ({len(exact)}) —— 客戶的 rename 會生效:")
            for x in exact:
                add(f"      {x}")
            if upper_only:
                add(f"    !! 只有轉大寫才對得上 ({len(upper_only)}) —— "
                    f"**客戶的 rename 不會生效**:")
                for x in upper_only:
                    add(f"      {x}")
                add("       pandas 的 rename 是精確比對, 字典的 key 要跟欄名一字不差。")
            # rename 之後會不會有兩個欄叫同一個名字 —— 客戶四支都沒有
            # columns.duplicated() 處理, 撞名的話 combined["X"] 會回傳
            # DataFrame 而不是 Series, to_number() 裡的 .str 直接
            # AttributeError, 整支腳本崩潰。用精確比對算 (客戶的行為)。
            after: Counter = Counter()
            for c in hdr:
                k = c.strip()
                if k:
                    after[mapping.get(k, k)] += 1
            dup = {k: v for k, v in after.items() if v > 1}
            if dup:
                add("    !! rename 之後會出現重複欄名 —— 客戶的腳本會崩潰:")
                for k, v in sorted(dup.items()):
                    src = [c.strip() for c in hdr
                           if mapping.get(c.strip(), c.strip()) == k]
                    add(f"      {k!r} <- {v} 個欄: {', '.join(src)}")
            add(f"    字典沒收的, 原樣保留 ({len(miss)}):")
            # 一行塞幾個, 66 欄逐行印太佔版面
            for i in range(0, len(miss), 3):
                add("      " + " | ".join(f"{c[:30]:<30}" for c in miss[i:i + 3]).rstrip())

    if not done:
        _no_hit(pats)


# --------------------------------------------------- 新資料進來先過這一遍
INTAKE_REQ = ["TKT_NUM", "FLIGHT_NO", "FLIGHT_DATE", "TKT_SET_NUM",
              "PSR_CURRENCY", "PSR AMT", "ORIGINAL PSR CURR", "ORIGINAL PSR AMT"]
INTAKE_SAMPLE = 200000     # 檔案按月排序, 取樣太少會只看到一月


def check_intake(root: Path, data: Path, pats: list[str] | None) -> None:
    """客戶給新資料就跑這個 —— 一個檔一個區塊, 回答三件事:

      1. 客戶的 required_columns 缺哪幾個 (缺 -> 被補成空 -> 金額算 0)
      2. 關鍵的選用欄在不在 (COUPON / Collected Tax / 月份 / 日差)
      3. 金額欄實際有沒有值 —— 尤其 ORIGINAL PSR AMT:
         Total Amount(HKD) = PSR Amount HKD + Original Amount HKD, 缺的欄會被
         算成 0。所以「新檔缺 ORIGINAL PSR AMT」到底有沒有少算, 要看**舊檔**
         那一欄是不是本來就多半是 0。這裡量的就是它。

    欄位盤點只讀表頭 (秒級), 只有金額比例需要讀列 (取樣 20 萬)。
    """
    mapping = load_client_mapping(root)

    def rn(h):
        return [mapping.get(c.strip().upper(), c.strip()) for c in h]

    folders = sorted({f.parent for f in data.rglob("*.xlsx")
                      if not f.name.startswith("~$")
                      and "duplicate_analysis" not in f.name.lower()},
                     key=lambda d: str(d).lower())

    add("")
    add("=" * 72)
    add("新資料盤點 —— 缺哪些欄 / 金額欄有沒有值")
    add("=" * 72)
    done = 0

    for d in folders:
        rel = d.relative_to(data).as_posix()
        # 大小寫不敏感 —— 實際的資料夾名稱之間有只差大小寫的情況,
        # 精確比對會靜靜漏掉一整個資料夾, 而且不會有任何警告。
        low = rel.lower()
        files = [f for f in sorted(d.glob("*.xlsx"), key=lambda x: x.name.lower())
                 if not f.name.startswith("~$")
                 and "duplicate_analysis" not in f.name.lower()
                 and (pats is None
                      or any(p.lower() in f.name.lower() or p.lower() in low
                             for p in pats))]
        if not files:
            continue
        done += 1
        add("")
        add("-" * 72)
        add(f"[{rel}]   {len(files)} 個檔")
        add("-" * 72)

        for f in files:
            t0 = time.time()
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            raw_hdr = [txt(v) for v in next(it, ())]
            hdr = rn(raw_hdr)
            ix = {c: i for i, c in enumerate(hdr)}
            rawset = {c.strip().lower() for c in raw_hdr}

            miss = [c for c in INTAKE_REQ if c not in ix]
            gap = [c for c in GAPCOLS if c in ix]
            ipa, ioa = ix.get("PSR AMT"), ix.get("ORIGINAL PSR AMT")
            ict = ix.get("Collected Tax")
            it_ = ix.get("TKT_NUM")
            n = 0
            pa = [0, 0]          # PSR AMT: [非零, 有值]
            oa = [0, 0]          # ORIGINAL PSR AMT
            ct = [0, 0]          # Collected Tax —— 新版 Conflict 3 只看「欄在不在」
                                 # 就拿它覆蓋 Total Amount, 沒看有沒有值。
                                 # 欄在但全空的檔, 金額會被覆蓋成 0。
            lens: Counter = Counter()
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                n += 1
                # 票號長度全掃 —— 這個統計很便宜, 而且「哪個檔有非 13 碼、
                # 有幾列」要精確, 取樣會低估 (檔案按月排序, 20 萬列只看得到前幾個月)。
                if it_ is not None and it_ < len(r):
                    t = txt(r[it_])
                    if t:
                        lens[len(t)] += 1
                if n > INTAKE_SAMPLE:
                    continue
                for i, acc in ((ipa, pa), (ioa, oa), (ict, ct)):
                    if i is None or i >= len(r):
                        continue
                    x = num(r[i])
                    if x is not None:
                        acc[1] += 1
                        if x != 0:
                            acc[0] += 1
            wb.close()

            shown = min(n, INTAKE_SAMPLE)
            add("")
            add(f"  --- {f.name} ---   {n:,} 列   ({len(hdr)} 欄)")
            add(f"    缺 required: " + (", ".join(miss) if miss else "無, 八個都在"))
            opt = []
            opt.append("COUPON_NUMBER " + ("O" if "COUPON_NUMBER" in ix else "X"))
            opt.append("Flight YearMonth " + ("O" if "Flight YearMonth" in ix else "X"))
            opt.append("Collected Tax " + ("O" if "collected tax" in rawset else "X"))
            opt.append("日差欄 " + (" + ".join(gap) if gap else "X"))
            add("    選用欄: " + " | ".join(opt))
            if "PSR AMT" in miss and "collected tax" in rawset:
                add("    -> 缺 PSR AMT 但有 Collected Tax, 新版 Conflict 3 會直接"
                    "拿它當 Total Amount(HKD)")
            if lens:
                add("    票號長度 (全掃): "
                    + ", ".join(f"{L}碼×{c:,}" + ("" if L == 13 else " <-非13碼")
                                for L, c in sorted(lens.items())))
                short = sum(c for L, c in lens.items() if L != 13)
                if short:
                    add(f"    !! 這個檔有 {short:,} 列票號不是 13 碼 ({short/n:.1%})")
            for name, acc, i in (("PSR AMT", pa, ipa),
                                 ("ORIGINAL PSR AMT", oa, ioa),
                                 ("Collected Tax", ct, ict)):
                if i is None:
                    add(f"    {name:<18} 沒有這一欄 -> 被補成空 -> 算 0")
                elif not acc[1]:
                    add(f"    {name:<18} 欄在, 但取樣 {shown:,} 列**全是空的**"
                        + ("   !! Conflict 3 會拿它覆蓋 Total Amount -> 全部變 0"
                           if name == "Collected Tax" else ""))
                else:
                    add(f"    {name:<18} 有值 {acc[1]:,}/{shown:,} "
                        f"({acc[1]/shown:.1%})   其中非零 {acc[0]:,} "
                        f"({acc[0]/acc[1]:.1%})")
            add(f"    ({time.time() - t0:.0f}s)")

    if not done:
        _no_hit(pats)
        return

    add("")
    add("判讀 ORIGINAL PSR AMT: 若舊檔那一欄本來就多半是 0 或全空,")
    add("則新檔缺這一欄**不會少算金額** (Total = PSR + Original, Original 本來就 0);")
    add("若舊檔非零比例高, 新檔就是真的少算了一塊, 那才非要客戶補不可。")


def check_outputs(data: Path) -> None:
    add("")
    add("=" * 72)
    add("客戶成果檔勘查 (只讀產出, 不碰原始資料)")
    add("=" * 72)
    outs = sorted((f for f in data.rglob("*.xlsx")
                   if "duplicate_analysis" in f.name.lower()
                   and not f.name.startswith("~$")),
                  key=lambda p: str(p).lower())
    if not outs:
        add("  找不到任何 Duplicate_Analysis.xlsx —— 客戶的腳本還沒跑?")
    for f in outs:
        add("")
        add("-" * 72)
        add(f"{f.relative_to(data).as_posix()}   {f.stat().st_size / 1e6:,.1f}MB")
        add("-" * 72)
        wb = load_workbook(f, read_only=True, data_only=True)
        add(f"  工作表: {', '.join(wb.sheetnames)}")

        # ---- Summary Pivot: 逐檔的列數與 flagged 數 ----
        for nm in wb.sheetnames:
            if "summary pivot" not in nm.lower():
                continue
            ws = wb[nm]
            rows = [[txt(v) for v in r] for r in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            add(f"\n  [{nm}]")
            for r in rows[:20]:
                add("    " + " | ".join(x for x in r if x))
            if len(rows) > 20:
                add(f"    ...另外 {len(rows) - 20} 列")

        # ---- Validation Issues: 依類型計數 ----
        for nm in wb.sheetnames:
            if "validation" not in nm.lower():
                continue
            ws = wb[nm]
            it = ws.iter_rows(values_only=True)
            hdr = [txt(v) for v in next(it, ())]
            ity = _find(hdr, "Issue_Type", "Issue Type")
            idt = _find(hdr, "Details")
            isf = _find(hdr, "Source_File", "Source File")
            byt: Counter = Counter()
            rows: list[tuple] = []
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                t = (txt(r[ity]) if ity is not None and ity < len(r) else "") or "(空)"
                dt = txt(r[idt]) if idt is not None and idt < len(r) else ""
                sf = txt(r[isf]) if isf is not None and isf < len(r) else ""
                byt[t] += 1
                rows.append((t, dt, sf))
            add(f"\n  [{nm}]  共 {len(rows):,} 筆")
            for k, v in byt.most_common():
                add(f"    {k:<28} {v:>6,}")
            # 這張表是拿去跟客戶要資料的依據, 所以逐筆印, 不做 top-N 截斷
            for t in sorted(byt):
                add(f"\n    --- {t} ({byt[t]:,} 筆) ---")
                for _, dt, sf in sorted(x for x in rows if x[0] == t):
                    add(f"      {dt:<34} {sf}")

        # ---- Rate Range Summary: 這是驗 C1/C2 的關鍵 ----
        for nm in wb.sheetnames:
            if "rate range" not in nm.lower():
                continue
            ws = wb[nm]
            it = ws.iter_rows(values_only=True)
            hdr = [txt(v) for v in next(it, ())]
            irr = _find(hdr, "Rate Range")
            imo = _find(hdr, "Flight YearMonth", "Month")
            irc = _find(hdr, "Row_Count", "Row Count")
            iamt = _find(hdr, "Total_Amount_HKD", "Total Amount(HKD)", "Total Amount HKD")
            by_rr: Counter = Counter()
            amt_rr: dict = {}
            by_year: dict = {}
            for r in it:
                if r is None or all(v is None for v in r):
                    continue
                rr = txt(r[irr]) if irr is not None and irr < len(r) else ""
                rr = rr or "(空白)"
                n = num(r[irc]) if irc is not None and irc < len(r) else None
                a = num(r[iamt]) if iamt is not None and iamt < len(r) else None
                mo = txt(r[imo]) if imo is not None and imo < len(r) else ""
                by_rr[rr] += int(n or 0)
                amt_rr[rr] = amt_rr.get(rr, 0.0) + (a or 0.0)
                y = mo[:4] if len(mo) >= 4 else "?"
                d = by_year.setdefault(y, {})
                d[rr] = d.get(rr, 0) + int(n or 0)
            add(f"\n  [{nm}]  依級距彙總")
            for rr in sorted(by_rr, key=lambda k: -by_rr[k]):
                add(f"    {rr:<12} {by_rr[rr]:>10,} 列   金額 {amt_rr.get(rr, 0):>18,.2f}")
            miss = by_rr.get("Missing FX", 0)
            add(f"    -> C1 檢查: Missing FX "
                + (f"!! 還有 {miss:,} 列" if miss else "0 列 (正確)"))
            add(f"\n  依年份 × 級距 (驗 C2: 2023-2025 該有 A/B/C 而不是空白)")
            rrs = sorted(by_rr)
            add("    年份  " + "".join(f"{x:>14}" for x in rrs))
            for y in sorted(by_year):
                add(f"    {y:<6}" + "".join(f"{by_year[y].get(x, 0):>14,}" for x in rrs))
        wb.close()

    # ---- Details CSV: Proposed Adjustment 分佈 ----
    add("")
    add("=" * 72)
    add("Proposed Adjustment 分佈 (這就是要交付的判斷結果)")
    add("=" * 72)
    for f in sorted((x for x in data.rglob("*.csv")
                     if "duplicate_details" in x.name.lower()),
                    key=lambda p: str(p).lower()):
        cnt: Counter = Counter()
        by_file: dict = {}
        n = 0
        t0 = time.time()
        for enc in ("utf-8-sig", "cp1252"):
            try:
                with open(f, encoding=enc, newline="") as fh:
                    rd = csv.reader(fh)
                    hdr = next(rd, [])
                    ipa = _find([h for h in hdr], "Proposed Adjustment")
                    isf = _find([h for h in hdr], "Source_File", "Source File")
                    # Adjust 的金額總和 = 這套方法認定的「多收了多少」,
                    # 拿去跟毛差額 (實收 − 應申報) 比, 才知道重複收費
                    # 到底解釋了差額的幾成。
                    iam = _find([h for h in hdr], "Total Amount(HKD)",
                                "Total Amount HKD", "Total_Amount_HKD",
                                "Total Amount")
                    for r in rd:
                        n += 1
                        v = (r[ipa].strip() if ipa is not None and ipa < len(r)
                             else "") or "(空白)"
                        cnt[v] += 1
                        sf = (r[isf].strip() if isf is not None and isf < len(r)
                              else "") or "(空白)"
                        e = by_file.setdefault(sf, [Counter(), {}])
                        e[0][v] += 1
                        a = num(r[iam]) if iam is not None and iam < len(r) else None
                        if a:
                            k = ("Adjust" if v.startswith("Adjust") else
                                 "Keep" if v.startswith("Keep") else "TBC")
                            e[1][k] = e[1].get(k, 0.0) + a
                break
            except UnicodeDecodeError:
                continue
        # 每個重複組的列是來自同一個檔, 還是橫跨多個檔? 第三條偵測是
        # TKT_SET_NUM + Rate Range 分組, 分組是**全域**的 —— 把兩個資料夾
        # 合起來跑, 原本各自只有 1 列的組會併成 2 列而被判成重複。
        # 這裡把「同檔內」與「跨檔」的重複拆開, 才知道 flagged 增加是
        # 真的抓到新東西, 還是只是分組母體變大。
        grp: dict = {}
        for enc in ("utf-8-sig", "cp1252"):
            try:
                with open(f, encoding=enc, newline="") as fh:
                    rd = csv.reader(fh)
                    h2 = next(rd, [])
                    ik = _find(h2, "Duplicate_Set_Key", "TKT_SET_NUM",
                               "Duplicate Set Key")
                    isf2 = _find(h2, "Source_File", "Source File")
                    ifd2 = _find(h2, "FLIGHT_DATE", "Flight Date")
                    if ik is not None and isf2 is not None:
                        for r in rd:
                            k = r[ik].strip() if ik < len(r) else ""
                            if not k:
                                continue
                            sf = r[isf2].strip() if isf2 < len(r) else ""
                            e = grp.setdefault(k, [0, set(), set()])
                            e[0] += 1
                            e[1].add(sf)
                            if ifd2 is not None and ifd2 < len(r):
                                e[2].add(day_of(r[ifd2]))
                break
            except UnicodeDecodeError:
                continue

        add("")
        add(f"  {f.relative_to(data).as_posix()}   {n:,} 列  ({time.time() - t0:.0f}s)")
        for k, v in cnt.most_common():
            add(f"    {k:<24} {v:>9,} ({v/n:>6.1%})" if n else f"    {k}")
        # 逐檔 —— 總計會被大檔蓋過去, 而新舊版的檔案範圍不同, 只有逐檔才比得了。
        # Adjust 才是實質結論: flagged 只代表「被抓出來看過」, Keep 等於維持原judgement。
        if len(by_file) > 1:
            add(f"      {'依 Source_File':<40} {'列數':>9} {'Keep':>9} "
                f"{'Adjust':>9} {'TBC':>6}   {'Adjust 金額(HKD)':>18}")
            tot_adj_amt = 0.0
            for sf in sorted(by_file):
                c, amt = by_file[sf]
                tot = sum(c.values())
                keep = sum(v for k, v in c.items() if k.startswith("Keep"))
                adj = sum(v for k, v in c.items() if k.startswith("Adjust"))
                tbc = sum(v for k, v in c.items() if k.startswith("TBC"))
                aa = amt.get("Adjust", 0.0)
                tot_adj_amt += aa
                add(f"      {sf[:40]:<40} {tot:>9,} {keep:>9,} "
                    f"{adj:>9,} {tbc:>6,}   {aa:>18,.2f}")
            add(f"      {'合計 Adjust 金額':<40} {'':>9} {'':>9} {'':>9} "
                f"{'':>6}   {tot_adj_amt:>18,.2f}")
            add("      ^ 這就是這套方法認定的「多收金額」, 拿去跟毛差額比對")
        if grp:
            one = [v for v in grp.values() if len(v[1]) == 1]
            many = [v for v in grp.values() if len(v[1]) > 1]
            add("")
            add("    重複組的來源 (第三條是全域分組, 母體越大組就越容易湊滿):")
            add(f"      同一個 Source_File 內 {len(one):>9,} 組"
                f"{sum(v[0] for v in one):>12,} 列")
            add(f"      橫跨 2 個以上檔案     {len(many):>9,} 組"
                f"{sum(v[0] for v in many):>12,} 列   <- 只有合併跑才會出現")
            # 決定性的一問: 被歸成同一組的列, 航班日期一樣嗎?
            # 一樣 = 可能真的是同一次離境被收兩次;
            # 不一樣 = 不同天各自從香港出發, 本來就該各收一次稅。
            if any(v[2] for v in grp.values()):
                for lab, sub in (("同檔內", one), ("跨檔案", many)):
                    same = [v for v in sub if len(v[2]) == 1]
                    diff = [v for v in sub if len(v[2]) > 1]
                    if not sub:
                        continue
                    add(f"        {lab}: 組內日期全同 {len(same):>8,} 組"
                        f"{sum(v[0] for v in same):>11,} 列"
                        f"   /  日期不同 {len(diff):>8,} 組"
                        f"{sum(v[0] for v in diff):>11,} 列")
                add("        ^ 日期不同 = 不同天各自從香港出發, 依規定該各收一次稅")


def check_outliers(data: Path, thr: float) -> None:
    """稅額異常大的那些列, 它們的其他欄位有沒有規律?

    若是外幣, 兩個訊號會同時出現:
      a. Collected Tax / amount 的比值會集中在某幾個匯率上
      b. 那些列會依航線或出票辦事處成群
    若兩者都散亂, 才有理由懷疑是資料本身的問題。
    """
    add("")
    add("=" * 72)
    add(f"離群列勘查 —— 稅額 > {thr:,.0f} 的那些列")
    add("=" * 72)
    for frag, amt_c, ccy_c, alt_c in ARA_CHECK:
        f = next((x for x in data.rglob("*.xlsx")
                  if frag in x.name and not x.name.startswith("~$")), None)
        if f is None:
            continue
        t0 = time.time()
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = [txt(v) for v in next(it, ())]
        ix = {c: i for i, c in enumerate(hdr)}
        # 欄名大小寫在各年度不一致 (amount / Amount), 找不到就再比一次小寫
        low = {c.lower(): i for i, c in enumerate(hdr)}

        def col(name):
            return ix.get(name, low.get(name.lower()))

        ia, il = col(amt_c), col(alt_c)
        ctx = [(c, col(c)) for c in CTX if col(c) is not None]
        n_out = 0
        rows = 0
        ratios: list[float] = []
        vals: dict[str, Counter] = {c: Counter() for c, _ in ctx}
        alt_of_out = Counter()
        for r in it:
            if r is None or all(v is None for v in r):
                continue
            rows += 1
            a = num(r[ia]) if ia is not None and ia < len(r) else None
            if a is None or a <= thr:
                continue
            n_out += 1
            l = num(r[il]) if il is not None and il < len(r) else None
            if l:
                ratios.append(a / l)
                alt_of_out[f"{l:,.2f}"] += 1
            for c, i in ctx:
                if i < len(r):
                    vals[c][txt(r[i]) or "(空)"] += 1
        wb.close()
        add("")
        add(f"--- {f.relative_to(data).as_posix()} ---")
        add(f"  {rows:,} 列裡有 {n_out:,} 列 > {thr:,.0f}  ({n_out/rows:.4%})")
        if not n_out:
            add(f"  ({time.time() - t0:.0f}s)")
            continue
        if ratios:
            ratios.sort()
            add(f"  {amt_c} / {alt_c} 的比值:")
            add(f"    最小 {ratios[0]:,.1f}   中位 {ratios[len(ratios)//2]:,.1f}   "
                f"最大 {ratios[-1]:,.1f}")
            rc = Counter(round(x, 1) if x < 100 else round(x / 10) * 10 for x in ratios)
            add("    最常出現的比值: " +
                ", ".join(f"{k:,}×{v}" for k, v in rc.most_common(8)))
            add(f"    (不同的比值共 {len(rc)} 種 —— 若是外幣, 應該集中在少數幾個)")
            for rate, name in FXHINT:
                hit = sum(1 for x in ratios if abs(x - rate) / rate < 0.15)
                if hit:
                    add(f"    ~{rate} ({name}) 附近: {hit} 列")
        add(f"  這些列的 {alt_c}: " +
            ", ".join(f"{k}×{v}" for k, v in alt_of_out.most_common(5)))
        add("  情境欄的分佈 (值域小的才有意義):")
        for c, _ in ctx:
            cnt = vals[c]
            if not cnt:
                continue
            top = ", ".join(f"{k}×{v}" for k, v in cnt.most_common(5))
            flag = "  <- 高度集中" if cnt.most_common(1)[0][1] / n_out > 0.8 else ""
            add(f"    {c:<22} {len(cnt):>4} 種   {top}{flag}")
        add(f"  ({time.time() - t0:.0f}s)")


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=".")
    ap.add_argument("--files", default=None,
                    help="檔名片段, 逗號分隔 (預設: 新系統那幾個檔 + 空列異常那個)")
    ap.add_argument("--all", action="store_true", help="全部檔案 (慢)")
    ap.add_argument("--sample", type=int, default=SAMPLE)
    ap.add_argument("--ara", action="store_true",
                    help="全掃那幾個新系統檔, 驗證 C2 的假設")
    ap.add_argument("--outliers", nargs="?", type=float, const=OUTLIER, default=None,
                    help=f"只看稅額 > N 的那些列有沒有規律 (預設 {OUTLIER:,.0f})")
    ap.add_argument("--dig", action="store_true",
                    help="深入查 H2 跨資料夾 / H4 轉機豁免 / H6 票號長度")
    ap.add_argument("--tkt", action="store_true",
                    help="票號前導零: 補得回來嗎 (check digit) / 客戶漏掉多少")
    ap.add_argument("--intake", action="store_true",
                    help="客戶給新資料就跑這個: 缺哪些欄 / 金額欄有沒有值")
    ap.add_argument("--cols", action="store_true",
                    help="只讀表頭, 印出每個檔的完整欄名 (秒級)")
    ap.add_argument("--recon", action="store_true",
                    help="原始檔列數 vs 產出 Summary Pivot 的列數, 逐檔對帳")
    ap.add_argument("--months", action="store_true",
                    help="每個檔涵蓋哪些月份, 區間內缺了哪幾個月")
    ap.add_argument("--samples", nargs="?", type=int, const=20, default=None,
                    help="抽 N 組重複明細寫成 dup_samples.txt 供肉眼確認 (預設 20)")
    ap.add_argument("--pair", action="store_true",
                    help="比對兩個檔的票號交集 (--files 要正好篩出 2 個檔)")
    ap.add_argument("--outputs", action="store_true",
                    help="讀客戶跑出來的成果檔, 驗 C1/C2 是否生效")
    ap.add_argument("--holes", action="store_true",
                    help="查客戶的重複偵測有沒有漏抓/誤判 (H1-H6)")
    args = ap.parse_args()

    data = Path(args.root).resolve() / DATA_DIRNAME
    if not data.is_dir():
        sys.exit(f"找不到 {data}")
    # 檔案清單只有「逐檔勘查」那個模式才用得到 —— 其他模式自己決定要讀什麼,
    # 先篩選會讓它們在還沒開始之前就因為「沒有檔案符合」而結束。
    other_mode = (args.dig or args.outputs or args.holes or args.ara or args.tkt
                  or args.intake or args.cols or args.pair or args.months
                  or args.recon or args.samples is not None
                  or args.outliers is not None)
    files: list[Path] = []
    if not other_mode:
        pats = (None if args.all
                else [p.strip() for p in (args.files or ",".join(DEFAULT_FILES)).split(",")])
        files = sorted((f for f in data.rglob("*.xlsx") if not f.name.startswith("~$")),
                       key=lambda p: str(p).lower())
        if pats is not None:
            files = [f for f in files if any(p in f.name for p in pats)]
        if not files:
            sys.exit(f"沒有檔案符合 {pats}")

    t0 = time.time()
    mode = ("intake" if args.intake else "tkt" if args.tkt else "dig" if args.dig
            else "outputs" if args.outputs else "holes" if args.holes
            else "cols" if args.cols else "pair" if args.pair
            else "months" if args.months else "recon" if args.recon else "samples" if args.samples is not None
            else "ara" if args.ara
            else "outliers" if args.outliers is not None else "profile")
    add(f"勘查 {data}")
    if args.samples is not None:
        check_samples(Path(args.root).resolve(), data,
                      [x.strip() for x in args.files.split(",")] if args.files else None,
                      args.samples)
    elif args.recon:
        check_recon(data,
                    [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.months:
        check_months(Path(args.root).resolve(), data,
                     [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.pair:
        check_pair(Path(args.root).resolve(), data,
                   [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.cols:
        check_cols(Path(args.root).resolve(), data,
                   [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.intake:
        check_intake(Path(args.root).resolve(), data,
                     [p.strip() for p in args.files.split(",")] if args.files else None)
    elif args.tkt:
        check_tkt(Path(args.root).resolve(), data,
                  [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.dig:
        check_dig(Path(args.root).resolve(), data,
                  [x.strip() for x in args.files.split(",")] if args.files else None)
    elif args.outputs:
        check_outputs(data)
    elif args.holes:
        check_holes(Path(args.root).resolve(), data)
    elif args.outliers is not None:
        check_outliers(data, args.outliers)
    elif args.ara:
        check_ara(data)
    else:
        add(f"對象 {len(files)} 個檔: " + ", ".join(f.name for f in files))
        for f in files:
            profile(f, f.relative_to(data).as_posix(), args.sample)
    add("")
    add(f"完成, 共 {time.time() - t0:.0f}s")
    # 每個模式各寫各的檔 —— 以前全部寫 survey.txt, 跑第二個模式就把第一個的
    # 結果蓋掉了, 而那份輸出常常是唯一還留著的 baseline。
    out = Path(args.root).resolve() / REPORT.format(mode=mode)
    out.write_text("\n".join(L), encoding="utf-8")
    print(f"\n報告: {out}")
    print("(只有欄名與統計值; 個資欄連高頻值都不印, 見 survey.py 的 PII)")


if __name__ == "__main__":
    main()
