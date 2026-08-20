#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx 明細勘查 + parquet 轉檔工具

給「一個資料夾裡有一堆結構相同的月報 xlsx, 要合併分析」這種情境用。
在寫聚合邏輯之前先掃一遍, 確認:

  1. 模板檔的欄位結構 (合併儲存格 / 兩層表頭)
  2. 每個明細檔的表頭是否一致, 有沒有哪個檔改過欄位
  3. 每個檔實際包含哪些月份 -> 有沒有跨檔重複或中間缺月
  4. key 欄位的空值狀況
  5. 同一個代碼在不同月份是否對到不同的名稱 / 分類 (會讓分組裂開)
  6. 數值欄有沒有非數值 (例如 "-") 或負數
  7. 聚合後大概會有多少列 (--full 模式才算)

順便可以把明細轉成 parquet。轉檔本身要花一次完整解析 xlsx 的時間,
所以設計成跟勘查同一趟做完, 不要分兩次跑。之後的分析直接讀 parquet,
從幾十分鐘變成幾秒。

用法:
    python inspect_xlsx.py                     # 快掃, 每檔只讀前 3000 列
    python inspect_xlsx.py --full              # 全掃
    python inspect_xlsx.py --to-parquet        # 全掃 + 轉 parquet
    python inspect_xlsx.py --root /path/to/dir --files "2024-*.xlsx"

預期的目錄結構:
    <root>/
        模板.xlsx          <- root 底下第一個 xlsx 會被當成模板
        data/*.xlsx        <- 明細檔
        parquet/           <- --to-parquet 的輸出

要換成你自己的資料 schema, 改下面 EXPECTED_COLS / GROUP_COLS / MEASURE_COLS 就好。
注意欄位名稱要跟來源檔完全一致 (包括簡繁體), 否則比對不到。
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("請先安裝 openpyxl:  pip install openpyxl")


# ================================================================= 設定
# 改這一段以符合你的資料。名稱要跟來源 xlsx 的表頭一致。
#
# 表頭比對是寬鬆的: 先去掉所有空白, 再依序試「完全相同 -> 開頭相同 -> 包含」,
# 所以 "营业日期 (月)" 對得到 "营业日期", "实收金额（含税）" 對得到 "实收金额"。
EXPECTED_COLS = [
    "营业日期", "门店名称", "门店编码", "城市", "公司", "销售渠道",
    "商品编码", "产品分类", "商品名称", "系统商品名称", "规格",
    "商品数量", "应收金额", "实收金额", "优惠金额",
]
# 做「包含」比對時, 長的先比, 避免 "商品名称" 誤吃掉 "系统商品名称"
_CONTAINS_ORDER = sorted(EXPECTED_COLS, key=len, reverse=True)

# 分組 key: 檢查空值、算 distinct、估聚合後列數都用這組
GROUP_COLS = ["门店编码", "销售渠道", "产品分类", "商品编码", "商品名称"]
# 數值欄: 檢查非數值與負數
MEASURE_COLS = ["商品数量", "应收金额", "实收金额"]
# 月份欄: 用來判斷每個檔涵蓋哪些期間
DATE_COL = "营业日期"
# 代碼欄 -> 名稱欄 / 分類欄, 用來偵測「同一代碼對到多個名稱」
CODE_COL, NAME_COL, CATEGORY_COL = "商品编码", "商品名称", "产品分类"
# 所有要存成 float 的欄位 (MEASURE_COLS 之外還想帶走的數值欄寫在這)
NUMERIC_COLS = MEASURE_COLS + ["优惠金额"]
# ================================================================= 設定結束

# parquet schema。字串欄靠 dictionary encoding 壓縮, 數值欄一律 float64。
# period / date_raw / source_file 是衍生欄, 用英文以便跟來源欄位區分。
# source_file 是 provenance —— 對數對不上時可以直接回推是哪個來源檔的問題。
PQ_NUM = list(NUMERIC_COLS)
PQ_STR = (["period", "date_raw"]
          + [c for c in EXPECTED_COLS if c != DATE_COL and c not in PQ_NUM]
          + ["source_file"])


class ParquetSink:
    """把明細逐批寫成 parquet。分批 flush 讓記憶體維持固定。

    金額欄轉不動的值 (例如 '-') 寫 null 而不是 0 —— 0 跟「本來就沒值」
    在加總時分不出來, null 之後要當 0 隨時可以, 反過來救不回。
    """

    def __init__(self, path: Path, batch: int = 500_000):
        import pyarrow as pa
        import pyarrow.parquet as pq
        self.pa, self.path = pa, path
        self.schema = pa.schema([(c, pa.string()) for c in PQ_STR]
                                + [(c, pa.float64()) for c in PQ_NUM])
        try:
            self.writer = pq.ParquetWriter(path, self.schema, compression="zstd")
            self.codec = "zstd"
        except Exception:
            self.writer = pq.ParquetWriter(path, self.schema, compression="snappy")
            self.codec = "snappy"
        self.buf: dict[str, list] = {c: [] for c in PQ_STR + PQ_NUM}
        self.batch, self.n = batch, 0

    def add(self, row: dict) -> None:
        for c in PQ_STR:
            v = row.get(c)
            self.buf[c].append(v if v else None)
        for c in PQ_NUM:
            self.buf[c].append(row.get(c))
        self.n += 1
        if self.n % self.batch == 0:
            self.flush()

    def flush(self) -> None:
        if not self.buf[PQ_STR[0]]:
            return
        cols = {c: self.pa.array(v, type=self.schema.field(c).type)
                for c, v in self.buf.items()}
        self.writer.write_table(self.pa.table(cols, schema=self.schema))
        for v in self.buf.values():
            v.clear()

    def close(self) -> None:
        self.flush()
        self.writer.close()


def norm_header(v) -> str:
    """表頭正規化: 去掉所有空白 (含全形空白) 與換行。"""
    if v is None:
        return ""
    return "".join(str(v).split()).replace("\u3000", "")


def canonical(header: str) -> str | None:
    """把實際表頭對到預期欄位名。先精準比對, 再前綴, 最後包含。"""
    h = norm_header(header)
    if not h:
        return None
    for t in EXPECTED_COLS:
        if h == t:
            return t
    for t in _CONTAINS_ORDER:
        if h.startswith(t):
            return t
    for t in _CONTAINS_ORDER:
        if t in h:
            return t
    return None


def to_month(v) -> str | None:
    """把營業日期正規化成 'YYYY-MM'。支援字串與 datetime 兩種存法。"""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return f"{v.year:04d}-{v.month:02d}"
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("/", "-").replace("年", "-").replace("月", "")
    parts = s.split("-")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].strip().isdigit():
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    return f"RAW:{s}"  # 認不得的原樣回報, 方便人工看


def to_number(v):
    """回傳 (數值, 是否成功)。'-'、''、None 視為 0 但標記為非數值。"""
    if v is None:
        return 0.0, False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), True
    s = str(v).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if s in ("", "-", "－", "—", "/", "N/A", "null", "NULL"):
        return 0.0, False
    try:
        return float(s), True
    except ValueError:
        return 0.0, False


def is_period(s: str) -> bool:
    return len(s) == 7 and s[4] == "-" and s[:4].isdigit() and s[5:].isdigit()


def analyse_template_sheet(rows: list[list[str]]) -> dict:
    """從前幾列推出模板結構: 哪一列是表頭、key 欄有幾個、幾個度量區塊、各區塊月份。

    不假設有合併儲存格 —— 區塊標題可能只放在該區塊第一格。
    """
    if not rows:
        return {}
    # 表頭 = 非空格最多的那一列
    hdr_i = max(range(len(rows)), key=lambda i: sum(1 for c in rows[i] if c))
    hdr = rows[hdr_i]
    label_row = rows[hdr_i - 1] if hdr_i > 0 else []

    periods = [(j, c) for j, c in enumerate(hdr) if is_period(c)]
    keys = [c for j, c in enumerate(hdr) if c and not is_period(c) and j < (periods[0][0] if periods else len(hdr))]

    # 區塊起點 = 上一列有標籤的欄位; 沒有標籤就用月份序列重新開始的位置推
    starts = [j for j, c in enumerate(label_row) if c]
    if not starts and periods:
        starts, prev = [periods[0][0]], None
        for j, c in periods:
            if prev is not None and c <= prev:
                starts.append(j)
            prev = c

    blocks = []
    for bi, s in enumerate(starts):
        end = starts[bi + 1] if bi + 1 < len(starts) else len(hdr)
        ms = [c for j, c in periods if s <= j < end]
        blocks.append({
            "label": label_row[s] if s < len(label_row) else "",
            "col_start": s + 1, "col_end": s + len(ms),
            "n_periods": len(ms),
            "first": ms[0] if ms else None, "last": ms[-1] if ms else None,
            "periods": ms,
        })
    same = len({tuple(b["periods"]) for b in blocks}) == 1 if blocks else False
    return {"header_row": hdr_i + 1, "label_row": hdr_i if hdr_i > 0 else None,
            "key_cols": keys, "n_blocks": len(blocks), "blocks_same_periods": same,
            "blocks": blocks, "data_starts_row": hdr_i + 2}


# ---------------------------------------------------------------- 模板勘查
def inspect_template(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    out = {"file": path.name, "sheets": []}
    for ws in wb.worksheets:
        rows = []
        for r in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            rows.append([("" if c is None else str(c)) for c in r])
        out["sheets"].append({
            "name": ws.title,
            "structure": analyse_template_sheet(rows),
            "dims": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "freeze_panes": ws.freeze_panes,
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "first_rows": rows,
        })
    wb.close()
    return out


# ---------------------------------------------------------------- 明細勘查
def inspect_data_file(path: Path, limit: int | None, full: bool,
                      acc: dict | None = None, pq_dir: Path | None = None) -> dict:
    t0 = time.time()
    wb = load_workbook(path, read_only=True, data_only=True)
    res = {"file": path.name, "size_mb": round(path.stat().st_size / 1e6, 1), "sheets": []}

    # 一個來源 xlsx 對一個 parquet, 某個檔壞掉可以單獨重跑
    sink = None
    if pq_dir is not None:
        sink = ParquetSink(pq_dir / f"{path.stem}.parquet")

    for ws in wb.worksheets:
        info: dict = {"name": ws.title, "declared_max_row": ws.max_row, "declared_max_col": ws.max_column}

        it = ws.iter_rows(values_only=True)
        header_row_idx, headers = None, None
        probe = []
        for i, row in enumerate(it, start=1):
            probe.append(row)
            hits = sum(1 for c in row if canonical(c))
            if hits >= 5:
                header_row_idx, headers = i, row
                break
            if i >= 10:
                break

        if headers is None:
            info["error"] = "前 10 列找不到表頭"
            info["probe"] = [[str(c) for c in (r or [])][:20] for r in probe[:5]]
            res["sheets"].append(info)
            continue

        colmap: dict[str, int] = {}
        raw_headers = []
        for idx, h in enumerate(headers):
            raw_headers.append(norm_header(h))
            c = canonical(h)
            if c and c not in colmap:
                colmap[c] = idx

        info["header_row"] = header_row_idx
        info["headers"] = raw_headers
        info["missing_cols"] = [c for c in EXPECTED_COLS if c not in colmap]
        info["unmapped_headers"] = [h for h in raw_headers if h and canonical(h) is None]

        months = Counter()
        blanks = Counter()
        nonnumeric = Counter()
        negatives = Counter()
        distinct = {c: set() for c in GROUP_COLS}
        name_by_code = defaultdict(set)
        cat_by_code = defaultdict(set)
        keymonth_hashes: set[int] = set()
        keys_only: set[int] = set()
        n = 0

        for row in it:
            n += 1
            if limit and n > limit:
                break
            g = lambda c: row[colmap[c]] if c in colmap and colmap[c] < len(row) else None

            m = to_month(g(DATE_COL))
            mkey = m if m else "<空白>"   # 統計一律用字串 key, 避免 None 混進來不能排序
            months[mkey] += 1

            vals = {}
            for c in GROUP_COLS:
                v = g(c)
                sv = "" if v is None else str(v).strip()
                vals[c] = sv
                if not sv:
                    blanks[c] += 1
                elif len(distinct[c]) < 200_000:
                    distinct[c].add(sv)

            code = vals.get(CODE_COL, "")
            if code and len(name_by_code) < 100_000:
                name_by_code[code].add(vals.get(NAME_COL, ""))
                cat_by_code[code].add(vals.get(CATEGORY_COL, ""))
                if acc is not None and len(acc["name"]) < 200_000:
                    acc["name"][code].add(vals.get(NAME_COL, ""))
                    acc["cat"][code].add(vals.get(CATEGORY_COL, ""))
                    acc["keys"].add(tuple(vals[c] for c in GROUP_COLS))

            nums = {}
            for c in NUMERIC_COLS:
                num, ok = to_number(g(c))
                nums[c] = num if ok else None      # 轉不動就 null, 不要塞 0
                if c in MEASURE_COLS:
                    if not ok:
                        nonnumeric[c] += 1
                    elif num < 0:
                        negatives[c] += 1
                    if acc is not None and ok:
                        # 控制總數: 之後拿來核對聚合結果有沒有掉資料
                        acc["totals"][mkey][c] += num
            if acc is not None:
                acc["totals"][mkey]["_rows"] += 1
                # 空白 key 的影響有多大 —— 用金額佔比判斷能不能直接丟掉
                for c in GROUP_COLS:
                    if not vals[c]:
                        acc["blank_impact"][c]["rows"] += 1
                        for mc in MEASURE_COLS:
                            if nums.get(mc):
                                acc["blank_impact"][c][mc] += nums[mc]

            if sink is not None:
                raw_date = g(DATE_COL)
                rec = {
                    "period": m if (m and not m.startswith("RAW:")) else None,
                    "date_raw": "" if raw_date is None else str(raw_date).strip(),
                    "source_file": path.name,
                }
                for c in PQ_STR:
                    if c not in rec:
                        v = vals.get(c) if c in vals else g(c)
                        rec[c] = "" if v is None else str(v).strip()
                rec.update(nums)
                sink.add(rec)

            if full:
                k = hash(tuple(vals[c] for c in GROUP_COLS))
                keys_only.add(k)
                keymonth_hashes.add(hash((k, m)))

        info["rows_scanned"] = n if not (limit and n > limit) else limit
        info["scanned_all"] = not (limit and n > limit)
        info["months"] = dict(sorted(months.items()))
        info["blank_key_cols"] = {k: v for k, v in blanks.items() if v}
        info["nonnumeric_measure"] = dict(nonnumeric)
        info["negative_measure"] = dict(negatives)
        info["distinct_counts"] = {c: len(s) for c, s in distinct.items()}
        drift_name = {k: sorted(v) for k, v in name_by_code.items() if len(v) > 1}
        drift_cat = {k: sorted(v) for k, v in cat_by_code.items() if len(v) > 1}
        info["name_drift_count"] = len(drift_name)
        info["name_drift_sample"] = dict(list(drift_name.items())[:10])
        info["category_drift_count"] = len(drift_cat)
        info["category_drift_sample"] = dict(list(drift_cat.items())[:10])
        if full:
            info["unique_group_keys"] = len(keys_only)
            info["unique_key_month"] = len(keymonth_hashes)
        res["sheets"].append(info)

    wb.close()
    if sink is not None:
        sink.close()
        res["parquet"] = {"path": str(sink.path), "rows": sink.n, "codec": sink.codec,
                          "size_mb": round(sink.path.stat().st_size / 1e6, 1)}
    res["seconds"] = round(time.time() - t0, 1)
    return res


# ---------------------------------------------------------------- 報告輸出
def write_report(report: dict, out_dir: Path) -> None:
    (out_dir / "inspect_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    L = []
    add = L.append
    tpl = report.get("template")
    if tpl:
        add(f"=== 模板 {tpl['file']} ===")
        for sh in tpl["sheets"]:
            add(f"  sheet '{sh['name']}'  dims={sh['dims']}  max_row={sh['max_row']} max_col={sh['max_col']}")
            add(f"  合併儲存格 ({len(sh['merged_cells'])}): {sh['merged_cells'][:20]}")
            for i, r in enumerate(sh["first_rows"], 1):
                nonempty = [(j + 1, v) for j, v in enumerate(r) if v]
                add(f"    row{i} 非空 {len(nonempty)} 格, 前 12: {nonempty[:12]}")
            st = sh.get("structure") or {}
            if st:
                add(f"  -> 推得結構: 標題列={st['label_row']} 表頭列={st['header_row']} "
                    f"資料從第 {st['data_starts_row']} 列開始")
                add(f"     key 欄 ({len(st['key_cols'])}): {st['key_cols']}")
                add(f"     度量區塊 {st['n_blocks']} 個, 各區塊月份序列一致: {st['blocks_same_periods']}")
                for b in st["blocks"]:
                    add(f"       欄 {b['col_start']}-{b['col_end']}  {b['n_periods']} 個月 "
                        f"({b['first']} ~ {b['last']})  標題: {b['label']}")
                if st["blocks"]:
                    add(f"     完整月份序列: {st['blocks'][0]['periods']}")
        add("")

    add("=== 明細檔 ===")
    all_months = Counter()
    for f in report["data"]:
        add(f"\n--- {f['file']}  ({f['size_mb']} MB, {f['seconds']}s)")
        for sh in f["sheets"]:
            if "error" in sh:
                add(f"  sheet '{sh['name']}': !! {sh['error']}")
                # 印出前幾列, 才分得出「空 sheet」跟「格式不同所以漏抓資料」
                probe = sh.get("probe", [])
                if not probe:
                    add("      -> 完全沒有資料列, 是空 sheet, 可以忽略")
                for pi, pr in enumerate(probe, 1):
                    cells = [(j + 1, v) for j, v in enumerate(pr) if v]
                    add(f"      row{pi} 非空 {len(cells)} 格: {cells[:8]}")
                if probe:
                    add("      -> 有資料! 這個 sheet 會被跳過, 確認是否需要處理")
                continue
            add(f"  sheet '{sh['name']}': 表頭在第 {sh['header_row']} 列, "
                f"掃描 {sh['rows_scanned']} 列 (完整={sh['scanned_all']})")
            if sh["missing_cols"]:
                add(f"    !! 缺欄位: {sh['missing_cols']}")
            if sh["unmapped_headers"]:
                add(f"    ?  多出來的欄位: {sh['unmapped_headers']}")
            add(f"    月份: {sh['months']}")
            for m, c in sh["months"].items():
                all_months[m] += c
            if sh["blank_key_cols"]:
                add(f"    !! key 欄位空值: {sh['blank_key_cols']}")
            if sh["nonnumeric_measure"]:
                add(f"    ?  金額非數值筆數: {sh['nonnumeric_measure']}")
            if sh["negative_measure"]:
                add(f"    ?  負數筆數 (退貨?): {sh['negative_measure']}")
            add(f"    distinct: {sh['distinct_counts']}")
            if sh["name_drift_count"]:
                add(f"    !! 同編碼多商品名稱: {sh['name_drift_count']} 個, 例: "
                    f"{list(sh['name_drift_sample'].items())[:3]}")
            if sh["category_drift_count"]:
                add(f"    !! 同編碼多產品分類: {sh['category_drift_count']} 個, 例: "
                    f"{list(sh['category_drift_sample'].items())[:3]}")
            if "unique_group_keys" in sh:
                add(f"    聚合後列數(本檔): {sh['unique_group_keys']}, key×月 組合: {sh['unique_key_month']}")

    add("\n=== 全部檔案月份彙總 ===")
    for m in sorted(all_months):
        add(f"  {m}: {all_months[m]:,} 列")
    valid = sorted(m for m in all_months if len(m) == 7 and m[:4].isdigit())
    add(f"  共 {len(all_months)} 個月份")
    if valid:
        add(f"  範圍: {valid[0]} ~ {valid[-1]}")
        # 期間內中斷的月份 -> 通常代表來源檔漏給了
        y, mm = int(valid[0][:4]), int(valid[0][5:])
        want = []
        while f"{y:04d}-{mm:02d}" <= valid[-1]:
            want.append(f"{y:04d}-{mm:02d}")
            y, mm = (y + 1, 1) if mm == 12 else (y, mm + 1)
        gaps = [m for m in want if m not in all_months]
        add(f"  連續期間應有 {len(want)} 個月, 缺少: {gaps if gaps else '無'}")
    bad = [m for m in all_months if m not in valid]
    if bad:
        add(f"  !! 無法解析的日期值: {bad}")

    # 同一月份出現在多個檔案 -> 會重複計算, 必須處理
    month_files = defaultdict(list)
    for f in report["data"]:
        for sh in f["sheets"]:
            for m in sh.get("months", {}):
                month_files[m].append(f["file"])
    overlap = {m: fs for m, fs in month_files.items() if len(set(fs)) > 1}
    add(f"\n=== 月份重複出現在多個檔案 (會重複計算!) ===\n  {overlap if overlap else '無'}")

    g = report.get("global")
    if g:
        add("\n=== 跨檔案一致性 ===")
        add(f"  同一 {CODE_COL} 對到多個 {NAME_COL}: {g['name_drift_count']} 個 "
            f"(其中只是空白 vs 有值: {g['name_drift_blank_only']}, 真的改名: {g['name_drift_real']})")
        for k, v in list(g["name_drift_sample"].items())[:15]:
            add(f"    {k}: {v}")
        add(f"  同一 {CODE_COL} 對到多個 {CATEGORY_COL}: {g['category_drift_count']} 個 "
            f"(空白: {g['category_drift_blank_only']}, 真的改: {g['category_drift_real']})")
        for k, v in list(g["category_drift_sample"].items())[:15]:
            add(f"    {k}: {v}")
        add(f"  全部檔案合計 unique 分組 key (= 輸出列數估計): {g['unique_group_keys']:,}")

        bi = g.get("blank_impact") or {}
        if bi:
            add("\n=== 空白 key 的影響 (決定要丟掉還是保留成一類) ===")
            for c, d in bi.items():
                parts = "  ".join(f"{mc}={d.get(mc, 0):,.0f}" for mc in MEASURE_COLS)
                add(f"  {c}: {int(d['rows']):,} 列   {parts}")

        tot = g.get("totals") or {}
        if tot:
            add("\n=== 各月控制總數 (拿來核對聚合結果, 不要弄丟) ===")
            add(f"  {'期間':<8}{'列數':>10}" + "".join(f"{mc:>18}" for mc in MEASURE_COLS))
            gt = {mc: 0.0 for mc in MEASURE_COLS}
            grows = 0
            for m in sorted(tot):
                d = tot[m]
                grows += int(d.get("_rows", 0))
                for mc in MEASURE_COLS:
                    gt[mc] += d.get(mc, 0.0)
                add(f"  {m:<8}{int(d.get('_rows', 0)):>10,}"
                    + "".join(f"{d.get(mc, 0.0):>18,.2f}" for mc in MEASURE_COLS))
            add(f"  {'合計':<8}{grows:>10,}"
                + "".join(f"{gt[mc]:>18,.2f}" for mc in MEASURE_COLS))

    pqs = [f["parquet"] for f in report["data"] if "parquet" in f]
    if pqs:
        src = sum(f["size_mb"] for f in report["data"] if "parquet" in f)
        dst = sum(p["size_mb"] for p in pqs)
        add(f"\n=== parquet 轉檔 ({pqs[0]['codec']}) ===")
        add(f"  {len(pqs)} 檔, 共 {sum(p['rows'] for p in pqs):,} 列")
        add(f"  xlsx {src:,.0f} MB -> parquet {dst:,.0f} MB "
            f"(壓到 {dst / src * 100:.0f}%)" if src else "")

    txt = "\n".join(L)
    (out_dir / "inspect_report.txt").write_text(txt, encoding="utf-8")
    print(txt)


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=str(Path(__file__).resolve().parent),
                    help="資料夾 (內含 data/ 子目錄與模板 xlsx), 預設為腳本所在目錄")
    ap.add_argument("--full", action="store_true", help="完整掃描所有列 (慢)")
    ap.add_argument("--limit", type=int, default=3000, help="快掃模式每個 sheet 讀幾列")
    ap.add_argument("--files", default="*.xlsx",
                    help="只掃符合的檔名, 例如 --files \"2024-*.xlsx\"; 先拿一個小檔測速很有用")
    ap.add_argument("--to-parquet", action="store_true",
                    help="同一趟順便把明細寫成 parquet (自動啟用 --full)")
    args = ap.parse_args()

    if args.to_parquet:
        args.full = True   # 半截的 parquet 沒有意義

    root = Path(args.root)
    data_dir = root / "data"
    if not data_dir.is_dir():
        sys.exit(f"找不到資料夾: {data_dir}")

    pq_dir = None
    if args.to_parquet:
        try:
            import pyarrow  # noqa: F401
        except ImportError:
            sys.exit("--to-parquet 需要 pyarrow:  pip install pyarrow")
        pq_dir = root / "parquet"
        pq_dir.mkdir(exist_ok=True)

    limit = None if args.full else args.limit
    report: dict = {"root": str(root), "mode": "full" if args.full else f"quick({args.limit})", "data": []}

    for tpl in sorted(root.glob("*.xlsx")):
        if tpl.name.startswith("~$"):
            continue
        report["template"] = inspect_template(tpl)
        break

    acc = {
        "name": defaultdict(set), "cat": defaultdict(set), "keys": set(),
        "totals": defaultdict(lambda: defaultdict(float)),
        "blank_impact": defaultdict(lambda: defaultdict(float)),
    }
    files = [p for p in sorted(data_dir.glob(args.files)) if not p.name.startswith("~$")]
    if not files:
        sys.exit(f"{data_dir} 裡沒有符合 {args.files} 的檔案")
    for i, p in enumerate(files, 1):
        print(f"[{i}/{len(files)}] 掃描 {p.name} ({p.stat().st_size/1e6:.0f} MB) ...",
              file=sys.stderr, flush=True)
        try:
            r = inspect_data_file(p, limit, args.full, acc, pq_dir)
            report["data"].append(r)
            rows = sum(sh.get("rows_scanned", 0) for sh in r["sheets"])
            secs = max(r["seconds"], 0.1)
            print(f"      {rows:,} 列 / {r['seconds']}s = {rows/secs:,.0f} 列/秒",
                  file=sys.stderr, flush=True)
        except Exception as e:  # 單檔壞掉不要整批停
            report["data"].append({"file": p.name, "size_mb": round(p.stat().st_size / 1e6, 1),
                                   "seconds": 0, "sheets": [{"name": "?", "error": f"{type(e).__name__}: {e}"}]})

    dn = {k: sorted(v) for k, v in acc["name"].items() if len(v) > 1}
    dc = {k: sorted(v) for k, v in acc["cat"].items() if len(v) > 1}
    # 「空白 vs 有名字」跟「真的改名」是兩回事, 處理方式也不同, 要分開看
    dn_blank = {k: v for k, v in dn.items() if "" in v and len(v) == 2}
    dn_real = {k: v for k, v in dn.items() if k not in dn_blank}
    dc_blank = {k: v for k, v in dc.items() if "" in v and len(v) == 2}
    dc_real = {k: v for k, v in dc.items() if k not in dc_blank}
    report["global"] = {
        "name_drift_count": len(dn),
        "name_drift_blank_only": len(dn_blank),
        "name_drift_real": len(dn_real),
        "name_drift_sample": dict(list(dn_real.items())[:20]),
        "category_drift_count": len(dc),
        "category_drift_blank_only": len(dc_blank),
        "category_drift_real": len(dc_real),
        "category_drift_sample": dict(list(dc_real.items())[:20]),
        "unique_group_keys": len(acc["keys"]),
        "totals": {m: dict(v) for m, v in sorted(acc["totals"].items())},
        "blank_impact": {c: dict(v) for c, v in acc["blank_impact"].items()},
    }

    write_report(report, root)
    print(f"\n報告已寫到 {root / 'inspect_report.txt'} 與 inspect_report.json", file=sys.stderr)


if __name__ == "__main__":
    main()
