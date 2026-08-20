#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把明細 parquet 聚合成「key × 期間」的結果, 存成 parquet。

刻意不直接寫進模板 —— 聚合是慢又容易錯的一步, 排版是快又常改的一步,
分開之後模板改版不必重跑聚合。輸出是長格式 (每列一個 key×期間),
後續要轉成任何寬表都很快。

用法:
    python build_aggregate.py --channel-map channel_map.csv
    python build_aggregate.py --dedup exact          # 去掉完全重複的列
    python build_aggregate.py --keys "门店编码,销售渠道,商品编码"   # 自訂分組

跑完會印對帳表: 來源總數 vs 聚合後總數。兩者必須完全相等, 不然就是掉了資料。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import duckdb
except ImportError:
    sys.exit("需要 duckdb:  pip install duckdb")

# ================================================================= 設定
DEFAULT_KEYS = ["门店编码", "销售渠道", "产品分类", "商品编码", "商品名称"]
MEASURE_COLS = ["商品数量", "应收金额", "实收金额"]
CODE_COL, NAME_COL, CHANNEL_COL = "商品编码", "商品名称", "销售渠道"
PERIOD_COL = "period"
BLANK_TOKENS = {"(空白)", "（空白）", "<空白>", "空白", ""}
BLANK_KEY = "<<BLANK>>"
# ================================================================= 設定結束


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def read_pairs(path: Path, col_a: str = None, col_b: str = None) -> list[tuple[str, str]]:
    """讀兩欄對照檔 (csv/xlsx)。給了欄名就照欄名抓, 否則取前兩欄。"""
    rows: list[list[str]] = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if col_a:                      # 依欄名找工作表與欄位
            for cand in wb.worksheets:
                first = next(cand.iter_rows(max_row=1, values_only=True), None)
                hdr = [("" if c is None else str(c).strip()) for c in (first or ())]
                if col_a in hdr:
                    ws = cand
                    break
        first = next(ws.iter_rows(max_row=1, values_only=True), None)
        hdr = [("" if c is None else str(c).strip()) for c in (first or ())]
        ia = hdr.index(col_a) if col_a and col_a in hdr else 0
        ib = hdr.index(col_b) if col_b and col_b in hdr else 1
        # 只取人工填的那一欄。不要退回工作表的「建议名称」—— 那是產生工作表當下
        # 算的舊值, 會蓋掉這支腳本較新的解析結果。
        for r in ws.iter_rows(min_row=2, values_only=True):
            def get(i):
                return "" if i is None or i >= len(r) or r[i] is None else str(r[i]).strip()
            rows.append([get(ia), get(ib)])
        wb.close()
    else:
        import csv
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                with path.open(encoding=enc, newline="") as f:
                    rows = [[c.strip() for c in r[:2]] for r in csv.reader(f) if r]
                break
            except UnicodeDecodeError:
                continue
    pairs = []
    for r in rows:
        if len(r) < 2 or not r[0]:
            continue
        orig = BLANK_KEY if r[0] in BLANK_TOKENS else r[0]
        pairs.append((orig, r[1]))
    return pairs


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=str(Path(__file__).resolve().parent))
    ap.add_argument("--channel-map", default=None, help="渠道合併對照 csv/xlsx")
    ap.add_argument("--keys", default=",".join(DEFAULT_KEYS), help="分組欄位, 逗號分隔")
    ap.add_argument("--dedup", choices=["none", "exact"], default="none",
                    help="none=不去重(預設); exact=移除所有欄位都相同的列")
    ap.add_argument("--out", default="aggregate.parquet")
    args = ap.parse_args()

    root = Path(args.root)
    pq = root / "parquet"
    if not sorted(pq.glob("*.parquet")):
        sys.exit(f"找不到 parquet: {pq}  (先跑 inspect_xlsx.py --to-parquet)")
    keys = [k.strip() for k in args.keys.split(",") if k.strip()]

    con = duckdb.connect()
    src = f"read_parquet('{pq.as_posix()}/*.parquet')"
    all_cols = [r[0] for r in con.sql(f"DESCRIBE SELECT * FROM {src}").fetchall()]
    L: list[str] = []
    add = L.append

    def totals(rel: str) -> tuple:
        ms = ", ".join(f"sum(COALESCE({q(c)}, 0))" for c in MEASURE_COLS)
        return con.sql(f"SELECT count(*), {ms} FROM {rel}").fetchone()

    t_src = totals(src)
    add("=== 來源 ===")
    add(f"  列數 {t_src[0]:,}")
    for c, v in zip(MEASURE_COLS, t_src[1:]):
        add(f"  {c}: {v:,.2f}")

    # ------------------------------------------------------------- 去重
    stage = src
    if args.dedup == "exact":
        body = ", ".join(q(c) for c in all_cols if c != "source_file")
        con.execute(f"CREATE TEMP TABLE deduped AS SELECT DISTINCT {body} FROM {src}")
        stage = "deduped"
        t_d = totals(stage)
        add(f"\n=== 去重 (--dedup exact) ===")
        add(f"  移除 {t_src[0] - t_d[0]:,} 列")
        for c, a, b in zip(MEASURE_COLS, t_src[1:], t_d[1:]):
            add(f"  {c}: {a:,.2f} -> {b:,.2f}   (差 {a - b:,.2f})")
    else:
        add("\n=== 去重: 未執行 (--dedup none), 總數與來源一致 ===")

    # ------------------------------------------------------------- 渠道合併
    sel_map: dict[str, str] = {}
    if args.channel_map:
        mp = Path(args.channel_map)
        pairs = read_pairs(mp if mp.is_absolute() else root / mp)
        con.execute("CREATE TEMP TABLE chmap(orig VARCHAR, mapped VARCHAR)")
        con.executemany("INSERT INTO chmap VALUES (?, ?)", pairs)
        norm = f"COALESCE(NULLIF(TRIM(s.{q(CHANNEL_COL)}), ''), '{BLANK_KEY}')"
        miss = con.sql(f"""
            SELECT count(*) FROM {stage} s LEFT JOIN chmap m ON {norm} = m.orig
            WHERE m.mapped IS NULL
        """).fetchone()[0]
        add(f"\n=== 渠道合併 ({len(pairs)} 筆對照) ===")
        if miss:
            add(f"  !! {miss:,} 列的渠道不在對照檔裡, 會保留原值")
        else:
            add("  對照檔完整覆蓋")
        sel_map[CHANNEL_COL] = f"COALESCE(m.mapped, s.{q(CHANNEL_COL)})"

    # ------------------------------------------------------------- 名稱解析
    # 每個編碼定一個名稱, 優先序:
    #   1. 最後一次出現的「非垃圾」商品名称
    #   2. 最後一次出現的「非垃圾」系统商品名称
    #   3. 空白 (不編造)
    # 垃圾值要排除, 否則像 '0' 這種值若剛好出現在最後, 會蓋掉真名稱。
    junk = ("({c} ~ '^[0-9]+(\\.[0-9]+)?$' OR length(trim({c})) <= 1 "
            "OR {c} IN ('-','--','null','NULL','N/A','#N/A') "
            "OR {c} LIKE '%测试%' OR lower({c}) LIKE '%test%')")
    jn, js = junk.format(c=q(NAME_COL)), junk.format(c='"系统商品名称"')
    con.execute(f"""
        CREATE TEMP TABLE resolved AS
        SELECT {q(CODE_COL)} AS code,
               COALESCE(
                 arg_max({q(NAME_COL)}, {q(PERIOD_COL)})
                     FILTER (WHERE {q(NAME_COL)} IS NOT NULL AND NOT {jn}),
                 arg_max("系统商品名称", {q(PERIOD_COL)})
                     FILTER (WHERE "系统商品名称" IS NOT NULL AND NOT {js})
               ) AS name,
               max(CASE WHEN {q(NAME_COL)} IS NOT NULL AND NOT {jn} THEN 1 ELSE 0 END) AS from_name
        FROM {stage} WHERE {q(CODE_COL)} IS NOT NULL GROUP BY 1
    """)
    rs = con.sql("""SELECT count(*), count(*) FILTER (WHERE from_name = 1),
                           count(*) FILTER (WHERE from_name = 0 AND name IS NOT NULL),
                           count(*) FILTER (WHERE name IS NULL) FROM resolved""").fetchone()
    add(f"\n=== 名稱解析 ===")
    add(f"  編碼總數 {rs[0]:,}")
    add(f"  用 {NAME_COL} {rs[1]:,}   退回 系统商品名称 {rs[2]:,}   兩者皆無留空 {rs[3]:,}")
    sel_map[NAME_COL] = f"r.name"

    # ------------------------------------------------------------- 聚合
    joins = f" LEFT JOIN resolved r ON s.{q(CODE_COL)} = r.code"
    if args.channel_map:
        joins += f" LEFT JOIN chmap m ON COALESCE(NULLIF(TRIM(s.{q(CHANNEL_COL)}), ''), '{BLANK_KEY}') = m.orig"

    key_exprs = [f"{sel_map.get(k, 's.' + q(k))} AS {q(k)}" for k in keys]
    sums = [f"sum(COALESCE(s.{q(c)}, 0)) AS {q(c)}" for c in MEASURE_COLS]
    group_by = ", ".join(str(i + 1) for i in range(len(keys) + 1))

    out_path = root / args.out
    con.execute(f"""
        COPY (
            SELECT {", ".join(key_exprs)}, s.{q(PERIOD_COL)} AS {q(PERIOD_COL)},
                   {", ".join(sums)}, count(*) AS 明细列数
            FROM {stage} s{joins}
            GROUP BY {group_by}
        ) TO '{out_path.as_posix()}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)

    agg = f"read_parquet('{out_path.as_posix()}')"
    t_agg = totals(agg)
    n_keys = con.sql(f"SELECT count(*) FROM (SELECT DISTINCT {', '.join(q(k) for k in keys)} FROM {agg})").fetchone()[0]

    add("\n=== 聚合結果 ===")
    add(f"  分組欄位: {keys}")
    add(f"  key × 期間 列數: {t_agg[0]:,}")
    add(f"  不重複 key 數 (= 寬表列數): {n_keys:,}")
    add(f"  檔案: {out_path.name}  {out_path.stat().st_size / 1e6:,.1f} MB")

    add("\n=== 對帳 (聚合後 vs 聚合前, 差額必須為 0) ===")
    base = t_d if args.dedup == "exact" else t_src
    ok = True
    for c, a, b in zip(MEASURE_COLS, base[1:], t_agg[1:]):
        diff = a - b
        if abs(diff) > 0.005:
            ok = False
        add(f"  {c}: {a:,.2f} vs {b:,.2f}   差 {diff:,.2f}{'' if abs(diff) <= 0.005 else '   !! 不符'}")
    add(f"  -> {'對帳通過' if ok else '!! 對帳失敗, 不要使用這份結果'}")

    txt = "\n".join(L)
    (root / "aggregate_report.txt").write_text(txt, encoding="utf-8")
    print(txt)
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
