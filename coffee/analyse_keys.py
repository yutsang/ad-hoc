#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
從 parquet 算出「聚合後到底會有幾列」, 以及不同分組規則各能省下多少。

inspect_xlsx.py --to-parquet 跑完之後用這支。讀 parquet 所以是秒級的,
可以反覆試不同規則, 不用再碰 xlsx。

用法:
    python analyse_keys.py
    python analyse_keys.py --root /path/to/dir
    python analyse_keys.py --master data/xxx.xlsx     # 順便抽出商品對照表

輸出全部是聚合統計, 不含任何明細列, 可以直接貼出來討論。
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import duckdb
except ImportError:
    sys.exit("需要 duckdb:  pip install duckdb")

# ================================================================= 設定
GROUP_COLS = ["门店编码", "销售渠道", "产品分类", "商品编码", "商品名称"]
MEASURE_COLS = ["商品数量", "应收金额", "实收金额"]
CODE_COL, NAME_COL, CATEGORY_COL, CHANNEL_COL = "商品编码", "商品名称", "产品分类", "销售渠道"
PERIOD_COL = "period"
PERIOD_HINT = "日期"          # 表頭含這個字 -> 是明細 sheet, 不是對照表
EXCEL_ROW_LIMIT = 1_048_576
BLANK_TOKENS = {"(空白)", "（空白）", "<空白>", "空白", ""}   # 對照檔裡代表「原值為空」的寫法
BLANK_KEY = "<<BLANK>>"       # 內部代表「原值為空」的哨符, 不可含控制字元 (會破壞 SQL 字面值)
# ================================================================= 設定結束


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def read_pairs(path: Path) -> list[tuple[str, str]]:
    """讀兩欄的對照檔 (csv 或 xlsx), 回傳 [(原值, 對應值)]。第一列若是標題會被略過。"""
    rows: list[list[str]] = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        for r in wb.active.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c).strip()) for c in (r or [])[:2]])
        wb.close()
    else:
        import csv
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                with path.open(encoding=enc, newline="") as f:
                    rows = [[c.strip() for c in r[:2]] for r in csv.reader(f) if r]
                break
            except UnicodeDecodeError:
                continue
    pairs = []
    for r in rows:
        if len(r) < 2 or not r[1]:
            continue
        orig = BLANK_KEY if r[0] in BLANK_TOKENS else r[0]
        pairs.append((orig, r[1]))
    # 不去猜哪一列是標題: 多餘的列對不到資料就是不會用到, 無害。
    # 真正要防的是「資料裡有、對照檔沒有」的值, 那個在下面會逐一列出。
    return pairs


def fmt_n(n) -> str:
    return f"{int(n):,}"


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=str(Path(__file__).resolve().parent))
    ap.add_argument("--master", default=None,
                    help="含商品對照表的 xlsx (例如 inspect report 裡那個被跳過的 sheet)")
    ap.add_argument("--channel-map", default=None,
                    help="渠道合併對照 csv/xlsx, 兩欄: 原渠道, 合併後渠道")
    args = ap.parse_args()

    root = Path(args.root)
    pq = root / "parquet"
    files = sorted(pq.glob("*.parquet"))
    if not files:
        sys.exit(f"找不到 parquet: {pq}  (先跑 inspect_xlsx.py --to-parquet)")

    con = duckdb.connect()
    src = f"read_parquet('{pq.as_posix()}/*.parquet')"
    L: list[str] = []
    add = L.append

    # ---------------------------------------------------------- 基本量體
    n_rows = con.sql(f"SELECT count(*) FROM {src}").fetchone()[0]
    add(f"=== 資料量 ===")
    add(f"  parquet {len(files)} 檔, {fmt_n(n_rows)} 列, "
        f"{sum(f.stat().st_size for f in files) / 1e6:,.0f} MB")

    dims = ", ".join(f"count(DISTINCT {q(c)}) AS {q(c)}" for c in GROUP_COLS)
    row = con.sql(f"SELECT count(DISTINCT {q(PERIOD_COL)}) AS 月份, {dims} FROM {src}").fetchone()
    cols = ["月份"] + GROUP_COLS
    add("  各維度 distinct: " + "  ".join(f"{c}={fmt_n(v)}" for c, v in zip(cols, row)))

    all_cols = [r[0] for r in con.sql(f"DESCRIBE SELECT * FROM {src}").fetchall()]
    key_period = [PERIOD_COL] + GROUP_COLS

    # ---------------------------------------------------------- 重複偵測
    # 聚合會把重複列一起加總, 所以真重複 = 金額被灌水。這裡分三種嚴重度來看。
    add("\n=== 重複偵測 ===")

    # (1) 整列完全一樣 (連來源檔都一樣) -> 幾乎確定是匯出時重覆
    body = ", ".join(q(c) for c in all_cols if c != "source_file")
    r = con.sql(f"""
        WITH d AS (SELECT {body}, count(*) AS c FROM {src} GROUP BY ALL)
        SELECT COALESCE(count(*) FILTER (WHERE c > 1), 0),
               COALESCE(sum(c - 1) FILTER (WHERE c > 1), 0)
        FROM d
    """).fetchone()
    add(f"  (1) 所有欄位都完全相同的列: {fmt_n(r[0])} 組, 多出 {fmt_n(r[1])} 列")
    if r[1]:
        amt = con.sql(f"""
            WITH d AS (SELECT {body}, count(*) AS c FROM {src} GROUP BY ALL)
            SELECT COALESCE(sum((c - 1) * COALESCE({q('实收金额')}, 0)), 0) FROM d WHERE c > 1
        """).fetchone()[0]
        add(f"      -> 若去重, 实收金额 會少 {amt:,.2f}")

        # 關鍵: 交易級明細裡「兩筆一模一樣的列」是合理的(兩個客人買了同樣的東西),
        # 去重會毀掉真資料。所以要看重複集中在哪些月份, 而不是無條件去重。
        add("\n      重複列的月份分布 (集中在少數月份 -> 可能是匯出問題;"
            " 平均分布在交易級月份 -> 可能是真交易):")
        per = con.sql(f"""
            WITH d AS (SELECT {body}, {q(PERIOD_COL)} AS p, count(*) AS c FROM {src} GROUP BY ALL),
                 tot AS (SELECT {q(PERIOD_COL)} AS p, count(*) AS n FROM {src} GROUP BY 1)
            SELECT d.p, sum(d.c - 1) AS excess, tot.n,
                   sum((d.c - 1) * COALESCE(d.{q('实收金额')}, 0)) AS amt
            FROM d JOIN tot ON d.p = tot.p
            WHERE d.c > 1 GROUP BY d.p, tot.n
            HAVING sum(d.c - 1) > 0 ORDER BY excess DESC LIMIT 12
        """).fetchall()
        add(f"      {'期間':<9}{'多出列數':>12}{'佔該月':>10}{'实收金额':>18}")
        for p, excess, ntot, a in per:
            add(f"      {p:<9}{excess:>12,}{excess / ntot * 100:>9.2f}%{a:>18,.0f}")
    else:
        add("      -> 沒有整列重複, 不會有灌水")

    # (2) key + 月份 + 三個度量都一樣 -> 分不出是真的兩筆還是重複
    dupcols = ", ".join(q(c) for c in key_period + MEASURE_COLS)
    r2 = con.sql(f"""
        WITH d AS (SELECT {dupcols}, count(*) AS c FROM {src} GROUP BY ALL)
        SELECT COALESCE(count(*) FILTER (WHERE c > 1), 0),
               COALESCE(sum(c - 1) FILTER (WHERE c > 1), 0),
               COALESCE(sum((c - 1) * COALESCE({q('实收金额')}, 0)) FILTER (WHERE c > 1), 0)
        FROM d
    """).fetchone()
    add(f"  (2) key+月份+金額都相同的列: {fmt_n(r2[0])} 組, 多出 {fmt_n(r2[1])} 列, "
        f"涉及 实收金额 {r2[2]:,.2f}")
    add("      -> 可能是同月同店同品的兩筆不同交易, 也可能是重複; 光看資料分不出來")

    # (3) 每個 key×月份 由幾列組成 -> 正常應該接近 1
    r3 = con.sql(f"""
        WITH d AS (SELECT {", ".join(q(c) for c in key_period)}, count(*) AS c FROM {src} GROUP BY ALL)
        SELECT max(c), avg(c), quantile_cont(c, 0.99) FROM d
    """).fetchone()
    add(f"  (3) 每個 key×月份 的列數: 平均 {r3[1]:.2f}, 99 分位 {r3[2]:.0f}, 最大 {fmt_n(r3[0])}")

    # ---------------------------------------------------------- 欄位邏輯一致性
    # 應收 - 優惠 = 實收 若成立, 代表欄位對應正確, 低實收率是真的折扣而不是讀錯欄。
    if all(c in all_cols for c in ("应收金额", "实收金额", "优惠金额")):
        add("\n=== 欄位邏輯一致性: 应收 - 优惠 = 实收 ? ===")
        r = con.sql(f"""
            SELECT count(*),
                   count(*) FILTER (WHERE abs(COALESCE({q('应收金额')},0)
                        - COALESCE({q('优惠金额')},0) - COALESCE({q('实收金额')},0)) > 0.01),
                   sum(COALESCE({q('应收金额')},0)), sum(COALESCE({q('优惠金额')},0)),
                   sum(COALESCE({q('实收金额')},0))
            FROM {src}
        """).fetchone()
        n, bad, ar, disc, real = r
        add(f"  不符的列: {fmt_n(bad)} / {fmt_n(n)}  ({bad / n * 100:.2f}%)")
        add(f"  应收 {ar:,.2f}  -  优惠 {disc:,.2f}  =  {ar - disc:,.2f}")
        add(f"  实收 {real:,.2f}   差額 {ar - disc - real:,.2f}")
        add(f"  实收 / 应收 = {real / ar * 100:.1f}%" if ar else "")
        add("  -> 若不符比例接近 0, 欄位對應正確, 低實收率是真實折扣; "
            "否則代表有欄位讀錯或另有扣項")

    # ---------------------------------------------------------- 逐月粒度
    # 某個月列數暴增, 若 date_raw 的 distinct 數也跟著暴增, 代表那個月是按日給的,
    # 不是重複。這是分辨「粒度不同」與「資料重複」的關鍵。
    add("\n=== 逐月粒度 (揪出哪個月的明細顆粒度跟別人不一樣) ===")
    extra = [c for c in ("date_raw", "规格", "系统商品名称", "门店名称") if c in all_cols]
    extra_sel = "".join(f", count(DISTINCT {q(c)}) AS {q(c)}" for c in extra)
    rows = con.sql(f"""
        SELECT {q(PERIOD_COL)} AS p, count(*) AS rows,
               count(DISTINCT ({", ".join(q(c) for c in GROUP_COLS)})) AS keys{extra_sel}
        FROM {src} GROUP BY 1 ORDER BY 1
    """).fetchall()
    hdr = f"  {'期間':<9}{'列數':>12}{'key 數':>12}{'列/key':>9}"
    hdr += "".join(f"{('distinct ' + c):>20}" for c in extra)
    add(hdr)
    ratios = []
    for rr in rows:
        p, nrows, nkeys = rr[0], rr[1], rr[2]
        ratio = nrows / nkeys if nkeys else 0
        ratios.append((ratio, p))
        line = f"  {p:<9}{nrows:>12,}{nkeys:>12,}{ratio:>9.2f}"
        line += "".join(f"{v:>20,}" for v in rr[3:])
        add(line)
    if ratios:
        med = sorted(r for r, _ in ratios)[len(ratios) // 2]
        odd = [p for r, p in ratios if med and r > med * 1.8]
        add(f"  中位數 列/key = {med:.2f}; 明顯偏高的月份: {odd if odd else '無'}")
        if odd:
            add("  -> 若這些月份的 distinct date_raw 也同步偏高, 就是按日給的(粒度不同),")
            add("     聚合後會正確合併; 若 date_raw 沒變多, 才要懷疑重複。")

            # 逐月比值偏高但 date_raw 沒變 -> 要查是「別的維度撐開」還是「同維度多筆交易」
            add("\n=== 異常月份鑽取: 多出來的列到底差在哪 ===")
            dim_cols = [c for c in all_cols
                        if c not in MEASURE_COLS + [PERIOD_COL, "source_file"]
                        and c not in ("优惠金额",)]
            kk = ", ".join(q(c) for c in GROUP_COLS)
            dd = ", ".join(q(c) for c in dim_cols)
            add(f"  {'期間':<9}{'列數':>12}{'照 key':>12}{'照全部維度':>14}{'判定':>26}")
            for p in odd:
                rr = con.sql(f"""
                    SELECT count(*), count(DISTINCT ({kk})), count(DISTINCT ({dd}))
                    FROM {src} WHERE {q(PERIOD_COL)} = '{p}'
                """).fetchone()
                nrows, by_key, by_dims = rr
                if by_dims >= nrows * 0.95:
                    verdict = "其他維度撐開的"
                elif by_dims <= nrows * 0.5:
                    verdict = "同維度多筆(交易級)"
                else:
                    verdict = "混合"
                add(f"  {p:<9}{nrows:>12,}{by_key:>12,}{by_dims:>14,}{verdict:>26}")
            add("  照全部維度 ≈ 列數  -> 是 规格/系统商品名称 等欄位把同一個 key 拆開, 聚合會合併")
            add("  照全部維度 << 列數  -> 同一組維度有多筆列, 是交易級明細, 聚合一樣會合併")
            add("  兩種都不會讓總額變多; 真正會灌水的只有上面「重複偵測(1)」那一項。")

    # ---------------------------------------------------------- 分組方案比較
    add("\n=== 聚合後列數: 不同分組規則的比較 ===")
    add(f"  (Excel 單一工作表上限 {fmt_n(EXCEL_ROW_LIMIT)} 列)")

    plans = [
        ("A 模板原樣 5 個 key", GROUP_COLS, ""),
        ("B 名稱不進 key (名稱改成由編碼推導)", [c for c in GROUP_COLS if c != NAME_COL], ""),
        ("C B + 分類也不進 key", [c for c in GROUP_COLS if c not in (NAME_COL, CATEGORY_COL)], ""),
        ("D C + 渠道也不進 key", [c for c in GROUP_COLS if c not in (NAME_COL, CATEGORY_COL, CHANNEL_COL)], ""),
    ]
    results = {}
    for label, keys, _ in plans:
        sel = ", ".join(q(c) for c in keys)
        n = con.sql(f"SELECT count(*) FROM (SELECT DISTINCT {sel} FROM {src})").fetchone()[0]
        results[label] = n
        flag = "放得進 Excel" if n <= EXCEL_ROW_LIMIT else f"!! 超出 {fmt_n(n - EXCEL_ROW_LIMIT)} 列"
        add(f"  {label:<38} {fmt_n(n):>12} 列   {flag}")

    # ---------------------------------------------------------- 渠道合併
    if args.channel_map:
        mp = Path(args.channel_map)
        pairs = read_pairs(mp if mp.is_absolute() else root / mp)
        con.execute("CREATE TEMP TABLE chmap(orig VARCHAR, mapped VARCHAR)")
        con.executemany("INSERT INTO chmap VALUES (?, ?)", pairs)
        norm = f"COALESCE(NULLIF(TRIM({q(CHANNEL_COL)}), ''), '{BLANK_KEY}')"
        joined = f"{src} s LEFT JOIN chmap m ON {norm.replace(q(CHANNEL_COL), 's.' + q(CHANNEL_COL))} = m.orig"

        add(f"\n=== 渠道合併 (對照檔 {mp.name}, {len(pairs)} 筆) ===")
        cov = con.sql(f"""
            SELECT COALESCE(m.mapped, '!! 未映射') AS ch,
                   count(DISTINCT COALESCE(NULLIF(TRIM(s.{q(CHANNEL_COL)}), ''), '(空白)')) AS n_orig,
                   count(*) AS rows, sum(COALESCE(s.{q('实收金额')}, 0)) AS amt
            FROM {joined} GROUP BY 1 ORDER BY rows DESC
        """).fetchall()
        add(f"  {'合併後':<14}{'原渠道數':>10}{'列數':>14}{'实收金额':>20}")
        for ch, n_orig, rows_, amt in cov:
            add(f"  {ch:<14}{n_orig:>10,}{rows_:>14,}{amt:>20,.0f}")

        miss = con.sql(f"""
            SELECT COALESCE(NULLIF(TRIM(s.{q(CHANNEL_COL)}), ''), '(空白)') AS ch,
                   count(*) AS rows, sum(COALESCE(s.{q('实收金额')}, 0)) AS amt
            FROM {joined} WHERE m.mapped IS NULL GROUP BY 1 ORDER BY rows DESC
        """).fetchall()
        if miss:
            add("  !! 資料裡有但對照檔沒有的渠道 (這些列會歸不到類):")
            for ch, rows_, amt in miss:
                add(f"     {ch}: {rows_:,} 列, 实收金额 {amt:,.2f}")
        else:
            add("  對照檔完整覆蓋資料裡所有渠道")

        others = [c for c in GROUP_COLS if c != CHANNEL_COL]
        for label, keys in [
            ("E 渠道合併 (其餘照模板)", others),
            ("F 渠道合併 + 名稱移出 key", [c for c in others if c != NAME_COL]),
        ]:
            sel = ", ".join("s." + q(c) for c in keys)
            n = con.sql(f"""
                SELECT count(*) FROM (
                    SELECT DISTINCT {sel}, COALESCE(m.mapped, '!! 未映射') FROM {joined})
            """).fetchone()[0]
            results[label] = n
            flag = "放得進 Excel" if n <= EXCEL_ROW_LIMIT else f"!! 超出 {fmt_n(n - EXCEL_ROW_LIMIT)} 列"
            add(f"  {label:<38} {fmt_n(n):>12} 列   {flag}")
        base = results.get("A 模板原樣 5 個 key")
        if base:
            e = results["E 渠道合併 (其餘照模板)"]
            add(f"  -> 相對方案 A 減少 {fmt_n(base - e)} 列 ({(base - e) / base * 100:.1f}%)")

    # 分類是否真的能由編碼推導 (能的話 B 跟 C 會一樣)
    n_code_multi_cat = con.sql(f"""
        SELECT count(*) FROM (
            SELECT {q(CODE_COL)} FROM {src}
            WHERE {q(CATEGORY_COL)} IS NOT NULL
            GROUP BY 1 HAVING count(DISTINCT {q(CATEGORY_COL)}) > 1)
    """).fetchone()[0]
    add(f"\n  一個 {CODE_COL} 對到多個 {CATEGORY_COL} 的個數: {fmt_n(n_code_multi_cat)}"
        f"  -> {'分類可由編碼唯一推導' if n_code_multi_cat == 0 else '分類無法唯一推導, 不能移出 key'}")

    n_code_multi_name = con.sql(f"""
        SELECT count(*) FROM (
            SELECT {q(CODE_COL)} FROM {src}
            WHERE {q(NAME_COL)} IS NOT NULL
            GROUP BY 1 HAVING count(DISTINCT {q(NAME_COL)}) > 1)
    """).fetchone()[0]
    add(f"  一個 {CODE_COL} 對到多個非空 {NAME_COL} 的個數: {fmt_n(n_code_multi_name)}"
        f"  -> 這些就是把 A 撐大的原因")

    # ---------------------------------------------------------- 長尾
    add("\n=== 長尾: 每個 key 實際有幾個月有銷售 ===")
    add("  (只出現一兩個月的 key 佔了多少列數與金額, 決定要不要設門檻)")
    sel = ", ".join(q(c) for c in GROUP_COLS)
    tail = con.sql(f"""
        WITH k AS (
            SELECT {sel},
                   count(DISTINCT {q(PERIOD_COL)}) AS n_months,
                   sum(COALESCE({q('实收金额')}, 0)) AS amt
            FROM {src} GROUP BY {sel})
        SELECT CASE WHEN n_months = 1 THEN '1 個月'
                    WHEN n_months <= 3 THEN '2-3 個月'
                    WHEN n_months <= 6 THEN '4-6 個月'
                    WHEN n_months <= 12 THEN '7-12 個月'
                    ELSE '13+ 個月' END AS 區間,
               min(n_months) AS srt, count(*) AS keys, sum(amt) AS amt
        FROM k GROUP BY 1 ORDER BY srt
    """).fetchall()
    tot_keys = sum(r[2] for r in tail) or 1
    tot_amt = sum(r[3] for r in tail) or 1
    add(f"  {'活躍月數':<12}{'key 數':>14}{'佔比':>9}{'实收金额':>20}{'佔比':>9}")
    for label, _, keys, amt in tail:
        add(f"  {label:<12}{fmt_n(keys):>14}{keys / tot_keys * 100:>8.1f}%"
            f"{amt:>20,.0f}{amt / tot_amt * 100:>8.2f}%")

    # 設門檻時要看的是「留下多少列 / 保住多少金額」, 所以要累計
    add("\n  設門檻的取捨 (只保留活躍月數 >= N 的 key):")
    add(f"  {'門檻':<10}{'保留 key 數':>16}{'vs Excel 上限':>18}{'保住金額佔比':>16}")
    cum_k, cum_a = tot_keys, tot_amt
    for label, srt, keys, amt in tail:
        fits = "放得進" if cum_k <= EXCEL_ROW_LIMIT else f"超出 {fmt_n(cum_k - EXCEL_ROW_LIMIT)}"
        add(f"  >= {srt:<7}{fmt_n(cum_k):>16}{fits:>18}{cum_a / tot_amt * 100:>15.2f}%")
        cum_k -= keys
        cum_a -= amt

    # ---------------------------------------------------------- 空白 key
    add("\n=== 空白 key 的影響 ===")
    for c in GROUP_COLS:
        r = con.sql(f"""
            SELECT count(*), sum(COALESCE({q('实收金额')}, 0))
            FROM {src} WHERE {q(c)} IS NULL
        """).fetchone()
        if r[0]:
            add(f"  {c} 為空: {fmt_n(r[0])} 列, 实收金额 {r[1]:,.0f}"
                f" ({r[1] / tot_amt * 100:.2f}%)")

    # 空白名稱能不能靠同編碼的其他月份補回來
    fillable = con.sql(f"""
        SELECT count(*) FROM (
            SELECT {q(CODE_COL)} FROM {src}
            GROUP BY 1
            HAVING count(*) FILTER (WHERE {q(NAME_COL)} IS NULL) > 0
               AND count(*) FILTER (WHERE {q(NAME_COL)} IS NOT NULL) > 0)
    """).fetchone()[0]
    orphan = con.sql(f"""
        SELECT count(*) FROM (
            SELECT {q(CODE_COL)} FROM {src}
            GROUP BY 1 HAVING count({q(NAME_COL)}) = 0)
    """).fetchone()[0]
    add(f"  空白 {NAME_COL} 可用同編碼其他列回填的編碼數: {fmt_n(fillable)}")
    add(f"  完全查不到名稱的編碼數: {fmt_n(orphan)}")

    # ---------------------------------------------------------- 控制總數
    add("\n=== 控制總數 (應與 inspect_report 完全一致) ===")
    ms = ", ".join(f"sum(COALESCE({q(c)}, 0))" for c in MEASURE_COLS)
    r = con.sql(f"SELECT count(*), {ms} FROM {src}").fetchone()
    add(f"  列數 {fmt_n(r[0])}")
    for c, v in zip(MEASURE_COLS, r[1:]):
        add(f"  {c}: {v:,.2f}")

    # ---------------------------------------------------------- 商品對照表
    if args.master:
        add("\n=== 商品對照表 ===")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(root / args.master if not Path(args.master).is_absolute()
                               else Path(args.master), read_only=True, data_only=True)
            found = False
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(max_row=2, values_only=True))
                if not rows:
                    continue
                hdr = [h for h in ("" if c is None else str(c).strip() for c in rows[0]) if h]
                # 對照表 = 欄位少、有商品編碼、且沒有日期欄 (有日期的是明細 sheet)
                is_master = (CODE_COL in hdr and len(hdr) <= 6
                             and not any(PERIOD_HINT in h for h in hdr))
                if is_master:
                    found = True
                    n = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
                    add(f"  sheet '{ws.title}' 欄位: {hdr}")
                    add(f"  資料列數: {fmt_n(n)}")
            if not found:
                add(f"  {args.master} 裡找不到含 {CODE_COL} 的對照表 sheet")
            wb.close()
        except Exception as e:
            add(f"  讀取失敗: {type(e).__name__}: {e}")

    txt = "\n".join(L)
    (root / "analyse_report.txt").write_text(txt, encoding="utf-8")
    print(txt)
    print(f"\n報告已寫到 {root / 'analyse_report.txt'}", file=sys.stderr)


if __name__ == "__main__":
    main()
