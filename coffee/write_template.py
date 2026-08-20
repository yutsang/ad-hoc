#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把 aggregate.parquet 填進聚合表模板, 產出最終交付檔。

模板結構 (照 聚合表模板.xlsx):
    row1  空
    row2  三個區塊標題, 在第 6 / 48 / 90 欄
    row3  欄位名: 1-5 是 key, 6-47 商品数量, 48-89 应收金额, 90-131 实收金额
    row4  起是資料

用法:
    python write_template.py
"""
import sys
import time
from pathlib import Path

import duckdb
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent
TEMPLATE = ROOT / "聚合表模板.xlsx"
AGG = ROOT / "aggregate.parquet"
OUT = ROOT / "聚合表_成果.xlsx"

KEYS = ["门店编码", "销售渠道", "产品分类", "商品编码", "商品名称"]
MEASURES = ["商品数量", "应收金额", "实收金额"]


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    # 聚合規則改過就得重跑, 不然這裡填的是舊資料, 而且對帳照樣會過
    builder = ROOT / "build_aggregate.py"
    if AGG.exists() and builder.exists() and AGG.stat().st_mtime < builder.stat().st_mtime:
        sys.exit(f"{AGG.name} 比 {builder.name} 舊, 是用舊規則產生的。\n"
                 f"請先重跑:  python build_aggregate.py --channel-map channel_map.csv")

    # ---- 讀模板前三列, 拿到月份順序與區塊位置 -------------------------
    wb_t = load_workbook(TEMPLATE, data_only=True)
    ws_t = wb_t.active
    rows123 = [[c for c in r] for r in ws_t.iter_rows(min_row=1, max_row=3, values_only=True)]
    header = ["" if c is None else str(c).strip() for c in rows123[2]]
    wb_t.close()

    periods = [h for h in header[len(KEYS):] if h]
    n_p = len(periods) // len(MEASURES)
    block_periods = periods[:n_p]
    # 三個區塊的月份序列必須一致, 不然欄位對應會錯位
    for i in range(len(MEASURES)):
        assert periods[i * n_p:(i + 1) * n_p] == block_periods, "區塊月份序列不一致"
    # period -> 該期間在各區塊裡的欄索引 (0-based, 相對整列)
    col_of = {p: len(KEYS) + i for i, p in enumerate(block_periods)}
    print(f"模板: {n_p} 個月 ({block_periods[0]} ~ {block_periods[-1]}), "
          f"{len(KEYS)} 個 key 欄, 共 {len(header)} 欄", file=sys.stderr)

    # ---- 讀聚合結果, 依 key 排序後串流 -------------------------------
    con = duckdb.connect()

    # 先驗證期間都在模板範圍內。寫到一半才發現的話, 錯誤會被 lxml 的串流例外蓋掉。
    stray = con.execute(f"""
        SELECT DISTINCT "period" FROM read_parquet('{AGG.as_posix()}')
        WHERE "period" IS NULL OR "period" NOT IN ({", ".join(repr(p) for p in block_periods)})
        ORDER BY 1
    """).fetchall()
    if stray:
        sys.exit(f"聚合結果有模板容納不了的期間: {[s[0] for s in stray]}\n"
                 f"模板只有 {block_periods[0]} ~ {block_periods[-1]}")

    key_sql = ", ".join(f'"{k}"' for k in KEYS)
    cur = con.execute(f"""
        SELECT {key_sql}, "period", {", ".join(f'"{m}"' for m in MEASURES)}
        FROM read_parquet('{AGG.as_posix()}')
        ORDER BY {key_sql}
    """)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(ws_t.title)
    for r in rows123:
        ws.append(list(r))

    n_keys = len(KEYS)
    row_len = len(header)
    totals = [0.0] * len(MEASURES)
    col_totals = [0.0] * row_len      # 逐欄加總 -> 才抓得到月份放錯欄
    n_rows = n_cells = 0
    t0 = time.time()

    cur_key = None
    buf = None

    def flush():
        nonlocal n_rows, n_cells
        if buf is not None:
            ws.append(buf)
            n_rows += 1
            n_cells += sum(1 for v in buf if v is not None)
            if n_rows % 100_000 == 0:
                print(f"  已寫 {n_rows:,} 列  ({time.time() - t0:.0f}s)",
                      file=sys.stderr, flush=True)

    while True:
        batch = cur.fetchmany(100_000)
        if not batch:
            break
        for rec in batch:
            key = rec[:n_keys]
            if key != cur_key:
                flush()
                cur_key = key
                buf = [None] * row_len
                buf[:n_keys] = list(key)
            base = col_of[rec[n_keys]]
            for i, m in enumerate(MEASURES):
                v = rec[n_keys + 1 + i] or 0.0
                # 這個 key×期間 有明細就寫值, 即使淨額是 0 (例如退貨剛好沖銷),
                # 空白只代表「那個月完全沒有這筆組合」, 兩者語意不同
                buf[base + i * n_p] = v
                totals[i] += v
                col_totals[base + i * n_p] += v
    flush()

    print(f"寫入 {n_rows:,} 列 / {n_cells:,} 個非空儲存格, 存檔中...", file=sys.stderr)
    wb.save(OUT)

    # ---- 對帳 ---------------------------------------------------------
    # 只比總數抓不到「月份寫錯欄」: 值在同一個區塊內挪位, 總數照樣相等。
    # 所以逐月逐度量比對, 這才是有效的檢查。
    exp = con.execute(f"""
        SELECT {", ".join(f'sum("{m}")' for m in MEASURES)}
        FROM read_parquet('{AGG.as_posix()}')
    """).fetchone()
    per = {r[0]: r[1:] for r in con.execute(f"""
        SELECT "period", {", ".join(f'sum("{m}")' for m in MEASURES)}
        FROM read_parquet('{AGG.as_posix()}') GROUP BY 1
    """).fetchall()}

    print("\n=== 對帳 1/2: 總數 ===")
    ok = True
    for m, got, want in zip(MEASURES, totals, exp):
        diff = got - (want or 0)
        if abs(diff) > 0.005:
            ok = False
        print(f"  {m}: {got:,.2f} vs {want:,.2f}   差 {diff:,.2f}"
              f"{'' if abs(diff) <= 0.005 else '   !! 不符'}")

    print("\n=== 對帳 2/2: 逐月逐度量 (抓欄位錯位) ===")
    # 關鍵: 這裡不能再用 col_of。用 col_of 寫、又用 col_of 驗, 是拿同一個映射
    # 自己對自己, 就算映射整個位移也永遠一致。改成從實際寫出的表頭 header[j]
    # 反推該欄是哪個期間, 驗證路徑才與寫入路徑獨立。
    bad = []
    got_by_period = {p: [0.0] * len(MEASURES) for p in block_periods}
    for i in range(len(MEASURES)):
        for off in range(n_p):
            j = len(KEYS) + i * n_p + off
            p = header[j]
            want = (per.get(p) or (0.0,) * len(MEASURES))[i] or 0.0
            got = col_totals[j]
            got_by_period[p][i] = got
            if abs(got - want) > 0.005:
                bad.append((p, MEASURES[i], got, want))
    lines = ["期间," + ",".join(MEASURES)]
    for p in block_periods:
        lines.append(p + "," + ",".join(f"{v:.2f}" for v in got_by_period[p]))
    if bad:
        ok = False
        for p, m, g, w in bad[:20]:
            print(f"  !! {p} {m}: {g:,.2f} vs {w:,.2f}")
    print(f"  {len(block_periods)} 個月 × {len(MEASURES)} 個度量 = "
          f"{len(block_periods) * len(MEASURES)} 項, 不符 {len(bad)} 項")

    rec = ROOT / "对数表.csv"
    rec.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"\n逐月對數表: {rec.name}  (拿去跟客戶系統的月報並排比)")
    print(f"  -> {'對帳通過' if ok else '!! 對帳失敗, 不要交付這份檔案'}")
    print(f"\n輸出: {OUT}  {OUT.stat().st_size / 1e6:,.1f} MB  "
          f"({time.time() - t0:.0f}s)")
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
