#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把寬表攤成 SKU 模板要的月度長表。

模板 (SKU template.xlsx / 工作表 "SKU 月度") 要的欄位:
    月份 | 门店编码 | 产品分类 | 商品名称 | 商品数量 | 应收金额 | 实收金额

模板 R4 標的取數範圍 F-AU / AV-CK / CL-EA 就是寬表的三個度量區塊
(第 6-47 / 48-89 / 90-131 欄), 所以這支腳本做的是把 126 個月份欄
攤成「一列一個月」。

粒度會變: 寬表的 key 是 门店编码×销售渠道×产品分类×商品编码×商品名称,
模板沒有 销售渠道 也沒有 商品编码 —— 同名不同編碼會被加總。
销售渠道 一樣不進 key (列數不變), 但三個總計欄旁邊會再多 6 欄把堂食/外卖
分開列: 商品数量-堂食/外卖、应收金额-堂食/外卖、实收金额-堂食/外卖。
欄數 7 -> 13, 列數不變。--periph 那份不拆, 維持 7 欄。
資料範圍只留 01咖啡类饮品 / 02非咖啡类饮品 (模板 R1 寫明)。

--periph 換成周邊那 8 類, 出成另一個檔。為什麼不併進同一個檔: 兩邊加起來
143 萬列, 一個工作表放不下 (analyse_periph.py 算過)。

輸入是 inspect_files.py 產的 宽表.parquet, 不是 xlsx —— 重跑一次只要幾秒,
不必再花三分鐘解析 105MB 的 xlsx。

用法:
    python build_sku_monthly.py
    python build_sku_monthly.py --root D:\\某資料夾
    python build_sku_monthly.py --periph            # 出周邊那 8 類
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from pathlib import Path

try:
    import duckdb
except ImportError:
    sys.exit("需要 duckdb:  pip install duckdb")
from openpyxl import Workbook, load_workbook

# ================================================================= 設定
PQ = "宽表.parquet"                  # inspect_files.py 的產物
TPL = "SKU template.xlsx"
OUT_XLSX = "SKU月度_成果.xlsx"
OUT_PQ = "SKU月度.parquet"
OUT_CSV = "SKU月度.csv"

MEASURES = ["商品数量", "应收金额", "实收金额"]
GROUP_KEYS = ["月份", "门店编码", "产品分类", "商品名称"]
CATS = ["01咖啡类饮品", "02非咖啡类饮品"]

# 渠道拆分。這兩個值不靠記憶寫死就算數: 跑的時候先 distinct 出來對, 對不上就停,
# 否則多一種渠道 (或空值) 只會讓兩欄加起來比總計少, 而且不會有人發現。
CHANNEL_COL = "销售渠道"
CHANNELS = ["堂食", "外卖"]

# --periph 用的那一組。這 8 個字串是 analyse_periph.py 從寬表 distinct 出來對過的,
# 跟項目組給的名單逐字相符 (15/20/21 是咖啡次卡/汉堡/油炸食品, 不是周邊)。
PERIPH_CATS = ["11餐盒", "12周边产品-杯子碗碟", "13周边产品-服装配饰",
               "14周边产品-包袋挂件", "16周边产品-咖啡器具", "17周边产品-生活用品",
               "18周边产品-其他货品", "19周边产品-特别限定门店"]
PERIPH_XLSX = "SKU月度_周边_results.xlsx"
PERIPH_PQ = "SKU月度_周边.parquet"
PERIPH_CSV = "SKU月度_周边.csv"
PERIPH_LABEL = "、".join(PERIPH_CATS)   # 抬頭裡 01/02 那串字樣要換成這個
TPL_HEAD_ROWS = 5                    # 模板前 5 列是抬頭, 資料從 R6 開始
EXCEL_MAX_ROWS = 1_048_576           # Excel 的硬上限
# 只有真的放不下才切。這個值必須跟「放得下嗎」那段的 cap 一致, 否則
# --dry-run 說放得下、正式跑卻切成兩個工作表。
SHEET_ROWS = EXCEL_MAX_ROWS - TPL_HEAD_ROWS
# ================================================================= 設定結束


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def lit(v: str) -> str:
    return "'" + v.replace("'", "''") + "'"


def relabel(v):
    """模板抬頭把範圍寫死成 01/02 —— 出周邊檔時照抄會變成標題跟內容不符。

    只動那兩個分類名, 其餘一個字都不改; 沒出現就原樣回傳。改了哪一格會印出來,
    不會靜靜地改掉交付檔的抬頭。

    整串一起換而不是逐個換: 客戶模板寫的是「01咖啡类饮品和02非咖啡类饮品」,
    逐個換會變成「周邊...和周邊...」同一句講兩次。所以連接詞要一起吃掉,
    但只吃分類「之間」的, 最後一個分類後面的標點要留著。
    """
    if not isinstance(v, str):
        return v
    one = "|".join(re.escape(c) for c in CATS)
    sep = r"\s*(?:[、,，/／&]|和|与|與|及|跟)\s*"
    return re.sub(rf"(?:{one})(?:{sep}(?:{one}))*", PERIPH_LABEL, v)


def parse_cols(cols: list[str]) -> tuple[dict, list[str]]:
    """寬表欄名長這樣: '"商品数量"按“营业日期”加总|2023-01' -> {度量: {月份: 欄名}}"""
    out: dict[str, dict[str, str]] = {m: {} for m in MEASURES}
    for c in cols:
        if "|" not in c:
            continue
        title, period = c.rsplit("|", 1)
        for m in MEASURES:
            if m in title:
                out[m][period] = c
                break
    periods = sorted(out[MEASURES[0]])
    for m in MEASURES:
        if sorted(out[m]) != periods:
            sys.exit(f"{m} 的月份欄跟 {MEASURES[0]} 對不起來, "
                     f"{len(out[m])} vs {len(periods)} 個")
    if not periods:
        sys.exit(f"{PQ} 裡找不到 '度量|月份' 形式的欄位, 這不是 inspect_files.py 的產物")
    return out, periods


def plan_sheets(per_month: list[tuple[str, int]], cap: int) -> list[list[str]]:
    """依月份邊界切工作表 —— 同一個月不跨表, 對帳時才好一段一段核。"""
    groups: list[list[str]] = []
    cur: list[str] = []
    n = 0
    for p, c in per_month:
        if cur and n + c > cap:
            groups.append(cur)
            cur, n = [], 0
        cur.append(p)
        n += c
    if cur:
        groups.append(cur)
    return groups


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=".", help="資料夾 (預設: 目前目錄)")
    ap.add_argument("--sheet-rows", type=int, default=SHEET_ROWS,
                    help=f"每個工作表最多幾列 (預設 {SHEET_ROWS:,})")
    ap.add_argument("--dry-run", action="store_true",
                    help="只算列數與對帳, 不寫檔 (幾秒)")
    ap.add_argument("--csv", action="store_true",
                    help="另外輸出一份 csv (預設不產, 交付的是 xlsx)")
    ap.add_argument("--periph", action="store_true",
                    help=f"改出周邊那 8 類 -> {PERIPH_XLSX}")
    args = ap.parse_args()

    cats = PERIPH_CATS if args.periph else CATS
    out_xlsx = PERIPH_XLSX if args.periph else OUT_XLSX
    out_pq_name = PERIPH_PQ if args.periph else OUT_PQ
    out_csv_name = PERIPH_CSV if args.periph else OUT_CSV
    # 只有 01/02 那份拆渠道。extra_pairs 是空的時候, 底下一切跟改動前完全一樣。
    extra_pairs = [] if args.periph else [(m, ch) for m in MEASURES for ch in CHANNELS]
    extra = [f"{m}-{ch}" for m, ch in extra_pairs]
    out_measures = MEASURES + extra
    print(f"範圍: {len(cats)} 類 {cats}")
    print(f"度量: {len(out_measures)} 欄" +
          (f" = 總計 {len(MEASURES)} + 渠道拆分 {len(extra)}" if extra else " (不拆渠道)"))

    root = Path(args.root).resolve()
    src_pq, tpl = root / PQ, root / TPL
    if not src_pq.exists():
        sys.exit(f"找不到 {src_pq}  (先跑 inspect_files.py)")
    if not tpl.exists():
        sys.exit(f"找不到 {tpl}")
    t0 = time.time()

    con = duckdb.connect()
    src = f"read_parquet('{src_pq.as_posix()}')"
    cols = [r[0] for r in con.sql(f"DESCRIBE SELECT * FROM {src}").fetchall()]
    cmap, periods = parse_cols(cols)
    print(f"寬表 {len(cols)} 欄 -> {len(MEASURES)} 個度量 × {len(periods)} 個月 "
          f"({periods[0]} ~ {periods[-1]})")

    # ---- 42 段 UNION ALL 攤平。三個度量全空的月份不產生列 ----------------
    # (刻意不用 UNPIVOT: 空值要不要留是這裡的重點, 自己寫 WHERE 才看得見規則)
    in_cats = ", ".join(lit(c) for c in cats)

    # ---- 渠道值先 distinct 出來對過, 不靠上次的結論 ----------------------
    if extra_pairs:
        rows = con.sql(f"SELECT COALESCE({q(CHANNEL_COL)}, '(NULL)'), count(*) "
                       f'FROM {src} WHERE "产品分类" IN ({in_cats}) '
                       f"GROUP BY 1 ORDER BY 2 DESC").fetchall()
        print(f"\n{CHANNEL_COL} 分佈 (只看要出的分類):")
        for ch, n in rows:
            print(f"  {ch}: {n:,} 列")
        found = sorted(ch for ch, _ in rows)
        if found != sorted(CHANNELS):
            sys.exit(f"\n渠道不是預期的 {CHANNELS}, 實際是 {found} —— "
                     f"照現在的寫法, 名單外的列會兩欄都不進, "
                     f"先確認多出來的要歸到哪一邊再跑")

    parts = []
    for p in periods:
        vals = ", ".join(f"{q(cmap[m][p])} AS {q(m)}" for m in MEASURES)
        nn = " OR ".join(f"{q(cmap[m][p])} IS NOT NULL" for m in MEASURES)
        parts.append(f'SELECT "门店编码", "产品分类", "商品名称", {q(CHANNEL_COL)},'
                     f' {lit(p)} AS "月份",'
                     f' {vals} FROM {src}'
                     f' WHERE "产品分类" IN ({in_cats}) AND ({nn})')
    long_sql = "\nUNION ALL\n".join(parts)

    keys = ", ".join(q(k) for k in GROUP_KEYS)
    # 排序要用「寫進檔案的值」而不是原值: 商品名称 是 NULL 的列在 xlsx 裡是空字串,
    # 而 DuckDB 把 NULL 排在最後、空字串排在最前 —— 不 COALESCE 的話, 檔案看起來
    # 就是沒排序的, 交付前的排序檢查會被這件事絆倒。
    order = ", ".join(f"COALESCE({q(k)}, '')" for k in GROUP_KEYS)
    # FILTER 一列都沒中時回傳的是 NULL 不是 0, 外面再包一層 COALESCE ——
    # 不然「這個月這家店沒有外卖」在檔案裡會是空格, 跟總計欄的處理不一致。
    sums = ", ".join(
        [f'sum(COALESCE({q(m)}, 0)) AS {q(m)}' for m in MEASURES]
        + [f'COALESCE(sum(COALESCE({q(m)}, 0)) '
           f'FILTER (WHERE {q(CHANNEL_COL)} = {lit(ch)}), 0) AS {q(f"{m}-{ch}")}'
           for m, ch in extra_pairs])
    out_pq = root / out_pq_name
    print("攤平 + 聚合中...")
    con.execute(f"""
        COPY (
            SELECT {keys}, {sums}
            FROM ({long_sql})
            GROUP BY {", ".join(str(i + 1) for i in range(len(GROUP_KEYS)))}
            ORDER BY {order}
        ) TO '{out_pq.as_posix()}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    res = f"read_parquet('{out_pq.as_posix()}')"

    n_rows = con.sql(f"SELECT count(*) FROM {res}").fetchone()[0]
    if not n_rows:
        sys.exit(f"篩不到任何資料。寬表的 产品分类 裡沒有 {cats} —— "
                 f"先確認分類名稱是否完全一致")
    n_store, n_name, n_blank = con.sql(f"""
        SELECT count(DISTINCT "门店编码"), count(DISTINCT "商品名称"),
               count(*) FILTER (WHERE "商品名称" IS NULL OR "商品名称" = '')
        FROM {res}""").fetchone()
    print(f"\n結果 {n_rows:,} 列   不重複 门店编码 {n_store:,}   商品名称 {n_name:,}")
    if n_blank:
        print(f"  !! 有 {n_blank:,} 列的 商品名称 是空的 (寬表本來就沒解析出名稱)")
    for cat, n in con.sql(f'SELECT "产品分类", count(*) FROM {res} '
                          f"GROUP BY 1 ORDER BY 2 DESC").fetchall():
        print(f"  {cat}: {n:,} 列")

    # ---- 對帳: 逐月逐度量 ---------------------------------------------
    # 來源那邊是「橫向把該月三個欄加起來」, 結果那邊是「縱向把該月的列加起來」,
    # 兩條路徑獨立, 所以月份錯位、漏月、重複計算都會現形。
    exprs = ", ".join(f'sum(COALESCE({q(cmap[m][p])}, 0))'
                      for p in periods for m in MEASURES)
    flat = con.sql(f"SELECT {exprs} FROM {src} "
                   f"WHERE \"产品分类\" IN ({in_cats})").fetchone()
    want = {p: flat[i * len(MEASURES):(i + 1) * len(MEASURES)]
            for i, p in enumerate(periods)}
    got = {r[0]: r[1:] for r in con.sql(
        f'SELECT "月份", {", ".join(f"sum({q(m)})" for m in MEASURES)} '
        f"FROM {res} GROUP BY 1").fetchall()}

    bad = []
    for p in periods:
        g = got.get(p, (0.0,) * len(MEASURES))
        for i, m in enumerate(MEASURES):
            if abs((g[i] or 0) - (want[p][i] or 0)) > 0.005:
                bad.append((p, m, g[i] or 0, want[p][i] or 0))
    print(f"\n=== 對帳: {len(periods)} 個月 × {len(MEASURES)} 個度量 = "
          f"{len(periods) * len(MEASURES)} 項, 不符 {len(bad)} 項 ===")
    for p, m, g, w in bad[:20]:
        print(f"  !! {p} {m}: 結果 {g:,.2f} vs 寬表 {w:,.2f}")
    tot = [sum(want[p][i] for p in periods) for i in range(len(MEASURES))]
    for m, v in zip(MEASURES, tot):
        print(f"  {m} 總計 {v:,.2f}")
    if bad:
        sys.exit("\n對帳失敗, 不要交付這份檔案")

    # ---- 對帳 2: 逐列 堂食 + 外卖 = 總計 ---------------------------------
    # 上面那條驗的是「攤平有沒有漏月」, 這條驗的是「渠道有沒有漏值」: 只要有
    # 第三種渠道或空值, 兩欄加起來就會比總計少, 這裡當場現形。
    if extra_pairs:
        chk = " OR ".join(
            f'abs({q(m)} - ' + " - ".join(q(f"{m}-{ch}") for ch in CHANNELS)
            + ") > 0.005" for m in MEASURES)
        n_ch_bad = con.sql(f"SELECT count(*) FROM {res} WHERE {chk}").fetchone()[0]
        print(f"\n=== 對帳: {' + '.join(CHANNELS)} = 總計, "
              f"{n_rows:,} 列裡不符 {n_ch_bad:,} 列 ===")
        if n_ch_bad:
            for r in con.sql(f"SELECT {keys}, "
                             f'{", ".join(q(m) for m in out_measures)} '
                             f"FROM {res} WHERE {chk} LIMIT 5").fetchall():
                print(f"  !! {r}")
            sys.exit("\n渠道拆分加不回總計, 不要交付這份檔案")
        tot_ch = con.sql(
            f'SELECT {", ".join(f"sum({q(c)})" for c in extra)} FROM {res}'
        ).fetchone()
        for c, v in zip(extra, tot_ch):
            print(f"  {c} 總計 {v:,.2f}")

    # ---- 放不放得進一個工作表 -------------------------------------------
    cap = EXCEL_MAX_ROWS - TPL_HEAD_ROWS
    print(f"\n=== 列數 vs Excel 上限 ===")
    print(f"  結果 {n_rows:,} 列   單一工作表扣掉 {TPL_HEAD_ROWS} 列抬頭可放 {cap:,} 列")
    if n_rows <= cap:
        print(f"  -> 放得下, 不用切 (還剩 {cap - n_rows:,} 列餘裕)")
    else:
        print(f"  -> 超出 {n_rows - cap:,} 列, 一個工作表放不下")
    for y, c in con.sql(f'SELECT left("月份", 4), count(*) FROM {res} '
                        f"GROUP BY 1 ORDER BY 1").fetchall():
        print(f"    {y} 年: {c:,} 列")

    if args.dry_run:
        print(f"\n(--dry-run: 只算了數, 沒寫 xlsx/csv。中間檔 {out_pq_name} 留著,"
              f" 正式跑的時候不用重算)")
        return

    # ---- 寫 xlsx (照模板抬頭), 太多列就按月份切工作表 --------------------
    n_col = len(GROUP_KEYS) + len(out_measures)
    wt = load_workbook(tpl, data_only=True)
    wst = wt["SKU 月度"] if "SKU 月度" in wt.sheetnames else wt.worksheets[0]
    head = [list(r) for r in wst.iter_rows(min_row=1, max_row=TPL_HEAD_ROWS,
                                           max_col=n_col, values_only=True)]
    head = [r + [None] * (n_col - len(r)) for r in head]   # 模板短就補空格
    tpl_title = wst.title
    wt.close()

    # 模板的抬頭只到第 7 欄, 新增那 6 欄要自己補標題。找「第一格是 月份」的那一列,
    # 不寫死 R5 —— 模板改版時抬頭列數會變, 寫死會把標題填到別的地方。
    if extra:
        i = next((i for i, r in enumerate(head)
                  if str(r[0]).strip() == GROUP_KEYS[0]), None)
        if i is None:
            print(f"  !! 抬頭裡找不到欄位名那一列 (第一格是 {GROUP_KEYS[0]!r} 的), "
                  f"新增的 {len(extra)} 欄沒有標題, 交付前要人工補")
        else:
            for j, name in enumerate(extra):
                head[i][len(GROUP_KEYS) + len(MEASURES) + j] = name
            print(f"  抬頭 R{i + 1} 補上 {len(extra)} 個欄位名: {', '.join(extra)}")

    if args.periph:
        # 抬頭寫的是 01/02 的範圍, 內容卻是周邊 —— 這種檔交出去會被當成錯的
        n_fix = 0
        for i, r in enumerate(head):
            for j, v in enumerate(r):
                nv = relabel(v)
                if nv != v:
                    print(f"  抬頭 R{i + 1}C{j + 1} 改寫: {v!r}\n"
                          f"                 -> {nv!r}")
                    if nv.count(PERIPH_LABEL) > 1:
                        # 連接詞沒吃乾淨的話會變成同一句講兩次, 出過一次這個包
                        print("  !! 這一格的範圍字樣出現了 "
                              f"{nv.count(PERIPH_LABEL)} 次, 抬頭要人工看過再送")
                    r[j] = nv
                    n_fix += 1
        print(f"  模板抬頭改了 {n_fix} 格"
              if n_fix else
              "  !! 模板抬頭沒有出現 01/02 的字樣, 沒有改動 —— "
              "抬頭若另有寫死的範圍說明, 要人工確認")

    per_month = con.sql(f'SELECT "月份", count(*) FROM {res} '
                        f"GROUP BY 1 ORDER BY 1").fetchall()
    groups = plan_sheets(per_month, args.sheet_rows)
    if len(groups) > 1:
        print(f"\n{n_rows:,} 列放不進一個工作表 (上限 {args.sheet_rows:,}), "
              f"按月份切成 {len(groups)} 個:")

    cur = con.execute(f'SELECT {keys}, {", ".join(q(m) for m in out_measures)} '
                      f"FROM {res} ORDER BY {order}")
    wb = Workbook(write_only=True)
    written = 0
    for g in groups:
        title = tpl_title if len(groups) == 1 else f"{g[0]}~{g[-1]}"
        ws = wb.create_sheet(title[:31])
        for r in head:
            ws.append(r)
        n = sum(c for p, c in per_month if p in set(g))
        left = n
        while left > 0:
            batch = cur.fetchmany(min(50_000, left))
            if not batch:
                break
            for row in batch:
                ws.append(["" if v is None else v for v in row])
            left -= len(batch)
        written += n - left
        if len(groups) > 1:
            print(f"  [{title}]  {n:,} 列")
    assert written == n_rows, f"寫出 {written} 列, 應該是 {n_rows}"

    out = root / out_xlsx
    print("\n存檔中 (大檔要一兩分鐘)...")
    wb.save(out)
    if args.csv:
        # 自己寫而不用 COPY: utf-8-sig 的 BOM 是 Excel 雙擊開檔不亂碼的關鍵
        cur2 = con.execute(f'SELECT {keys}, {", ".join(q(m) for m in out_measures)} '
                           f"FROM {res} ORDER BY {order}")
        with open(root / out_csv_name, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(GROUP_KEYS + out_measures)
            while True:
                batch = cur2.fetchmany(50_000)
                if not batch:
                    break
                w.writerows(batch)

    print(f"\n輸出:")
    print(f"  {out_xlsx}  {out.stat().st_size / 1e6:,.1f} MB  "
          f"{n_rows:,} 列 / {len(groups)} 個工作表   <- 交付這個")
    if args.csv:
        print(f"  {out_csv_name}   "
              f"{(root / out_csv_name).stat().st_size / 1e6:,.1f} MB")
    print(f"  {out_pq_name}  (中間檔)")
    print(f"\n完成, 共 {time.time() - t0:.0f}s")


if __name__ == "__main__":
    main()
