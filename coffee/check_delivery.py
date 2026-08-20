#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
送出前的最後一道核對: 拿成果 xlsx 直接跟原始寬表 xlsx 對, 並看資料品質。

為什麼不共用 build_sku_monthly.py 的程式碼: 這支的價值就在於獨立。
共用常數或函式的話, 兩邊會一起錯而且對帳照樣過, 那就白驗了。所以這裡
自己重讀兩個 xlsx、自己重算, 唯一相同的只有「答案應該是什麼」。

驗四件事:
  1. 成果檔真的寫進去了 —— 讀回來重算列數與合計
  2. 逐月逐度量 vs 原始寬表 (獨立路徑, 抓得到月份錯位)
  3. 資料品質 —— 負數、重複 key、空名稱、值域、实收>应收
  4. 逐月核算表, 印出來給人眼睛掃

用法:
    python check_delivery.py
    python check_delivery.py --skip-source     # 跳過重掃原始 xlsx (省 3 分鐘)
    python check_delivery.py --periph          # 改核對周邊那份成果檔
"""
from __future__ import annotations

import argparse
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# ================================================================= 設定
SRC = "聚合表_result.xlsx"           # 原始寬表
OUT = "SKU月度_成果.xlsx"            # 要送出的成果
REPORT = "交付核对.txt"

CATS = ["01咖啡类饮品", "02非咖啡类饮品"]

# --periph 用的那一組。這裡刻意再抄一次而不是 import build_sku_monthly:
# 兩邊各自寫死, 名單打錯才會在「混進了不該有的分類」那關現形。
PERIPH_OUT = "SKU月度_周边_results.xlsx"
PERIPH_REPORT = "交付核对_周边.txt"
PERIPH_CATS = ["11餐盒", "12周边产品-杯子碗碟", "13周边产品-服装配饰",
               "14周边产品-包袋挂件", "16周边产品-咖啡器具", "17周边产品-生活用品",
               "18周边产品-其他货品", "19周边产品-特别限定门店"]
MEASURES = ["商品数量", "应收金额", "实收金额"]
# 01/02 那份在三個總計欄後面還有 6 欄渠道拆分 (H-M 欄)。順序是每個度量的
# 堂食 在前、外卖 在後。跟 build_sku_monthly.py 一樣各自寫死不共用 ——
# 共用的話兩邊會一起錯而且核帳照樣過。周邊那份沒有這 6 欄。
CHANNELS = ["堂食", "外卖"]
CH_COLS = [f"{m}-{ch}" for m in MEASURES for ch in CHANNELS]
SRC_HEAD_ROWS = 3                    # 寬表: 前 3 列是抬頭
SRC_N_KEYS = 5                       # 寬表: 门店编码/销售渠道/产品分类/商品编码/商品名称
SRC_CH_COL = 1                       # 寬表: 销售渠道 在第 2 欄 (0-based 1)
SRC_CAT_COL = 2                      # 寬表: 产品分类 在第 3 欄 (0-based 2)
OUT_HEAD_ROWS = 5                    # 成果: 前 5 列是模板抬頭
OUT_N_KEYS = 4                       # 成果: 月份/门店编码/产品分类/商品名称
TOL = 0.005
# ================================================================= 設定結束

L: list[str] = []


def add(s: str = "") -> None:
    L.append(s)
    print(s, flush=True)


def num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def pad(t: str, w: int) -> str:
    n = sum(2 if ord(c) > 0x2E80 else 1 for c in t)
    return t + " " * max(0, w - n)


# --------------------------------------------------- 1. 讀成果檔
def read_out(path: Path, split: bool) -> dict:
    """串流讀成果檔, 順便做品質檢查。資料已排序, 所以重複 key 只要比相鄰列。

    split=True 時多讀 H-M 那 6 欄渠道拆分, 並逐列驗 堂食 + 外卖 = 總計。
    """
    per: dict = defaultdict(lambda: [0.0, 0.0, 0.0])
    per_ch: dict = defaultdict(lambda: [0.0] * len(CH_COLS))
    ch_bad = 0                       # 堂食 + 外卖 加不回總計的列數
    n_rows = 0
    cats: Counter = Counter()
    stores = set()
    blank_name = 0
    blank_sum = [0.0, 0.0, 0.0]
    neg = [0, 0, 0]
    dup = 0
    unsorted_n = 0
    over = 0            # 实收 > 应收
    zero_qty_amt = 0    # 数量 0 但金额不是 0
    prev_key = None
    sheets = []
    t0 = time.time()

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        n_sheet = 0
        for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= OUT_HEAD_ROWS:
                continue
            if r is None or all(v is None for v in r[:4]):
                continue
            month, store, cat, name = (txt(r[0]), txt(r[1]), txt(r[2]), txt(r[3]))
            vals = [num(r[4]) or 0.0, num(r[5]) or 0.0, num(r[6]) or 0.0]
            n_rows += 1
            n_sheet += 1
            cats[cat] += 1
            stores.add(store)
            if not name:
                blank_name += 1
                for j in range(3):
                    blank_sum[j] += vals[j]
            for j in range(3):
                if vals[j] < 0:
                    neg[j] += 1
            if vals[2] > vals[1] + TOL:
                over += 1
            if abs(vals[0]) < 1e-9 and (abs(vals[1]) > TOL or abs(vals[2]) > TOL):
                zero_qty_amt += 1
            key = (month, store, cat, name)
            if key == prev_key:
                dup += 1
            elif prev_key is not None and key < prev_key:
                # 重複 key 是比相鄰列驗出來的, 沒排序的話那個檢查就不完整
                unsorted_n += 1
            prev_key = key
            t = per[month]
            for j in range(3):
                t[j] += vals[j]
            if split:
                cv = []
                for k in range(len(CH_COLS)):
                    c = OUT_N_KEYS + len(MEASURES) + k
                    cv.append(num(r[c]) or 0.0 if c < len(r) else 0.0)
                tc = per_ch[month]
                for k in range(len(CH_COLS)):
                    tc[k] += cv[k]
                nc = len(CHANNELS)
                for j in range(len(MEASURES)):
                    if abs(vals[j] - sum(cv[j * nc:(j + 1) * nc])) > TOL:
                        ch_bad += 1
                        break
            if n_rows % 200_000 == 0:
                print(f"    已讀 {n_rows:,} 列  ({time.time() - t0:.0f}s)",
                      file=sys.stderr, flush=True)
        sheets.append((ws.title, n_sheet))
    wb.close()
    return {"per": per, "n_rows": n_rows, "cats": cats, "stores": stores,
            "blank_name": blank_name, "blank_sum": blank_sum, "neg": neg, "dup": dup, "over": over,
            "zero_qty_amt": zero_qty_amt, "sheets": sheets,
            "unsorted": unsorted_n, "per_ch": per_ch, "ch_bad": ch_bad,
            "secs": time.time() - t0}


# --------------------------------------------------- 2. 重掃原始寬表
def read_src(path: Path, cats: list[str], split: bool) -> dict:
    """只累加指定的分類, 回傳 {月份: [三個度量]}。表頭的月份自己重讀一次。

    split=True 時另外按 销售渠道 分開累加一份, 用來獨立驗成果檔的 H-M 欄。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    head = [next(rows) for _ in range(SRC_HEAD_ROWS)]
    hdr = [txt(c) for c in head[SRC_HEAD_ROWS - 1]]
    while hdr and not hdr[-1]:
        hdr.pop()
    periods = hdr[SRC_N_KEYS:]
    n_p = len(periods) // len(MEASURES)
    if n_p * len(MEASURES) != len(periods):
        sys.exit(f"原始寬表的期間欄數 {len(periods)} 切不成 {len(MEASURES)} 個區塊")
    months = periods[:n_p]

    per: dict = defaultdict(lambda: [0.0, 0.0, 0.0])
    per_ch: dict = defaultdict(lambda: [0.0] * len(CH_COLS))
    ch_other: Counter = Counter()    # 名單外的渠道 (含空白) 各幾列
    ch_idx = {c: i for i, c in enumerate(CHANNELS)}
    nc = len(CHANNELS)
    n_rows = 0
    t0 = time.time()
    want = set(cats)
    for r in rows:
        if len(r) <= SRC_CAT_COL or txt(r[SRC_CAT_COL]) not in want:
            continue
        n_rows += 1
        ci = None
        if split:
            ch = txt(r[SRC_CH_COL]) if len(r) > SRC_CH_COL else ""
            ci = ch_idx.get(ch)
            if ci is None:
                ch_other[ch or "(空白)"] += 1
        for j in range(len(MEASURES)):
            base = SRC_N_KEYS + j * n_p
            for k in range(n_p):
                v = r[base + k] if base + k < len(r) else None
                if v is not None:
                    x = num(v)
                    if x:
                        per[months[k]][j] += x
                        if ci is not None:
                            per_ch[months[k]][j * nc + ci] += x
        if n_rows % 50_000 == 0:
            print(f"    已掃 {n_rows:,} 列  ({time.time() - t0:.0f}s)",
                  file=sys.stderr, flush=True)
    wb.close()
    return {"per": per, "n_rows": n_rows, "months": months,
            "per_ch": per_ch, "ch_other": ch_other,
            "secs": time.time() - t0}


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=".")
    ap.add_argument("--skip-source", action="store_true",
                    help="跳過重掃原始寬表 (省時間, 但就少了獨立驗證)")
    ap.add_argument("--periph", action="store_true",
                    help=f"改核對 {PERIPH_OUT}")
    args = ap.parse_args()

    out_name = PERIPH_OUT if args.periph else OUT
    cats = PERIPH_CATS if args.periph else CATS
    report = PERIPH_REPORT if args.periph else REPORT
    split = not args.periph          # 周邊那份沒有渠道拆分欄

    root = Path(args.root).resolve()
    out_p, src_p = root / out_name, root / SRC
    if not out_p.exists():
        sys.exit(f"找不到 {out_p}")
    ok = True

    add(f"核對時間 {datetime.now():%Y-%m-%d %H:%M}")
    add(f"成果檔 {out_name}  {out_p.stat().st_size / 1e6:,.1f} MB")
    add(f"範圍 {len(cats)} 類 {cats}")

    add("")
    add("=== 1. 成果檔讀回 ===")
    print("  讀成果檔中...", file=sys.stderr, flush=True)
    o = read_out(out_p, split)
    add(f"  資料列數 {o['n_rows']:,}   工作表 {len(o['sheets'])} 個   "
        f"({o['secs']:.0f}s)")
    for name, n in o["sheets"]:
        add(f"    [{name}] {n:,} 列")
    tot = [sum(v[j] for v in o["per"].values()) for j in range(3)]
    for m, v in zip(MEASURES, tot):
        add(f"  {m} 合計 {v:,.2f}")
    add(f"  月份 {len(o['per'])} 個   门店编码 {len(o['stores']):,} 個")

    add("")
    add("=== 2. 對原始寬表獨立核帳 ===")
    s = None
    if args.skip_source:
        add("  (--skip-source: 沒做。這是唯一能抓出「讀寬表就讀錯了」的檢查)")
    elif not src_p.exists():
        add(f"  !! 找不到 {SRC}, 沒辦法做獨立核帳")
        ok = False
    else:
        print("  重掃原始寬表中 (幾分鐘)...", file=sys.stderr, flush=True)
        s = read_src(src_p, cats, split)
        add(f"  原始寬表裡 {cats} 共 {s['n_rows']:,} 列  ({s['secs']:.0f}s)")
        bad = []
        for mo in s["months"]:
            w = s["per"].get(mo, [0.0, 0.0, 0.0])
            g = o["per"].get(mo, [0.0, 0.0, 0.0])
            for j in range(3):
                if abs(g[j] - w[j]) > TOL:
                    bad.append((mo, MEASURES[j], g[j], w[j]))
        extra = set(o["per"]) - set(s["months"])
        if extra:
            ok = False
            add(f"  !! 成果檔有原始寬表沒有的月份: {sorted(extra)}")
        add(f"  {len(s['months'])} 個月 × {len(MEASURES)} 個度量 = "
            f"{len(s['months']) * len(MEASURES)} 項, 不符 {len(bad)} 項")
        for mo, m, g, w in bad[:20]:
            add(f"    !! {mo} {m}: 成果 {g:,.2f} vs 原始 {w:,.2f}  "
                f"差 {g - w:,.2f}")
        if bad:
            ok = False

    # ------------------------------------------------- 2b. 渠道拆分
    if split:
        add("")
        add(f"=== 2b. 渠道拆分核對 ({'/'.join(CHANNELS)}) ===")
        # 第一關只用成果檔自己: 拆分加不回總計就是錯的, 不必看原始寬表
        add(f"  成果檔內部: {' + '.join(CHANNELS)} = 總計, "
            f"{o['n_rows']:,} 列裡不符 {o['ch_bad']:,} 列")
        if o["ch_bad"]:
            ok = False
        if s is None:
            add("  (沒重掃原始寬表, 少了「拆分值本身對不對」的獨立驗證)")
        else:
            if s["ch_other"]:
                # 名單外的渠道不會進任何一個拆分欄, 但總計欄有 —— 兩邊就對不上
                ok = False
                add(f"  !! 原始寬表裡有 {sum(s['ch_other'].values()):,} 列的 "
                    f"销售渠道 不在 {CHANNELS} 裡:")
                for ch, n in s["ch_other"].most_common(10):
                    add(f"       {ch}: {n:,} 列")
            cbad = []
            for mo in s["months"]:
                w = s["per_ch"].get(mo, [0.0] * len(CH_COLS))
                g = o["per_ch"].get(mo, [0.0] * len(CH_COLS))
                for k, cname in enumerate(CH_COLS):
                    if abs(g[k] - w[k]) > TOL:
                        cbad.append((mo, cname, g[k], w[k]))
            add(f"  對原始寬表: {len(s['months'])} 個月 × {len(MEASURES)} 個度量 "
                f"× {len(CHANNELS)} 個渠道 = "
                f"{len(s['months']) * len(CH_COLS)} 項, 不符 {len(cbad)} 項")
            for mo, cname, g, w in cbad[:20]:
                add(f"    !! {mo} {cname}: 成果 {g:,.2f} vs 原始 {w:,.2f}  "
                    f"差 {g - w:,.2f}")
            if cbad:
                ok = False
        for k, cname in enumerate(CH_COLS):
            add(f"  {cname} 合計 {sum(v[k] for v in o['per_ch'].values()):,.2f}")

    add("")
    add("=== 3. 資料品質 ===")
    add(f"  产品分类值域: {dict(o['cats'])}")
    stray = [c for c in o["cats"] if c not in cats]
    if stray:
        ok = False
        add(f"  !! 混進了不該有的分類: {stray}")
    else:
        add(f"  -> 只有 {cats}, 正確")
    missing = [c for c in cats if c not in o["cats"]]
    if missing:
        add(f"  !! 這些分類在成果檔裡一列都沒有: {missing}")
    if o["unsorted"]:
        ok = False
        add(f"  !! 有 {o['unsorted']:,} 處排序不對 —— 重複 key 的檢查會不完整")
    else:
        add("  排序正確 (月份->门店->分类->名称)")
    add(f"  重複 key (月份+门店+分类+名称): {o['dup']:,}"
        + ("   <- 應該是 0" if o["dup"] else "   (0, 正確)"))
    if o["dup"]:
        ok = False
    if o["blank_name"]:
        # 光看列數判斷不了重不重要, 要看它佔多少金額
        add(f"  商品名称 空白: {o['blank_name']:,} 列 "
            f"({o['blank_name'] / o['n_rows'] * 100:.4f}%)"
            f"   <- 寬表本來就沒解析出名稱, 不是這一步造成的")
        for m, v, t in zip(MEASURES, o["blank_sum"], tot):
            add(f"    其中 {m} {v:,.2f}"
                + (f"  占全部的 {v / t * 100:.4f}%" if t else ""))
    else:
        add("  商品名称 空白: 0 列")
    for m, n in zip(MEASURES, o["neg"]):
        add(f"  {m} 為負: {n:,} 列" + ("   (退貨/沖銷, 合理但要知道)" if n else ""))
    add(f"  实收金额 > 应收金额: {o['over']:,} 列"
        + ("   <- 一般會有優惠所以實收較低, 這些值得抽看" if o["over"] else ""))
    add(f"  商品数量 為 0 但金額不為 0: {o['zero_qty_amt']:,} 列")

    add("")
    add("=== 4. 逐月核算表 (拿去跟客戶月報並排比) ===")
    add(f"  {pad('月份', 9)}{'商品数量':>16}{'应收金额':>20}{'实收金额':>20}")
    for mo in sorted(o["per"]):
        v = o["per"][mo]
        add(f"  {pad(mo, 9)}{v[0]:>16,.2f}{v[1]:>20,.2f}{v[2]:>20,.2f}")
    add(f"  {pad('合計', 9)}{tot[0]:>16,.2f}{tot[1]:>20,.2f}{tot[2]:>20,.2f}")

    add("")
    add(f"=== 結論: {'可以送出' if ok else '!! 有問題, 先不要送出'} ===")
    (root / report).write_text("\n".join(L), encoding="utf-8")
    print(f"\n報告: {root / report}", file=sys.stderr)
    if not ok:
        sys.exit(1)


if __name__ == "__main__":
    main()
