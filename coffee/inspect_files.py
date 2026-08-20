#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
勘查一個資料夾裡的 xlsx: 認出大寬表的來歷與結構, 並把小模板檔整個攤開。

一趟回答四件事:
  1. 每個 xlsx 是誰產的 —— docProps 指紋 (openpyxl 產的 vs Excel 存過的)
  2. 大寬表的結構是不是 write_template.py 的輸出 (5 個 key 欄 + 3 個度量區塊 × N 個月)
  3. 大寬表的产品分类分佈: 每類幾列、三個度量各多少 —— 決定後續篩選範圍
  4. 小模板檔每一格寫了什麼 (含合併儲存格、註解、填色標記)

第 3 步要完整解析一次 xlsx (慢)。既然成本已經付了, 同一趟順便把寬表寫成
parquet, 之後任何重排都是秒級, 不必再解析第二次。結構認不出來就不寫,
免得產出一個對不上的檔案。

用法:
    python inspect_files.py                    # 掃目前目錄
    python inspect_files.py --root D:\\某資料夾
    python inspect_files.py --no-scan          # 只看結構, 跳過全掃 (快)
    python inspect_files.py --no-parquet       # 全掃但不寫 parquet
    python inspect_files.py --rows 50000       # 全掃只讀前 N 列 (試跑用)
"""
from __future__ import annotations

import argparse
import re
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("需要 openpyxl:  pip install openpyxl")

# ================================================================= 設定
# 照 write_template.py 的輸出格式。認不出來就會據實回報, 不會硬套。
KEYS = ["门店编码", "销售渠道", "产品分类", "商品编码", "商品名称"]
MEASURES = ["商品数量", "应收金额", "实收金额"]
CATEGORY_COL = "产品分类"
BIG_MB = 5.0                      # 超過當寬表, 以下當模板
MAX_TPL_ROWS, MAX_TPL_COLS = 200, 60
MAX_CAT_ROWS = 40                 # 分類分佈最多印幾類
PQ_NAME = "宽表.parquet"
REPORT = "inspect_report.txt"
# ================================================================= 設定結束

L: list[str] = []


def add(s: str = "") -> None:
    L.append(s)
    print(s, flush=True)


def s(v) -> str:
    """儲存格值轉成好比對的字串。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def pad(t: str, w: int) -> str:
    """中文字佔兩格, ljust 會對不齊, 自己算顯示寬度。"""
    n = sum(2 if ord(ch) > 0x2E80 else 1 for ch in t)
    return t + " " * max(0, w - n)


def width(t: str) -> int:
    return sum(2 if ord(ch) > 0x2E80 else 1 for ch in t)


def to_number(v):
    """數值欄: 轉不動就回 None (別回 0, 0 跟「沒值」在加總後分不出來)。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


# ------------------------------------------------------------------ 指紋
def props(path: Path) -> dict:
    """從 zip 直接讀 docProps —— 不必載入整個活頁簿, 105MB 的檔也是瞬間。"""
    out: dict[str, str] = {}
    want = {
        "docProps/core.xml": ("dc:creator", "cp:lastModifiedBy",
                              "dcterms:created", "dcterms:modified"),
        "docProps/app.xml": ("Application", "AppVersion", "Company"),
    }
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        for part, tags in want.items():
            if part not in names:
                continue
            xml = z.read(part).decode("utf-8", "replace")
            for tag in tags:
                m = re.search(rf"<{tag}[^>]*>(.*?)</{tag}>", xml, re.S)
                if m:
                    out[tag.split(":")[-1]] = m.group(1).strip()
        out["_sheets"] = str(sum(1 for n in names
                                 if n.startswith("xl/worksheets/sheet")))
        out["_styles"] = "有" if "xl/styles.xml" in names else "無"
    return out


def show_props(path: Path) -> None:
    p = props(path)
    st = path.stat()
    add(f"  檔案大小 {st.st_size / 1e6:,.1f} MB   "
        f"檔案時間 {datetime.fromtimestamp(st.st_mtime):%Y-%m-%d %H:%M}")
    creator = p.get("creator", "(無)")
    app = p.get("Application", "(無)")
    add(f"  產生者 creator={creator}   Application={app}   "
        f"lastModifiedBy={p.get('lastModifiedBy', '(無)')}")
    add(f"  建立時間 {p.get('created', '(無)')}   修改時間 {p.get('modified', '(無)')}")
    add(f"  工作表數 {p['_sheets']}   樣式表 {p['_styles']}")
    if creator == "openpyxl":
        add("  -> creator 是 openpyxl: 由 Python 腳本產生, 之後沒被 Excel 另存過")
    elif app.startswith("Microsoft Excel"):
        add("  -> Application 是 Excel: 這個檔被 Excel 開過並存檔 (或本來就是 Excel 產的)")


# ------------------------------------------------------------- 寬表結構
def head_rows(path: Path, n: int = 4) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=n, values_only=True):
        rows.append([s(c) for c in r])
    wb.close()
    return rows


def analyse_wide(rows: list[list[str]]) -> dict:
    """從前幾列推出寬表結構, 並跟 write_template.py 的輸出格式比對。"""
    r = {"ok": False, "notes": []}
    if not rows:
        r["notes"].append("讀不到任何列")
        return r

    hdr_i = max(range(len(rows)), key=lambda i: sum(1 for c in rows[i] if c))
    hdr = rows[hdr_i]
    while hdr and not hdr[-1]:
        hdr.pop()
    r["hdr_row"] = hdr_i + 1          # 1-based
    r["n_cols"] = len(hdr)
    r["keys"] = hdr[:len(KEYS)]
    r["keys_ok"] = hdr[:len(KEYS)] == KEYS

    periods = hdr[len(KEYS):]
    if not periods:
        r["notes"].append(f"key 欄之後沒有其他欄位 (共 {len(hdr)} 欄), 這不是寬表")
        return r
    if any(not p for p in periods):
        r["notes"].append(f"期間欄裡有 {sum(1 for p in periods if not p)} 個空表頭")
    if len(periods) % len(MEASURES) != 0:
        r["notes"].append(f"期間欄數 {len(periods)} 不是 {len(MEASURES)} 的倍數, "
                          f"切不成度量區塊")
        return r

    n_p = len(periods) // len(MEASURES)
    blocks = [periods[i * n_p:(i + 1) * n_p] for i in range(len(MEASURES))]
    same = all(b == blocks[0] for b in blocks)
    if not same:
        r["notes"].append("三個區塊的月份序列不一致 —— 不是本 repo 產的格式")
    r["n_periods"] = n_p
    r["periods"] = blocks[0]
    r["blocks_same"] = same

    # 區塊標題在表頭上一列, 每段第一格
    titles = []
    if hdr_i > 0:
        above = rows[hdr_i - 1]
        for i in range(len(MEASURES)):
            j = len(KEYS) + i * n_p
            titles.append(above[j] if j < len(above) else "")
    r["titles"] = titles
    r["titles_ok"] = [t or MEASURES[i] for i, t in enumerate(titles)]
    r["ok"] = r["keys_ok"] and same and n_p > 0
    return r


def report_wide(path: Path, rows: list[list[str]], r: dict) -> None:
    add(f"  表頭在第 {r.get('hdr_row', '?')} 列, 共 {r.get('n_cols', 0)} 欄")
    add(f"  前 {len(KEYS)} 欄: {r.get('keys')}")
    add(f"       預期: {KEYS}   -> {'相同' if r.get('keys_ok') else '!! 不同'}")
    if r.get("periods"):
        ps = r["periods"]
        add(f"  度量區塊 {len(MEASURES)} 個 × {r['n_periods']} 個期間 = "
            f"{r['n_periods'] * len(MEASURES)} 欄")
        add(f"  區塊標題: {r['titles']}")
        add(f"  期間範圍: {ps[0]} ~ {ps[-1]}   (前 3 個 {ps[:3]})")
    for n in r["notes"]:
        add(f"  !! {n}")
    add("")
    add(f"  === 結論: {'吻合 write_template.py 的輸出格式' if r['ok'] else '不吻合, 需人工判讀'} ===")
    if not r["ok"]:
        # 認不出來就把前幾列攤開, 至少能看出這到底是什麼表
        add("  前 4 列 × 前 15 欄:")
        for i, row in enumerate(rows, 1):
            cs = [f"{get_column_letter(j + 1)}={v}"
                  for j, v in enumerate(row[:15]) if v]
            add(f"    R{i} " + (" | ".join(cs) if cs else "(空)"))


# ------------------------------------------------------------- 全掃 + parquet
class Sink:
    """逐批寫 parquet, 記憶體維持固定。"""

    def __init__(self, path: Path, names: list[str], n_keys: int, batch: int = 100_000):
        import pyarrow as pa
        import pyarrow.parquet as pq
        self.pa = pa
        self.schema = pa.schema([(n, pa.string()) for n in names[:n_keys]]
                                + [(n, pa.float64()) for n in names[n_keys:]])
        try:
            self.w = pq.ParquetWriter(path, self.schema, compression="zstd")
            self.codec = "zstd"
        except Exception:
            self.w = pq.ParquetWriter(path, self.schema, compression="snappy")
            self.codec = "snappy"
        self.n_keys, self.batch, self.n = n_keys, batch, 0
        self.cols: list[list] = [[] for _ in names]
        self.path = path

    def add(self, keys, vals) -> None:
        for i in range(self.n_keys):
            self.cols[i].append(keys[i] or None)
        for i, v in enumerate(vals):
            self.cols[self.n_keys + i].append(v)
        self.n += 1
        if self.n % self.batch == 0:
            self.flush()

    def flush(self) -> None:
        if not self.cols[0]:
            return
        tbl = self.pa.table(
            {f.name: self.pa.array(c, type=f.type)
             for f, c in zip(self.schema, self.cols)}, schema=self.schema)
        self.w.write_table(tbl)
        for c in self.cols:
            c.clear()

    def close(self) -> None:
        self.flush()
        self.w.close()


def pq_names(r: dict) -> list[str]:
    """期間欄名在三個區塊裡重複, 加區塊前綴才能當 parquet 欄名。"""
    names = list(KEYS)
    for i, title in enumerate(r["titles_ok"]):
        for p in r["periods"]:
            names.append(f"{title}|{p}")
    seen: Counter = Counter()
    out = []
    for n in names:
        seen[n] += 1
        out.append(n if seen[n] == 1 else f"{n}#{seen[n]}")
    return out


def scan_wide(path: Path, r: dict, limit: int | None, sink: Sink | None) -> None:
    n_keys, n_p = len(KEYS), r["n_periods"]
    first_data = r["hdr_row"] + 1
    cat_i = KEYS.index(CATEGORY_COL)

    cnt: Counter = Counter()                                   # 分類 -> 列數
    sums: dict = defaultdict(lambda: [0.0] * len(MEASURES))    # 分類 -> 三個度量
    channels: Counter = Counter()
    stores, codes = set(), set()
    blank_keys = 0
    n_rows = 0
    t0 = time.time()

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=first_data, values_only=True):
        keys = [s(v) for v in row[:n_keys]]
        if not any(keys):
            continue
        n_rows += 1
        vals = [to_number(v) for v in row[n_keys:n_keys + n_p * len(MEASURES)]]
        if len(vals) < n_p * len(MEASURES):
            vals += [None] * (n_p * len(MEASURES) - len(vals))

        cat = keys[cat_i] or "(空白)"
        cnt[cat] += 1
        tot = sums[cat]
        for i in range(len(MEASURES)):
            seg = vals[i * n_p:(i + 1) * n_p]
            tot[i] += sum(v for v in seg if v is not None)

        channels[keys[1] or "(空白)"] += 1
        stores.add(keys[0])
        codes.add(keys[3])
        if not all(keys):
            blank_keys += 1
        if sink is not None:
            sink.add(keys, vals)

        if n_rows % 50_000 == 0:
            print(f"    已讀 {n_rows:,} 列  ({time.time() - t0:.0f}s)",
                  file=sys.stderr, flush=True)
        if limit and n_rows >= limit:
            break
    wb.close()
    if sink is not None:
        sink.close()

    add(f"  資料列數 {n_rows:,}" + (f"  (--rows 限制在前 {limit:,} 列)" if limit else "")
        + f"   ({time.time() - t0:.0f}s)")
    add(f"  不重複 门店编码 {len(stores):,}   商品编码 {len(codes):,}   "
        f"销售渠道 {len(channels):,}")
    if blank_keys:
        add(f"  !! 有 {blank_keys:,} 列的 key 欄有空值")

    add("")
    add(f"  === {CATEGORY_COL} 分佈 (寬表全域合計) ===")
    w = max([width(c) for c in cnt] + [8])
    add(f"  {pad('分类', w)}  {'行数':>10}  {MEASURES[0]:>16}  "
        f"{MEASURES[1]:>18}  {MEASURES[2]:>18}")
    for cat, n in cnt.most_common(MAX_CAT_ROWS):
        t = sums[cat]
        add(f"  {pad(cat, w)}  {n:>12,}  {t[0]:>18,.2f}  {t[1]:>20,.2f}  {t[2]:>20,.2f}")
    if len(cnt) > MAX_CAT_ROWS:
        rest = sum(n for _, n in cnt.most_common()[MAX_CAT_ROWS:])
        add(f"  ...另外 {len(cnt) - MAX_CAT_ROWS} 個分類, 合計 {rest:,} 列")
    add(f"  分類總數 {len(cnt)}")

    add("")
    add(f"  === 销售渠道 分佈 ===")
    for ch, n in channels.most_common(20):
        add(f"    {ch}: {n:,} 列")
    if len(channels) > 20:
        add(f"    ...另外 {len(channels) - 20} 個渠道")

    if sink is not None:
        add("")
        add(f"  === parquet ({sink.codec}) ===")
        add(f"  {sink.path.name}  {sink.n:,} 列  "
            f"{sink.path.stat().st_size / 1e6:,.1f} MB  "
            f"(xlsx {path.stat().st_size / 1e6:,.1f} MB)")
        add("  -> 之後要重排格式直接讀這個檔, 不必再解析一次 xlsx")


# ------------------------------------------------------------------ 模板
def dump_template(path: Path) -> None:
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        add("")
        add(f"  --- 工作表 [{ws.title}]  {ws.max_row} 列 × {ws.max_column} 欄 "
            f"{'(隱藏)' if ws.sheet_state != 'visible' else ''} ---")
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            add(f"  合併儲存格 ({len(merged)}): {', '.join(sorted(merged)[:30])}"
                + (" ..." if len(merged) > 30 else ""))

        n_r = min(ws.max_row or 0, MAX_TPL_ROWS)
        n_c = min(ws.max_column or 0, MAX_TPL_COLS)
        notes = []
        for row in ws.iter_rows(min_row=1, max_row=n_r, max_col=n_c):
            cells = []
            for c in row:
                v = s(c.value)
                if not v:
                    continue
                mark = ""
                f = c.fill
                if f is not None and f.patternType and \
                        getattr(f.fgColor, "rgb", None) not in (None, "00000000"):
                    mark = "*"           # 有填色 -> 通常是人工標記
                cells.append(f"{c.column_letter}{mark}={v}")
                if c.comment is not None:
                    notes.append(f"    註解 {c.coordinate}: "
                                 f"{c.comment.text.strip()[:300]}")
            if cells:
                add(f"  R{row[0].row:<3} " + " | ".join(cells))
        if notes:
            add("  儲存格註解:")
            for n in notes:
                add(n)
        if (ws.max_row or 0) > MAX_TPL_ROWS or (ws.max_column or 0) > MAX_TPL_COLS:
            add(f"  (只印了前 {MAX_TPL_ROWS} 列 × {MAX_TPL_COLS} 欄)")
    wb.close()
    add("")
    add("  註: 欄位代號後面的 * 表示該儲存格有填色 (多半是人工標記)")


# ------------------------------------------------------------------ main
def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=".", help="要掃的資料夾 (預設: 目前目錄)")
    ap.add_argument("--no-scan", action="store_true", help="跳過寬表全掃 (最慢的一步)")
    ap.add_argument("--no-parquet", action="store_true", help="全掃但不寫 parquet")
    ap.add_argument("--rows", type=int, default=0, help="全掃只讀前 N 列")
    ap.add_argument("--big-mb", type=float, default=BIG_MB,
                    help=f"超過幾 MB 當寬表看待 (預設 {BIG_MB})")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    files = sorted([p for p in root.glob("*.xlsx") if not p.name.startswith("~$")],
                   key=lambda p: p.stat().st_size, reverse=True)
    if not files:
        sys.exit(f"{root} 底下沒有 xlsx")

    add(f"資料夾: {root}")
    add(f"檔案 {len(files)} 個")

    for path in files:
        big = path.stat().st_size / 1e6 >= args.big_mb
        add("")
        add("=" * 78)
        add(f"{path.name}   [{'寬表' if big else '模板/小檔'}]")
        add("=" * 78)
        try:
            one(path, big, args, root)
        except Exception as e:      # 一個檔壞掉不要拖垮其他檔
            add(f"  !! 讀這個檔出錯: {type(e).__name__}: {e}")

    out = root / REPORT
    out.write_text("\n".join(L), encoding="utf-8")
    print(f"\n報告: {out}", file=sys.stderr)


def one(path: Path, big: bool, args, root: Path) -> None:
    show_props(path)
    add("")

    if not big:
        dump_template(path)
        return

    print("  讀表頭中 (大檔要先載入字串表, 稍等)...", file=sys.stderr, flush=True)
    rows = head_rows(path)
    r = analyse_wide(rows)
    report_wide(path, rows, r)

    if args.no_scan:
        add("  (--no-scan: 跳過全掃, 沒有分類分佈)")
        return
    if not r["ok"]:
        add("  結構認不出來, 不做全掃也不寫 parquet —— 先看上面的表頭再決定")
        return

    sink = None
    if not args.no_parquet:
        try:
            import pyarrow  # noqa: F401
        except ImportError:
            add("  (沒有 pyarrow, 跳過 parquet: pip install pyarrow)")
        else:
            sink = Sink(root / PQ_NAME, pq_names(r), len(KEYS))
    add("")
    print("  全掃中...", file=sys.stderr, flush=True)
    scan_wide(path, r, args.rows or None, sink)


if __name__ == "__main__":
    main()
