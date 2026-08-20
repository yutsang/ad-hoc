#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
比較 商品名称 與 系统商品名称, 決定模板上該用哪一個。

順便找出「看起來有名稱、其實是垃圾值」的情況 (例如名稱是 "0"),
因為「取最後出現的名稱」遇到垃圾值會選到垃圾。

用法:
    python check_names.py
    python check_names.py 0229622868 011002      # 指定編碼鑽取
"""
import sys
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
PQ = ROOT / "parquet"
MASTER = ROOT / "data" / "2023-10-12.xlsx"      # 含商品對照表的檔

NAME, SYS = "商品名称", "系统商品名称"
CODE, AMT, PERIOD = "商品编码", "实收金额", "period"

# 一看就不是真名稱的值
_J = ("({c} ~ '^[0-9]+(\\.[0-9]+)?$' OR length(trim({c})) <= 1"
      " OR {c} IN ('-','--','null','NULL','N/A','#N/A')"
      " OR {c} LIKE '%测试%' OR lower({c}) LIKE '%test%')")
JUNK = _J.format(c=NAME)
JUNK_SYS = _J.format(c=SYS)


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    con = duckdb.connect()
    src = f"read_parquet('{PQ.as_posix()}/*.parquet')"
    out = []
    add = out.append

    # ---------------------------------------------------------- 指定編碼鑽取
    codes = sys.argv[1:] or ["0229622868"]
    add("=== 指定編碼鑽取 ===")
    for c in codes:
        rows = con.execute(f"""
            SELECT {NAME}, {SYS}, min({PERIOD}), max({PERIOD}),
                   count(DISTINCT {PERIOD}), count(*), sum(COALESCE({AMT},0))
            FROM {src} WHERE {CODE} = ? GROUP BY 1,2 ORDER BY 7 DESC
        """, [c]).fetchall()
        add(f"\n  编码 {c}:  {len(rows)} 種 (商品名称, 系统商品名称) 組合")
        if not rows:
            add("    (資料裡沒有這個編碼)")
            continue
        add(f"    {'商品名称':<28}{'系统商品名称':<28}{'首月':>9}{'末月':>9}"
            f"{'月数':>6}{'列数':>10}{'实收金额':>16}")
        for nm, sy, f, l, nmo, nr, amt in rows:
            add(f"    {str(nm):<28}{str(sy):<28}{f:>9}{l:>9}{nmo:>6}{nr:>10,}{amt:>16,.2f}")

    # ---------------------------------------------------------- 覆蓋率
    add("\n\n=== 覆蓋率: 每個編碼有沒有名稱 ===")
    r = con.execute(f"""
        WITH k AS (
            SELECT {CODE},
                   count({NAME}) AS n_name,
                   count({SYS})  AS n_sys,
                   count(*) FILTER (WHERE {NAME} IS NOT NULL AND NOT {JUNK}) AS n_good,
                   sum(COALESCE({AMT},0)) AS amt
            FROM {src} WHERE {CODE} IS NOT NULL GROUP BY 1)
        SELECT count(*),
               count(*) FILTER (WHERE n_name > 0),
               count(*) FILTER (WHERE n_sys  > 0),
               count(*) FILTER (WHERE n_name = 0 AND n_sys > 0),
               count(*) FILTER (WHERE n_name = 0 AND n_sys = 0),
               sum(amt) FILTER (WHERE n_name = 0 AND n_sys > 0),
               sum(amt) FILTER (WHERE n_name = 0 AND n_sys = 0),
               count(*) FILTER (WHERE n_good = 0 AND n_sys > 0)
        FROM k
    """).fetchone()
    tot, hn, hs, only_sys, neither, amt_only_sys, amt_neither, junk_rescued = r
    add(f"  編碼總數                                {tot:,}")
    add(f"  有 {NAME}                              {hn:,}  ({hn/tot*100:.1f}%)")
    add(f"  有 {SYS}                            {hs:,}  ({hs/tot*100:.1f}%)")
    add(f"  只有 {SYS} (用它就能救回)            {only_sys:,}   实收 {amt_only_sys or 0:,.2f}")
    add(f"  兩個都沒有 (真的無解)                    {neither:,}   实收 {amt_neither or 0:,.2f}")
    add(f"  {NAME} 全是垃圾值但有 {SYS}      {junk_rescued:,}")

    # ---------------------------------------------------------- 穩定度
    add("\n=== 穩定度: 哪個名稱欄比較不會漂移 ===")
    r = con.execute(f"""
        WITH k AS (
            SELECT {CODE}, count(DISTINCT {NAME}) AS dn, count(DISTINCT {SYS}) AS ds
            FROM {src} WHERE {CODE} IS NOT NULL GROUP BY 1)
        SELECT count(*) FILTER (WHERE dn > 1), count(*) FILTER (WHERE ds > 1),
               max(dn), max(ds), avg(dn), avg(ds)
        FROM k
    """).fetchone()
    add(f"  一個編碼對到多個 {NAME}:    {r[0]:,} 個  (最多 {r[2]} 種, 平均 {r[4]:.2f})")
    add(f"  一個編碼對到多個 {SYS}:  {r[1]:,} 個  (最多 {r[3]} 種, 平均 {r[5]:.2f})")
    add("  (只看『有幾個編碼漂移』會誤導 —— 還要看最壞情況有多壞)")

    # 一個編碼對到幾百種名稱, 代表那個編碼是萬用/雜項碼, 拿它的名稱當代表沒有意義
    for col in (NAME, SYS):
        rows = con.execute(f"""
            SELECT {CODE}, count(DISTINCT {col}) AS n, count(*) AS rows,
                   sum(COALESCE({AMT},0)) AS amt
            FROM {src} WHERE {CODE} IS NOT NULL AND {col} IS NOT NULL
            GROUP BY 1 HAVING count(DISTINCT {col}) > 5 ORDER BY n DESC LIMIT 5
        """).fetchall()
        add(f"\n  {col} 變體最多的編碼 (>5 種):")
        if not rows:
            add("    無")
        for c, n, nr, amt in rows:
            sample = con.execute(f"""
                SELECT {col} FROM {src} WHERE {CODE} = ? AND {col} IS NOT NULL
                GROUP BY 1 ORDER BY count(*) DESC LIMIT 3
            """, [c]).fetchall()
            add(f"    {c}: {n} 種, {nr:,} 列, 实收 {amt:,.2f}")
            add(f"      前 3 名: {[s[0] for s in sample]}")

    # ---------------------------------------------------------- 垃圾值
    add("\n=== 疑似垃圾的 商品名称 ===")
    rows = con.execute(f"""
        SELECT {NAME}, count(DISTINCT {CODE}), count(*), sum(COALESCE({AMT},0))
        FROM {src} WHERE {NAME} IS NOT NULL AND {JUNK}
        GROUP BY 1 ORDER BY 3 DESC LIMIT 20
    """).fetchall()
    if rows:
        add(f"  {'值':<24}{'编码数':>9}{'列数':>12}{'实收金额':>16}")
        for v, nc, nr, amt in rows:
            add(f"  {repr(v):<24}{nc:>9,}{nr:>12,}{amt:>16,.2f}")
    else:
        add("  無")

    # 「取最後出現的名稱」會不會選到垃圾
    r = con.execute(f"""
        WITH k AS (
            SELECT {CODE}, arg_max({NAME}, {PERIOD}) FILTER (WHERE {NAME} IS NOT NULL) AS latest
            FROM {src} WHERE {CODE} IS NOT NULL GROUP BY 1)
        SELECT count(*) FROM k WHERE latest IS NOT NULL
          AND (latest ~ '^[0-9]+(\\.[0-9]+)?$' OR length(trim(latest)) <= 1
               OR latest LIKE '%测试%' OR lower(latest) LIKE '%test%')
    """).fetchone()[0]
    add(f"\n  !! 「取最後出現的名稱」會選到垃圾值的編碼數: {r:,}")

    # ---------------------------------------------------------- 兩欄一致性
    add("\n=== 兩欄何時不一致 ===")
    r = con.execute(f"""
        SELECT count(*) FILTER (WHERE {NAME} IS NOT NULL AND {SYS} IS NOT NULL
                                  AND {NAME} = {SYS}),
               count(*) FILTER (WHERE {NAME} IS NOT NULL AND {SYS} IS NOT NULL
                                  AND {NAME} <> {SYS}),
               count(*) FILTER (WHERE {NAME} IS NULL AND {SYS} IS NOT NULL),
               count(*) FILTER (WHERE {NAME} IS NOT NULL AND {SYS} IS NULL),
               count(*) FILTER (WHERE {NAME} IS NULL AND {SYS} IS NULL)
        FROM {src}
    """).fetchone()
    labels = ["兩欄相同", "兩欄不同", f"只有 {SYS}", f"只有 {NAME}", "兩欄都空"]
    total = sum(r)
    for lab, v in zip(labels, r):
        add(f"  {lab:<22}{v:>14,}  ({v/total*100:5.2f}%)")

    # ---------------------------------------------------------- 對照表能補多少
    if MASTER.exists():
        m = {}
        wb = load_workbook(MASTER, read_only=True, data_only=True)
        for ws in wb.worksheets:
            first = next(ws.iter_rows(max_row=1, values_only=True), None)
            hdr = [("" if c is None else str(c).strip()) for c in (first or ())]
            named = [h for h in hdr if h]
            if CODE in hdr and SYS in hdr and len(named) <= 6:
                ci, ni = hdr.index(CODE), hdr.index(SYS)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if ci < len(row) and row[ci] is not None:
                        c = str(row[ci]).strip()
                        n = "" if ni >= len(row) or row[ni] is None else str(row[ni]).strip()
                        if c and n:
                            m.setdefault(c, n)
        wb.close()
        con.execute("CREATE TEMP TABLE master(code VARCHAR, name VARCHAR)")
        con.executemany("INSERT INTO master VALUES (?, ?)", list(m.items()))
        r = con.execute(f"""
            WITH k AS (
                SELECT {CODE} AS code,
                       count(*) FILTER (WHERE {NAME} IS NOT NULL AND NOT {JUNK}) AS n_good,
                       -- 系统商品名称 也要濾垃圾: 只數「有值」會高估能救回的數量
                       count(*) FILTER (WHERE {SYS} IS NOT NULL AND NOT {JUNK_SYS}) AS n_sys,
                       sum(COALESCE({AMT},0)) AS amt
                FROM {src} WHERE {CODE} IS NOT NULL GROUP BY 1)
            SELECT count(*) FILTER (WHERE k.n_good = 0),
                   count(*) FILTER (WHERE k.n_good = 0 AND k.n_sys > 0),
                   count(*) FILTER (WHERE k.n_good = 0 AND k.n_sys = 0 AND ms.name IS NOT NULL),
                   count(*) FILTER (WHERE k.n_good = 0 AND k.n_sys = 0 AND ms.name IS NULL),
                   sum(k.amt) FILTER (WHERE k.n_good = 0 AND k.n_sys = 0 AND ms.name IS NULL)
            FROM k LEFT JOIN master ms ON k.code = ms.code
        """).fetchone()
        add(f"\n=== 沒有可用 {NAME} 的編碼, 能被補回多少 ===")
        add(f"  對照表載入 {len(m):,} 筆")
        add(f"  沒有可用 {NAME} 的編碼          {r[0]:,}")
        add(f"    其中可用明細的 {SYS} 補     {r[1]:,}")
        add(f"    其中可用對照表補                {r[2]:,}")
        add(f"    完全補不回來                    {r[3]:,}   实收 {r[4] or 0:,.2f}")

    txt = "\n".join(out)
    (ROOT / "name_check.txt").write_text(txt, encoding="utf-8")
    print(txt)


if __name__ == "__main__":
    main()
