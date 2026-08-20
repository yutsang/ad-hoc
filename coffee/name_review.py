#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
列出名稱需要複核的商品編碼: 排除垃圾值 (空白 / '0' / 單字元 / 測試品) 之後,
每個編碼實際會拿到什麼名稱、來自哪一欄。

全自動規則, 沒有人工填寫。這支只是把結果攤開來給人看, 確認自動解析合理。

用法:
    python name_review.py

產出 名称复核.xlsx:
    工作表「需复核」      有垃圾值 / 空白 / 多個有效名稱的編碼
    工作表「变体最多」    一個編碼對到很多名稱的情況 (通常是萬用碼或髒資料)
"""
import sys
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PQ = ROOT / "parquet"
OUT = ROOT / "名称复核.xlsx"

NAME, SYS = "商品名称", "系统商品名称"
CODE, AMT, PERIOD = "商品编码", "实收金额", "period"
CAT = "产品分类"

JUNK = ("({c} ~ '^[0-9]+(\\.[0-9]+)?$' OR length(trim({c})) <= 1 "
        "OR {c} IN ('-','--','null','NULL','N/A','#N/A') "
        "OR {c} LIKE '%测试%' OR lower({c}) LIKE '%test%')")
JN, JS = JUNK.format(c=f'"{NAME}"'), JUNK.format(c=f'"{SYS}"')


def join(vals, n=6):
    """把名稱清單接成一格。太多就截斷, 免得 711 種塞爆儲存格。"""
    vals = [v for v in (vals or []) if v is not None]
    if not vals:
        return ""
    s = " | ".join(map(str, vals[:n]))
    return s + (f"  …共 {len(vals)} 種" if len(vals) > n else "")


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    if not sorted(PQ.glob("*.parquet")):
        sys.exit(f"找不到 parquet: {PQ}")

    con = duckdb.connect()
    src = f"read_parquet('{PQ.as_posix()}/*.parquet')"

    rows = con.execute(f"""
        SELECT "{CODE}"                                                        AS code,
               list(DISTINCT "{NAME}")  FILTER (WHERE "{NAME}" IS NOT NULL)    AS all_names,
               list(DISTINCT "{NAME}")  FILTER (WHERE "{NAME}" IS NOT NULL AND {JN})     AS junk_names,
               list(DISTINCT "{NAME}")  FILTER (WHERE "{NAME}" IS NOT NULL AND NOT {JN}) AS good_names,
               list(DISTINCT "{SYS}")   FILTER (WHERE "{SYS}"  IS NOT NULL AND NOT {JS}) AS good_sys,
               arg_max("{NAME}", "{PERIOD}") FILTER (WHERE "{NAME}" IS NOT NULL AND NOT {JN}) AS pick_name,
               arg_max("{SYS}",  "{PERIOD}") FILTER (WHERE "{SYS}"  IS NOT NULL AND NOT {JS}) AS pick_sys,
               count(*) FILTER (WHERE "{NAME}" IS NULL)                        AS n_blank,
               count(*)                                                        AS n_rows,
               count(DISTINCT "{PERIOD}")                                      AS n_months,
               min("{PERIOD}") AS first_p, max("{PERIOD}") AS last_p,
               any_value("{CAT}") FILTER (WHERE "{CAT}" IS NOT NULL)          AS category,
               sum(COALESCE("{AMT}", 0))                                       AS amt
        FROM {src} WHERE "{CODE}" IS NOT NULL GROUP BY 1
    """).fetchall()

    review, stats = [], {"无名称": 0, "取系统名": 0, "多名称": 0, "有垃圾值": 0, "正常": 0}
    amt_by = {k: 0.0 for k in stats}
    for (code, all_n, junk_n, good_n, good_s, pick_n, pick_s,
         n_blank, n_rows, n_months, fp, lp, cat, amt) in rows:
        good_n, junk_n, good_s = good_n or [], junk_n or [], good_s or []
        final = pick_n or pick_s
        source = "商品名称" if pick_n else ("系统商品名称" if pick_s else "(空白)")

        if not final:
            st = "无名称"
        elif not good_n:
            st = "取系统名"
        elif len(good_n) > 1:
            st = "多名称"
        elif junk_n:
            st = "有垃圾值"
        else:
            st = "正常"
        stats[st] += 1
        amt_by[st] += amt or 0
        if st != "正常":
            review.append((code, st, cat or "", join(all_n), join(junk_n), join(good_n),
                           join(good_s), final or "", source, n_blank, n_rows,
                           n_months, fp, lp, round(amt or 0, 2)))

    review.sort(key=lambda r: (-r[14],))

    # ---- 主控台摘要 --------------------------------------------------
    print(f"編碼總數 {len(rows):,}")
    print(f"  {'狀態':<10}{'編碼數':>9}{'实收金额':>20}")
    for k in ("无名称", "取系统名", "多名称", "有垃圾值", "正常"):
        print(f"  {k:<10}{stats[k]:>9,}{amt_by[k]:>20,.2f}")
    print(f"\n需複核 {len(review):,} 個編碼, 前 15 名 (依实收金额):")
    print(f"  {'编码':<14}{'状态':<9}{'最终采用':<26}{'来源':<14}{'实收金额':>16}")
    for r in review[:15]:
        print(f"  {str(r[0]):<14}{r[1]:<9}{str(r[7])[:24]:<26}{r[8]:<14}{r[14]:>16,.2f}")

    # ---- 變體最多的編碼 ----------------------------------------------
    top = con.execute(f"""
        SELECT "{CODE}", count(DISTINCT "{NAME}"), count(DISTINCT "{SYS}"),
               count(*), sum(COALESCE("{AMT}",0))
        FROM {src} WHERE "{CODE}" IS NOT NULL GROUP BY 1
        HAVING count(DISTINCT "{NAME}") > 4 OR count(DISTINCT "{SYS}") > 4
        ORDER BY greatest(count(DISTINCT "{NAME}"), count(DISTINCT "{SYS}")) DESC LIMIT 30
    """).fetchall()
    print(f"\n變體最多的編碼 (任一欄 >4 種), 共 {len(top)} 個, 前 5:")
    for c, dn, ds, nr, amt in top[:5]:
        sample = con.execute(f"""
            SELECT "{SYS}" FROM {src} WHERE "{CODE}" = ? AND "{SYS}" IS NOT NULL
            GROUP BY 1 ORDER BY count(*) DESC LIMIT 3
        """, [c]).fetchall()
        print(f"  {c}: {NAME} {dn} 種 / {SYS} {ds} 種, {nr:,} 列, 实收 {amt:,.2f}")
        print(f"     {SYS} 前 3: {[s[0] for s in sample]}")

    # ---- 寫 xlsx -----------------------------------------------------
    wb = Workbook()
    hf, hfill = Font(bold=True), PatternFill("solid", fgColor="DDDDDD")
    ws = wb.active
    ws.title = "需复核"
    head = ["商品编码", "状态", "产品分类", "全部商品名称(含垃圾)", "其中垃圾值", "有效商品名称",
            "有效系统商品名称", "最终采用", "来源", "名称空白列数", "明细列数",
            "出现月数", "首月", "末月", "实收金额"]
    ws.append(head)
    for i in range(1, len(head) + 1):
        ws.cell(1, i).font, ws.cell(1, i).fill = hf, hfill
        ws.cell(1, i).alignment = Alignment(horizontal="center")
    for r in review:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in {1: 16, 2: 11, 3: 22, 4: 46, 5: 22, 6: 46, 7: 46, 8: 30,
                 9: 15, 10: 12, 11: 11, 12: 10, 13: 10, 14: 10, 15: 16}.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("变体最多")
    head2 = ["商品编码", f"{NAME} 种数", f"{SYS} 种数", "明细列数", "实收金额",
             f"{SYS} 前 5"]
    ws2.append(head2)
    for i in range(1, len(head2) + 1):
        ws2.cell(1, i).font, ws2.cell(1, i).fill = hf, hfill
    for c, dn, ds, nr, amt in top:
        s = con.execute(f"""
            SELECT "{SYS}" FROM {src} WHERE "{CODE}" = ? AND "{SYS}" IS NOT NULL
            GROUP BY 1 ORDER BY count(*) DESC LIMIT 5
        """, [c]).fetchall()
        ws2.append([c, dn, ds, nr, round(amt or 0, 2), join([x[0] for x in s], 5)])
    ws2.freeze_panes = "A2"
    for i, w in {1: 16, 2: 14, 3: 16, 4: 12, 5: 16, 6: 60}.items():
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"\n已寫出 {OUT.name} ({len(review):,} 列需複核)")


if __name__ == "__main__":
    main()
